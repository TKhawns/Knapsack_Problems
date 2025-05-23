import os
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from ortools.linear_solver import pywraplp
import matplotlib.pyplot as plt

id_counter = 0


def write_to_xlsx(result_dict):
    excel_results = []
    excel_results.append(result_dict)
    output_path = 'output/'

    if not os.path.exists(output_path):
        os.makedirs(output_path)

    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()

        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False):
            sheet.append(row)
        book.save(excel_file_path)

    else:
        df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

def solve_by_pysat(input, time, result, num_var, num_clause, name_lib):
    global id_counter
    id_counter += 1
    result_dict = {
        "ID": id_counter,
        "Problem": input,
        "Type": name_lib,
        "Time": time,
        "Result": result,
        "Variables": num_var,
        "Clauses": num_clause
    }
    write_to_xlsx(result_dict)

def max_profit_solution(rectangles, weights, profits, capacity, W, H):
    solver = pywraplp.Solver.CreateSolver('SAT')
    if not solver:
        print("MIP solver not available.")
        return None

    xs = [solver.BoolVar(f"x_{i}") for i in range(len(weights))]

    # Weight constraint
    solver.Add(solver.Sum(xs[i] * weights[i] for i in range(len(weights))) <= capacity)

    # Ham muc tieu
    solver.Maximize(solver.Sum(xs[i] * profits[i] for i in range(len(profits))))

    status = solver.Solve()
    elapsed_time = solver.wall_time() / 1000.0
    all_solutions = []
    result = []

    if status == pywraplp.Solver.OPTIMAL or status == pywraplp.Solver.FEASIBLE:
        solution = [int(xs[i].solution_value()) for i in range(len(weights))]
        all_solutions.append(solution)

        selected_items = [i for i, x in enumerate(solution) if x == 1]
        total_profit = sum(profits[i] for i in selected_items)
        result.append(selected_items + [total_profit])

    # Sort the results by profit
    sorted_result = sorted(result, key=lambda x: x[-1], reverse=True)
    num_solution = len(sorted_result)
    isSat = "unsat" if num_solution == 0 else "sat"

    for i in range(0, num_solution):
        input = sorted_result[i]
        size = len(input)
        list_rect = []

        for j in range(0, size-1):
            list_rect.append(rectangles[input[j]])

        max_profit = sorted_result[i][-1]

        status = mip_OPP(list_rect, W, H)
        elapsed_time += status[1]
        if status[0] == "sat":
            print("Max profit:", max_profit)
            isSat = "sat"
            break

    solve_by_pysat(len(weights), elapsed_time, isSat, 0, 0, "MIP")
    return elapsed_time

def mip_OPP(rectangles, W, H):
    solver = pywraplp.Solver.CreateSolver('SAT')
    if not solver:
        print("MIP solver not available.")
        return None

    n = len(rectangles)
    x = [solver.IntVar(0, W, f'x_{i}') for i in range(n)]
    y = [solver.IntVar(0, H, f'y_{i}') for i in range(n)]

    for i in range(n):
        wi, hi = rectangles[i]
        solver.Add(x[i] + wi <= W)
        solver.Add(y[i] + hi <= H)

    for i in range(n):
        for j in range(i + 1, n):
            wi, hi = rectangles[i]
            wj, hj = rectangles[j]
            no_overlap_1 = solver.BoolVar(f'nonoverlap_1_{i}_{j}')
            no_overlap_2 = solver.BoolVar(f'nonoverlap_2_{i}_{j}')
            no_overlap_3 = solver.BoolVar(f'nonoverlap_3_{i}_{j}')
            no_overlap_4 = solver.BoolVar(f'nonoverlap_4_{i}_{j}')

            M = max(W, H)
            solver.Add(x[i] + wi <= x[j] + M * (1 - no_overlap_1))
            solver.Add(x[j] + wj <= x[i] + M * (1 - no_overlap_2))
            solver.Add(y[i] + hi <= y[j] + M * (1 - no_overlap_3))
            solver.Add(y[j] + hj <= y[i] + M * (1 - no_overlap_4))
            solver.Add(no_overlap_1 + no_overlap_2 + no_overlap_3 + no_overlap_4 >= 1)

    solver.Maximize(solver.Sum([0]))

    status = solver.Solve()
    elapsed_time = solver.wall_time() / 1000.0

    if status == pywraplp.Solver.OPTIMAL:
        print("Solution found:")
        for i in range(n):
            wi, hi = rectangles[i]
            print(wi, hi)
            print(f"Rectangle {i+1}: Position (x, y) = ({x[i].solution_value()}, {y[i].solution_value()})"),    
        return ["sat", elapsed_time]
    else:
        return ["unsat", elapsed_time]

def solve_KPWL_CPSAT():
    with open('under500_input.txt', 'r') as file:
        lines = file.readlines()
        len_file = len(lines)
        for i in range(0, len_file, 6):
            bounds = list(map(int, lines[i].strip().split()))
            arr_strip = list(map(int, lines[i + 1].strip().split()))
            weights = list(map(int, lines[i + 2].strip().split()))
            profits = list(map(int, lines[i + 3].strip().split()))
            arr_width = list(map(int, lines[i + 4].strip().split()))
            arr_height = list(map(int, lines[i + 5].strip().split()))
            rectangles = [(arr_width[j], arr_height[j]) for j in range(len(arr_width))]
            weight_bound = bounds[0]


            max_profit_solution(rectangles, weights, profits, weight_bound, arr_strip[0], arr_strip[1])
            pos = [[0 for i in range(2)] for j in range(len(rectangles))]
            # display_solution(arr_strip, rectangles, pos)



def display_solution(strip, rectangles, pos_circuits):
    # define Matplotlib figure and axis
    fig, ax = plt.subplots()
    ax = plt.gca()
    plt.title(strip)

    if len(pos_circuits) > 0:
        for i in range(len(rectangles)):
            rect = plt.Rectangle(pos_circuits[i], *rectangles[i], edgecolor="#333")
            ax.add_patch(rect)

    ax.set_xlim(0, strip[0])
    ax.set_ylim(0, strip[1] + 1)
    ax.set_xticks(range(strip[0] + 1))
    ax.set_yticks(range(strip[1] + 1))
    ax.set_xlabel('width')
    ax.set_ylabel('height')
    # display plot
    plt.show()

solve_KPWL_CPSAT()
