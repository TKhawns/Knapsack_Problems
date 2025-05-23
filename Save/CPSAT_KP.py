from ortools.sat.python import cp_model
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

id_counter = 0

def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path =  'output/'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)

    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

def solve_by_pysat(input, time, result, num_var, num_clause, name_lib):
    global id_counter
    id_counter += 1
    result_dict = {
        "ID": id_counter,
        "Problem": input,
        "Type": name_lib,
        "Time": time,
        "Result": result,
        "Variables": num_var,
        "Clauses": num_clause
    }
    write_to_xlsx(result_dict)

def cpSAT_OPP(rectangles, W, H):
    model = cp_model.CpModel()

    n = len(rectangles)
    
    # Variable encoding (x[i], y[i]) = (w[i], h[i])
    x = [model.NewIntVar(0, W, f'x_{i}') for i in range(n)]
    y = [model.NewIntVar(0, H, f'y_{i}') for i in range(n)]

    # (1) Domain of rectangle (x[i], y[i])
    for i in range(n):
        wi, hi = rectangles[i]
        print(wi, hi)
        model.Add(x[i] + wi <= W) 
        model.Add(y[i] + hi <= H)

    #(2) Non-overlapping constraints
    for i in range(n):
        for j in range(i + 1, n):
            wi, hi = rectangles[i]
            wj, hj = rectangles[j]

            # Boolean variables
            no_overlap_1 = model.NewBoolVar(f'nonoverlap_1_{i}_{j}')  # x[i] + wi <= x[j]
            no_overlap_2 = model.NewBoolVar(f'nonoverlap_2_{i}_{j}')  # x[j] + wj <= x[i]
            no_overlap_3 = model.NewBoolVar(f'nonoverlap_3_{i}_{j}')  # y[i] + hi <= y[j]
            no_overlap_4 = model.NewBoolVar(f'nonoverlap_4_{i}_{j}')  # y[j] + hj <= y[i]

            model.Add(x[i] + wi <= x[j]).OnlyEnforceIf(no_overlap_1)
            model.Add(x[j] + wj <= x[i]).OnlyEnforceIf(no_overlap_2)
            model.Add(y[i] + hi <= y[j]).OnlyEnforceIf(no_overlap_3)
            model.Add(y[j] + hj <= y[i]).OnlyEnforceIf(no_overlap_4)

            # At least one constraints
            model.AddBoolOr([no_overlap_1, no_overlap_2, no_overlap_3, no_overlap_4])

    # CP Solver
    solver = cp_model.CpSolver()
    status = solver.Solve(model)
    elapse_time = solver.user_time

    if status == cp_model.FEASIBLE or status == cp_model.OPTIMAL:
        print('Solution found:')
        for i in range(n):
            print(f'Rectangle {i}: ({solver.Value(x[i])}, {solver.Value(y[i])})')
        return ["sat", elapse_time]
    else:
        print('No solution found.')
        return ["unsat", elapse_time]

def max_profit_solution(rectangles, weights, profits, capacity, W, H):

    model = cp_model.CpModel()
    xs = [model.NewBoolVar(f"x_{i}") for i in range(len(weights))]

    # Constrainst max Weight
    model.Add(sum(x * w for x, w in zip(xs, weights)) <= capacity)
    
    solver = cp_model.CpSolver()
    status = solver.Solve(model)
    solution = [solver.Value(x) for x in xs]

    all_solutions = []

    #  Array that store [solution, max profit of solution].
    result = []
    sum_profit = set()
    elapsed_time = 0
    n_items = len(profits)
    isSat = ""

    while True:
        status = solver.Solve(model)
        if (elapsed_time > 100):
            isSat = "timeout"
            solve_by_pysat(len(weights), elapsed_time, isSat, 0, 0, "CPSAT")
            return [all_solutions, "timeout"]
        
        elapsed_time = elapsed_time +  solver.user_time

        if status == cp_model.FEASIBLE or status == cp_model.OPTIMAL:
            solution = [solver.Value(x) for x in xs]
            all_solutions.append(solution)

            temp = []
            for i in range (0, n_items):
                if (solution[i] > 0):
                    temp.append(solution[i])

            selected_items = [i for i, x in enumerate(solution) if x == 1]

            total_profit = sum(profits[i] for i in selected_items)

            sum_profit.add(total_profit)
            if (selected_items != []):
                selected_items.append(total_profit)
                result.append(selected_items)

            # Remove current solution
            model.AddBoolOr([xs[i].Not() if solver.Value(xs[i]) == 1 else xs[i] for i in range(len(xs))])
        else: 
            break

    sorted_result = sorted(result, key=lambda x: x[-1], reverse=True)
    num_solution = len(sorted_result)

    for i in range(0, num_solution):
            input = sorted_result[i]
            size = len(input)
            list_rect = []

            for j in range(0, size-1):
                list_rect.append(rectangles[input[j]])

            max_profit = sorted_result[i][-1]

            status = cpSAT_OPP(list_rect, W , H)
            elapsed_time += status[1]
            if (status[0] == "sat"):
                print("Max profit", max_profit)
                isSat = "sat"
                break
            elif (status[0] == "unsat"):
                isSat = "unsat"
                continue

    solve_by_pysat(len(weights), elapsed_time, isSat, 0, 0, "CPSAT")
    return elapsed_time


def solve_KPWL_CPSAT():
    with open('under500_input.txt', 'r') as file:
        # max_weight, min_profit
        # strip
        # list weights
        # list profits
        # width_i
        # height_i
        lines = file.readlines()
        len_file = len(lines)
        for i in range(0, len_file,6):
            bounds = list(map(int,lines[i].strip().split()))
            arr_strip = list(map(int,lines[i+1].strip().split()))
            weights = list(map(int, lines[i + 2].strip().split()))
            profits = list(map(int, lines[i + 3].strip().split()))
            arr_width =  list(map(int, lines[i + 4].strip().split()))
            arr_height =  list(map(int, lines[i + 5].strip().split()))
            rectangles = []
            
            l = len(arr_width)
            for j in range(0, l):
                rectangles.append((arr_width[j], arr_height[j]))
            weight_bound = bounds[0]
            
            max_profit_solution(rectangles, weights, profits, weight_bound, arr_strip[0], arr_strip[1])

solve_KPWL_CPSAT()

# print(f"Number of solutions: {len(all_solutions)}")
