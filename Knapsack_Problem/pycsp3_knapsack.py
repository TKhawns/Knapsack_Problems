from pycsp3 import *
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from threading import Timer
import time as timelib

id_counter = 0

def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path =  'output/'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)

    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

def solve_by_pysat(input, time, result):
    global id_counter
    id_counter += 1
    result_dict = {
        "ID": id_counter,
        "Problem": input,
        "Type": "PyCSP3",
        "Time": time,
        "Result": result,
        "Variables": "unknown",
        "Clauses": "unknown"
    }
    write_to_xlsx(result_dict)

def pycsp_solution():
    with open('base_input.txt', 'r') as file:
        lines = file.readlines()
        len_file = len(lines)

        for i in range(0, len_file,3):
            result = ""
            bounds = list(map(int,lines[i].strip().split()))
            weights = list(map(int, lines[i + 1].strip().split()))
            profits = list(map(int, lines[i + 2].strip().split()))
            w_limit = bounds[0]
            p_limit = bounds[1]
            nItems = len(weights)
            print(weights)

            variable = VarArray(size=nItems, dom={0, 1}, id = "x" + str(i))
            print("Array of variable x: ", variable)

            #start = timelib.time()
            satisfy(
                Knapsack(variable, weights=weights, wlimit=w_limit, profits=profits) >= p_limit
            )
            result_status = solve(sols=ALL, options="-t=60s")
            #elapse_time = timelib.time() - start
            print(result_status)
            if result_status is UNKNOWN:
                result = "unsolve"
            else: 
                if result_status is UNSAT:
                    result = "unsat"
                else:
                    if result_status is SAT:
                        result = "sat"
                        for j in range(n_solutions()):
                            print(f"Solution {j+1}: {values(variable, sol=j)} of profit {sum(value(variable[k], sol=j)*profits[k] for k in range(nItems))}")
            #solve_by_pysat(str(weights), str(elapse_time), result)
            clear()

pycsp_solution()