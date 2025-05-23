from pysat.pb import PBEnc
from pysat.solvers import Glucose3
from pypblib import pblib
from pypblib.pblib import PBConfig, Pb2cnf
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from threading import Timer

time_budget = 30
id_counter = 0

def interrupt(s): s.interrupt()

def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path =  'output/'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)

    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

def solve_by_pysat(input, time, result, num_var, num_clause, name_lib):
    global id_counter
    id_counter += 1
    result_dict = {
        "ID": id_counter,
        "Problem": input,
        "Type": name_lib,
        "Time": time,
        "Result": result,
        "Variables": num_var,
        "Clauses": num_clause
    }
    write_to_xlsx(result_dict)

def find_all_solutions(solver):
    all_solutions = []
    solutions_set = set() 
    
    while solver.solve():
        # Get the current solution
        solution = solver.get_model()
        temp = []
        for i in range (0, 10):
            if (solution[i] > 0):
                temp.append(solution[i])
        solutions_set.add(tuple(temp))
        
        # Block the current solution
        blocking_clause = [-lit for lit in solution]
        solver.add_clause(blocking_clause)
    
    return solutions_set

def pysat_solution():


    # <Loop> First line is bound weight and bound profit, two next lines is weight and height of each item
    with open('under500_input.txt', 'r') as file:
        lines = file.readlines()
        len_file = len(lines)

        for i in range(0, len_file,3):
            result = ""
            time = 0
            bounds = list(map(int,lines[i].strip().split()))
            weights = list(map(int, lines[i + 1].strip().split()))
            profits = list(map(int, lines[i + 2].strip().split()))
            weight_bound = bounds[0]
            profit_bound = bounds[1]
            num_items = len(weights)
            formula = []
            vars = list(range(1, num_items + 1))

            pbConfig = PBConfig()
            pbConfig.set_PB_Encoder(pblib.PB_BEST)
            pb2 = Pb2cnf(pbConfig)

            max_var = pb2.encode_leq(weights, vars, weight_bound, formula, num_items+1)
            # print(max_var)
            # max_var = pb2.encode_geq(profits, vars, profit_bound, formula, max_var+1)
            
            solver = Glucose3(use_timer=True)

            
            for clause in formula:
                solver.add_clause(clause)

            # print(find_all_solutions(solver))

            num_vars = solver.nof_vars()
            num_clauses = solver.nof_clauses()

            # print("Variables: " + str(num_vars))
            # print("Clauses:" + str(num_clauses))

            timer = Timer(time_budget, interrupt, [solver])
            timer.start()

            sat_status = solver.solve_limited(expect_interrupt = True)

            if sat_status is False:
                elapsed_time = format(solver.time())
                result = "unsat"
                time = elapsed_time
                # print("No solutions found")
            else:
                solution = solver.get_model()
                print(solution)
                if solution is None:
                    result = "timeout"
                    time = time_budget
                else:
                    elapsed_time = format(solver.time())
                    result = "sat"
                    time = elapsed_time
                    selected_items = [i for i in range(1, num_items + 1) if solution[i-1] > 0]
                    print("\nSelected Items:")
                    print(selected_items)

            timer.cancel()
            solver.delete()

            solve_by_pysat(str(num_items), str(time), result, num_vars, num_clauses, "PB_SORTINGNETWORKS")            
pysat_solution()