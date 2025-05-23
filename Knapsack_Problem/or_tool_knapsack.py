from ortools.algorithms.python import knapsack_solver
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import time


id_counter = 0
time_budget = 600

def interrupt(s): s.interrupt()

def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path =  'output/'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)

    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

def solve_by_or(input, time, result):
    global id_counter
    id_counter += 1
    result_dict = {
        "ID": id_counter,
        "Problem": "OR"  + input,
        "Type": "OR-Tool",
        "Time": time,
        "Result": result,
        "Variables": "unknown",
        "Clauses": "unknown"
    }
    write_to_xlsx(result_dict)

def or_solutions():
    result = ""
    # <Loop> First line is bound weight and bound profit, two next lines is weight and height of each item
    with open('base_input.txt', 'r') as file:
        lines = file.readlines()
        print(lines)
        len_file = len(lines)
        print(len_file)
        for i in range(0, len_file,3):
            bounds = list(map(int,lines[i].strip().split()))
            weights = [list(map(int, lines[i + 1].strip().split()))]
            values = list(map(int, lines[i + 2].strip().split()))
            capacities = [bounds[0]]
            print(capacities)   

            packed_items = []
            packed_weights = []
            total_weight = 0
            profit = 0

            # Start timer
            start = time.time()
            solver = knapsack_solver.KnapsackSolver(
                knapsack_solver.SolverType.KNAPSACK_MULTIDIMENSION_BRANCH_AND_BOUND_SOLVER,
                "KnapsackExample",)
            solver.init(values, weights, capacities)
            computed_value = solver.solve()
            total_time = (time.time() - start)

            # print("Total value =", computed_value)
            for i in range(len(values)):
                if solver.best_solution_contains(i):
                    packed_items.append(i)
                    packed_weights.append(weights[0][i])
                    total_weight += weights[0][i]
                    profit += values[i]
                    result = "SAT"
                if computed_value == 0:
                    result = "UNSAT"
                
            print("Total weight:", total_weight)
            print("Packed items:", packed_items)
            print("Packed_weights:", packed_weights)
            print("Total values: ", profit)
            solve_by_or(str(weights),str(total_time), result)

or_solutions()
