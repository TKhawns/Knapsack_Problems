from pysat.formula import CNF
from pysat.solvers import Solver
from pysat.solvers import Glucose3
from pypblib import pblib
from pypblib.pblib import PBConfig, Pb2cnf
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

id_counter = 0

def positive_range(end):
    if (end < 0):
        return []
    return range(end)

def opp_solution(rectangles, strip):
    cnf = CNF()
    variables = {}
    counter = 1
    for i in range(len(rectangles)):
        for j in range(len(rectangles)):
            variables[f"lr{i + 1},{j + 1}"] = counter  # lri,rj
            counter += 1
            variables[f"ud{i + 1},{j + 1}"] = counter  # uri,rj
            counter += 1
        for e in positive_range(strip[0] - rectangles[i][0] + 2):
            variables[f"px{i + 1},{e}"] = counter  # pxi,e
            counter += 1
        for f in positive_range(strip[1] - rectangles[i][1] + 2):
            variables[f"py{i + 1},{f}"] = counter  # pyi,f
            counter += 1

    # Add the 2-literal axiom clauses
    for i in range(len(rectangles)):
        for e in range(strip[0] - rectangles[i][0] + 1):  # -1 because we're using e+1 in the clause
            cnf.append([-variables[f"px{i + 1},{e}"], variables[f"px{i + 1},{e + 1}"]])
        for f in range(strip[1] - rectangles[i][1] + 1):  # -1 because we're using f+1 in the clause
            cnf.append([-variables[f"py{i + 1},{f}"], variables[f"py{i + 1},{f + 1}"]])

    # Add the 4-literal axiom clauses
    for i in range(len(rectangles)):
        for j in range(i + 1, len(rectangles)):
            cnf.append([variables[f"lr{i + 1},{j + 1}"], variables[f"lr{j + 1},{i + 1}"], variables[f"ud{i + 1},{j + 1}"],
                        variables[f"ud{j + 1},{i + 1}"]])

    # Add the 3-literal non-overlapping constraints
    for i in range(len(rectangles)):
        for j in range(i + 1, len(rectangles)):
            for e in positive_range(strip[0] - rectangles[i][0]) :
                if f"px{j + 1},{rectangles[i][0] - 1}" in variables:
                    cnf.append([-variables[f"lr{i + 1},{j + 1}"], -variables.get(f"px{j + 1},{rectangles[i][0] - 1}", 0)])
                if(strip[0] - rectangles[i][0] - rectangles[j][0] > 0) and f"px{i + 1},{strip[0] - rectangles[i][0] - rectangles[j][0]}" in variables:
                    cnf.append([-variables[f"lr{i + 1},{j + 1}"],
                                variables[f"px{i + 1},{strip[0] - rectangles[i][0] - rectangles[j][0]}"]])
                if f"px{j + 1},{e + rectangles[i][0]}" in variables:
                    cnf.append([-variables[f"lr{i + 1},{j + 1}"], variables[f"px{i + 1},{e}"],
                                -variables.get(f"px{j + 1},{e + rectangles[i][0]}", 0)])
                if f"px{i + 1},{e + rectangles[j][0]}" in variables:
                    cnf.append([-variables[f"lr{j + 1},{i + 1}"], variables.get(f"px{j + 1},{e}", 0),
                                -variables.get(f"px{i + 1},{e + rectangles[j][0]}", 0)])
                    
            for f in positive_range(strip[1] - rectangles[i][1]):
                if f"py{j + 1},{rectangles[i][1] - 1}" in variables:
                    cnf.append([-variables[f"ud{i + 1},{j + 1}"], -variables.get(f"py{j + 1},{rectangles[i][1] - 1}", 0)])
                if strip[1] - rectangles[i][1] - rectangles[j][1] > 0:
                    cnf.append([-variables[f"ud{i + 1},{j + 1}"],
                                variables[f"py{i + 1},{strip[1] - rectangles[i][1] - rectangles[j][1]}"]])
                if f"py{j + 1},{f + rectangles[i][1]}" in variables:
                    cnf.append([-variables[f"ud{i + 1},{j + 1}"], variables[f"py{i + 1},{f}"],
                            -variables.get(f"py{j + 1},{f + rectangles[i][1]}", 0)])

                if f"py{i + 1},{f + rectangles[j][1]}" in variables:
                    cnf.append([-variables[f"ud{j + 1},{i + 1}"], variables.get(f"py{j + 1},{f}", 0),
                                -variables.get(f"py{i + 1},{f + rectangles[j][1]}", 0)])

    for i in range(len(rectangles)):
        if f"px{i + 1},{strip[0] - rectangles[i][0]}" in variables:
            cnf.append([variables[f"px{i + 1},{strip[0] - rectangles[i][0]}"]])  # px(i, W-wi)
        if f"py{i + 1},{strip[1] - rectangles[i][1]}" in variables:
            cnf.append([variables[f"py{i + 1},{strip[1] - rectangles[i][1]}"]])  # py(i, H-hi)

    for i in range(len(rectangles)):
        for j in range(i + 1, len(rectangles)):
            # if indomain(len(px[j]) - 1, width_rects[i] - 1)
            if strip[0] - rectangles[i][0] + 1 >= rectangles[j][0] - 1:
                cnf.append([-variables[f"lr{i + 1},{j + 1}"], -variables[f"px{j + 1},{rectangles[i][0] - 1}"]])
            else:
                if f"px{j + 1},{strip[0] - rectangles[i][0] + 1}" in variables:
                    cnf.append([-variables[f"lr{i + 1},{j + 1}"], variables[f"px{j + 1},{strip[0] - rectangles[i][0] + 1}"]])
            # if indomain(len(px[i] - 1, width_rects[j] - 1)
            if strip[0] - rectangles[j][0] + 1 >= rectangles[i][0] - 1:
                cnf.append([-variables[f"lr{j + 1},{i + 1}"], -variables[f"px{i + 1},{rectangles[j][0] - 1}"]])
            else:
                if f"px{i + 1},{strip[0] - rectangles[j][0] + 1}" in variables:
                    cnf.append([-variables[f"lr{j + 1},{i + 1}"], variables[f"px{i + 1},{strip[0] - rectangles[j][0] + 1}"]])
            # if indomain(len(py[j]) - 1, height_rects[i] - 1)
            if strip[1] - rectangles[i][1] + 1 >= rectangles[j][1] - 1:
                cnf.append([-variables[f"ud{i + 1},{j + 1}"], -variables[f"py{j + 1},{rectangles[i][1] - 1}"]])
            else:
                if f"py{j + 1},{strip[1] - rectangles[i][1] + 1}" in variables:
                    cnf.append([-variables[f"ud{i + 1},{j + 1}"], variables[f"py{j + 1},{strip[1] - rectangles[i][1] + 1}"]])
            # if indomain(len(py[i]) - 1, height_rects[j] - 1)
            if strip[1] - rectangles[j][1] + 1 >= rectangles[i][1] - 1:
                cnf.append([-variables[f"ud{j + 1},{i + 1}"], -variables[f"py{i + 1},{rectangles[j][1] - 1}"]])
            else:
                if f"py{i + 1},{strip[1] - rectangles[j][1] + 1}" in variables:
                    cnf.append([-variables[f"ud{j + 1},{i + 1}"], variables[f"py{i + 1},{strip[1] - rectangles[j][1] + 1}"]])


    with Solver(name="mc") as solver:
        solver.append_formula(cnf)
        if solver.solve():
            pos = [[0 for i in range(2)] for j in range(len(rectangles))]
            model = solver.get_model()
            result = {}
            for var in model:
                if var > 0:
                    result[list(variables.keys())[list(variables.values()).index(var)]] = True
                else:
                    result[list(variables.keys())[list(variables.values()).index(-var)]] = False

            for i in range(len(rectangles)):
                for e in range(strip[0] - rectangles[i][0] + 1):
                    if result[f"px{i + 1},{e}"] == False and result[f"px{i + 1},{e + 1}"] == True:
                        # print(f"x{i + 1} = {e + 1}")
                        pos[i][0] = e + 1
                    if e == 0 and result[f"px{i + 1},{e}"] == True:
                        # print(f"x{i + 1} = 0")
                        pos[i][0] = 0
                for f in range(strip[1] - rectangles[i][1] + 1):
                    if result[f"py{i + 1},{f}"] == False and result[f"py{i + 1},{f + 1}"] == True:
                        # print(f"y{i + 1} = {f + 1}")
                        pos[i][1] = f + 1
                    if f == 0 and result[f"py{i + 1},{f}"] == True:
                        # print(f"y{i + 1} = 0")
                        pos[i][1] = 0
                        
            print(["sat", pos])
            return "sat"
        else:
            print("unsat")
            return "unsat"
    

def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path =  'output/'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)

    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

def solve_by_pysat(input, time, result, num_var, num_clause, name_lib):
    global id_counter
    id_counter += 1
    result_dict = {
        "ID": id_counter,
        "Problem": input,
        "Type": name_lib,
        "Time": time,
        "Result": result,
        "Variables": num_var,
        "Clauses": num_clause
    }
    write_to_xlsx(result_dict)

#  find all solutions of clauses
def find_all_solutions(solver, n_items):
    solutions_set = set() 
    elapsed_time = 0
    while solver.solve():
        # print(solver.time())
        elapsed_time = elapsed_time +  solver.time()
        solution = solver.get_model()
        print("Solution is: ", solution)
        temp = []
        for i in range (0, n_items):
            if (solution[i] > 0):
                temp.append(solution[i])
        solutions_set.add(tuple(temp))
        
        # Block the current solution
        blocking_clause = [-lit for lit in solution[0:10]]
        solver.add_clause(blocking_clause)
        
    print(elapsed_time)
    print(solutions_set)
    return [solutions_set, elapsed_time]

def max_profit_solution(rectangles, weights, profits, weight_bound, strip):
    # input data is [width, height, weight, profit] + [MAX_WEIGHT].
    # rectangles = [(1, 2), (1, 2), (2, 1), (1, 1),(1, 2), (1, 2), (2, 1), (1, 1),(3,4),(1,4)]   #width, height
    # weights = [10, 2, 6, 11, 21, 4, 8, 3, 8, 10]
    # profits = [20, 4, 3, 9, 13, 2, 3, 4, 7, 8]
    # weight_bound = 20
    formula = []
    isSat = ""
    num_items = len(weights)
    print(num_items)
    vars = list(range(1, num_items + 1))

    pbConfig = PBConfig()
    pbConfig.set_PB_Encoder(pblib.PB_BDD)
    pb2 = Pb2cnf(pbConfig)

    # Encoding by PB_Lib or SCPB
    pb2.encode_leq(weights, vars, weight_bound, formula, num_items+1)
    print("Formula: ",formula)
    solver = Glucose3(use_timer=True)
        
    for clause in formula:
        print(clause)
        solver.add_clause(clause)

    num_vars = solver.nof_vars()
    num_clauses = solver.nof_clauses()
    

    print("Number of variables: ", num_vars)
    print("Number of clauses: ", num_clauses)
    
    # all solutions that satify leq encoding with Weight include empty solution []
    arr_solution = find_all_solutions(solver, num_items)
    # print(arr_solution)
    all_solution = arr_solution[0]
    elapse_time = arr_solution[1]

    if (all_solution == {()}):
        isSat = "unsat"
        solve_by_pysat(len(weights), elapse_time, isSat, num_vars, num_clauses, "PB_ADDER")
        return
        
    #  Array that store [solution, max profit of solution].
    result = []
    sum_profit = set()

    for sol in all_solution:
        sum = 0
        temp = []
        if (sol):
            size = len(sol)
            for i in range(0, size):
                sum += profits[sol[i]-1]
            sum_profit.add(sum)
            temp = list(sol)
            temp.append(sum)
            result.append(temp)
    
    print("result:")
    sorted_result = sorted(result, key=lambda x: x[-1], reverse=True)
    print(sorted_result)

    sorted_rectangles = sorted(sum_profit, reverse=True)
    print(sorted_rectangles)
    

    num_solution = len(sorted_result)


    for i in range(0, num_solution):
        input = sorted_result[i]
        size = len(input)
        list_rect = []
        for j in range(0, size-1):
            if input[j]-1 <= len(rectangles):
                list_rect.append(rectangles[input[j]-1])

        status = opp_solution(list_rect, strip)
        if (status == "sat"):
            print("Sat")
            isSat = "sat"
            break
        else:
            isSat = "unsat"

    solve_by_pysat(len(weights), elapse_time, isSat, num_vars, num_clauses, "PB_BEST")

def solve_KPWL_Pblib():
    with open('test_input.txt', 'r') as file:
        # max_weight, min_profit
        # strip
        # list weights
        # list profits
        # width_i
        # height_i
        lines = file.readlines()
        len_file = len(lines)
        for i in range(0, len_file,6):
            bounds = list(map(int,lines[i].strip().split()))
            arr_strip = list(map(int,lines[i+1].strip().split()))
            weights = list(map(int, lines[i + 2].strip().split()))
            profits = list(map(int, lines[i + 3].strip().split()))
            arr_width =  list(map(int, lines[i + 4].strip().split()))
            arr_height =  list(map(int, lines[i + 5].strip().split()))
            rectangles = []
            
            l = len(arr_width)
            for j in range(0, l):
                rectangles.append((arr_width[j], arr_height[j]))
            weight_bound = bounds[0]
            

            strip = (arr_strip[0], arr_strip[1])

            max_profit_solution(rectangles, weights, profits, weight_bound, strip)

solve_KPWL_Pblib()
