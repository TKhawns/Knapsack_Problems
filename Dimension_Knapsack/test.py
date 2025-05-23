from pysat.formula import CNF
from pysat.solvers import Glucose3
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import time as timead

time_budget = 100
id_counter = 0

def interrupt(s): s.interrupt()

def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path =  'output/'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)

    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

def solve_by_pysat(input, time, result, num_var, num_clause, name_lib):
    global id_counter
    id_counter += 1
    result_dict = {
        "ID": id_counter,
        "Problem": input,
        "Type": name_lib,
        "Time": time,
        "Result": result,
        "Variables": num_var,
        "Clauses": num_clause
    }
    write_to_xlsx(result_dict)


def positive_range(end):
    if (end < 0):
        return []
    return range(end)

def opp_solution(rectangles, strip, length):
    cnf = CNF()
    variables = {}
    counter = 1
    for i in range(len(rectangles)):
        for j in range(len(rectangles)):
            variables[f"lr{i + 1},{j + 1}"] = counter  # lri,rj
            counter += 1
            variables[f"ud{i + 1},{j + 1}"] = counter  # uri,rj
            counter += 1
        for e in positive_range(strip[0] - rectangles[i][0] + 2):
            variables[f"px{i + 1},{e}"] = counter  # pxi,e
            counter += 1
        for f in positive_range(strip[1] - rectangles[i][1] + 2):
            variables[f"py{i + 1},{f}"] = counter  # pyi,f
            counter += 1

    # Add the 2-literal axiom clauses
    for i in range(len(rectangles)):
        for e in range(strip[0] - rectangles[i][0] + 1):  # -1 because we're using e+1 in the clause
            cnf.append([-variables[f"px{i + 1},{e}"], variables[f"px{i + 1},{e + 1}"]])
        for f in range(strip[1] - rectangles[i][1] + 1):  # -1 because we're using f+1 in the clause
            cnf.append([-variables[f"py{i + 1},{f}"], variables[f"py{i + 1},{f + 1}"]])

    # Add the 4-literal axiom clauses
    for i in range(len(rectangles)):
        for j in range(i + 1, len(rectangles)):
            cnf.append([variables[f"lr{i + 1},{j + 1}"], variables[f"lr{j + 1},{i + 1}"], variables[f"ud{i + 1},{j + 1}"],
                        variables[f"ud{j + 1},{i + 1}"]])

    # Add the 3-literal non-overlapping constraints
    for i in range(len(rectangles)):
        for j in range(i + 1, len(rectangles)):
            for e in positive_range(strip[0] - rectangles[i][0]) :
                if f"px{j + 1},{rectangles[i][0] - 1}" in variables:
                    cnf.append([-variables[f"lr{i + 1},{j + 1}"], -variables.get(f"px{j + 1},{rectangles[i][0] - 1}", 0)])
                if(strip[0] - rectangles[i][0] - rectangles[j][0] > 0) and f"px{i + 1},{strip[0] - rectangles[i][0] - rectangles[j][0]}" in variables:
                    cnf.append([-variables[f"lr{i + 1},{j + 1}"],
                                variables[f"px{i + 1},{strip[0] - rectangles[i][0] - rectangles[j][0]}"]])
                if f"px{j + 1},{e + rectangles[i][0]}" in variables:
                    cnf.append([-variables[f"lr{i + 1},{j + 1}"], variables[f"px{i + 1},{e}"],
                                -variables.get(f"px{j + 1},{e + rectangles[i][0]}", 0)])
                if f"px{i + 1},{e + rectangles[j][0]}" in variables:
                    cnf.append([-variables[f"lr{j + 1},{i + 1}"], variables.get(f"px{j + 1},{e}", 0),
                                -variables.get(f"px{i + 1},{e + rectangles[j][0]}", 0)])
                    
            for f in positive_range(strip[1] - rectangles[i][1]):
                if f"py{j + 1},{rectangles[i][1] - 1}" in variables:
                    cnf.append([-variables[f"ud{i + 1},{j + 1}"], -variables.get(f"py{j + 1},{rectangles[i][1] - 1}", 0)])
                if strip[1] - rectangles[i][1] - rectangles[j][1] > 0:
                    cnf.append([-variables[f"ud{i + 1},{j + 1}"],
                                variables[f"py{i + 1},{strip[1] - rectangles[i][1] - rectangles[j][1]}"]])
                if f"py{j + 1},{f + rectangles[i][1]}" in variables:
                    cnf.append([-variables[f"ud{i + 1},{j + 1}"], variables[f"py{i + 1},{f}"],
                            -variables.get(f"py{j + 1},{f + rectangles[i][1]}", 0)])

                if f"py{i + 1},{f + rectangles[j][1]}" in variables:
                    cnf.append([-variables[f"ud{j + 1},{i + 1}"], variables.get(f"py{j + 1},{f}", 0),
                                -variables.get(f"py{i + 1},{f + rectangles[j][1]}", 0)])

    for i in range(len(rectangles)):
        if f"px{i + 1},{strip[0] - rectangles[i][0]}" in variables:
            cnf.append([variables[f"px{i + 1},{strip[0] - rectangles[i][0]}"]])  # px(i, W-wi)
        if f"py{i + 1},{strip[1] - rectangles[i][1]}" in variables:
            cnf.append([variables[f"py{i + 1},{strip[1] - rectangles[i][1]}"]])  # py(i, H-hi)

    for i in range(len(rectangles)):
        for j in range(i + 1, len(rectangles)):
            # if indomain(len(px[j]) - 1, width_rects[i] - 1)
            if strip[0] - rectangles[i][0] + 1 >= rectangles[j][0] - 1:
                cnf.append([-variables[f"lr{i + 1},{j + 1}"], -variables[f"px{j + 1},{rectangles[i][0] - 1}"]])
            else:
                if f"px{j + 1},{strip[0] - rectangles[i][0] + 1}" in variables:
                    cnf.append([-variables[f"lr{i + 1},{j + 1}"], variables[f"px{j + 1},{strip[0] - rectangles[i][0] + 1}"]])
            # if indomain(len(px[i] - 1, width_rects[j] - 1)
            if strip[0] - rectangles[j][0] + 1 >= rectangles[i][0] - 1:
                cnf.append([-variables[f"lr{j + 1},{i + 1}"], -variables[f"px{i + 1},{rectangles[j][0] - 1}"]])
            else:
                if f"px{i + 1},{strip[0] - rectangles[j][0] + 1}" in variables:
                    cnf.append([-variables[f"lr{j + 1},{i + 1}"], variables[f"px{i + 1},{strip[0] - rectangles[j][0] + 1}"]])
            # if indomain(len(py[j]) - 1, height_rects[i] - 1)
            if strip[1] - rectangles[i][1] + 1 >= rectangles[j][1] - 1:
                cnf.append([-variables[f"ud{i + 1},{j + 1}"], -variables[f"py{j + 1},{rectangles[i][1] - 1}"]])
            else:
                if f"py{j + 1},{strip[1] - rectangles[i][1] + 1}" in variables:
                    cnf.append([-variables[f"ud{i + 1},{j + 1}"], variables[f"py{j + 1},{strip[1] - rectangles[i][1] + 1}"]])
            # if indomain(len(py[i]) - 1, height_rects[j] - 1)
            if strip[1] - rectangles[j][1] + 1 >= rectangles[i][1] - 1:
                cnf.append([-variables[f"ud{j + 1},{i + 1}"], -variables[f"py{i + 1},{rectangles[j][1] - 1}"]])
            else:
                if f"py{i + 1},{strip[1] - rectangles[j][1] + 1}" in variables:
                    cnf.append([-variables[f"ud{j + 1},{i + 1}"], variables[f"py{i + 1},{strip[1] - rectangles[j][1] + 1}"]])


    solver = Glucose3(use_timer=True)
    solver.append_formula(cnf)
    
    sat_status = solver.solve_limited(expect_interrupt = True)

    if sat_status is False:
        elapsed_time = format(solver.time())
        result = "unsat"
        time = elapsed_time   
        return "unsat"
        # print("No solutions found")
    else:
        solution = solver.get_model()
        if solution is None:
            result = "timeout"
            return "timeout" 
        else:
            pos = [[0 for i in range(2)] for j in range(len(rectangles))]
            result = {}
            for var in solution:
                if var > 0:
                    result[list(variables.keys())[list(variables.values()).index(var)]] = True
                else:
                    result[list(variables.keys())[list(variables.values()).index(-var)]] = False
            #print(result)

            for i in range(len(rectangles)):
                for e in range(strip[0] - rectangles[i][0] + 1):
                    if result[f"px{i + 1},{e}"] == False and result[f"px{i + 1},{e + 1}"] == True:
                        # print(f"x{i + 1} = {e + 1}")
                        pos[i][0] = e + 1
                    if e == 0 and result[f"px{i + 1},{e}"] == True:
                        # print(f"x{i + 1} = 0")
                        pos[i][0] = 0
                for f in range(strip[1] - rectangles[i][1] + 1):
                    if result[f"py{i + 1},{f}"] == False and result[f"py{i + 1},{f + 1}"] == True:
                        # print(f"y{i + 1} = {f + 1}")
                        pos[i][1] = f + 1
                    if f == 0 and result[f"py{i + 1},{f}"] == True:
                        # print(f"y{i + 1} = 0")
                        pos[i][1] = 0
                        
            print(["sat", pos])
            elapsed_time = format(solver.time())
            result = "sat"
            time = elapsed_time
            # return "sat"
            
            # timer.cancel()
            # solver.delete()

            # solve_by_pysat(str(length), str(time), result, num_vars, num_clauses, "Glucose3")
            return "sat" 
                
        # if solver.solve():
        #     pos = [[0 for i in range(2)] for j in range(len(rectangles))]
        #     model = solver.get_model()
        #     # print("SAT")

        #     result = {}
        #     for var in model:
        #         if var > 0:
        #             result[list(variables.keys())[list(variables.values()).index(var)]] = True
        #         else:
        #             result[list(variables.keys())[list(variables.values()).index(-var)]] = False
        #     #print(result)

        #     for i in range(len(rectangles)):
        #         for e in range(strip[0] - rectangles[i][0] + 1):
        #             if result[f"px{i + 1},{e}"] == False and result[f"px{i + 1},{e + 1}"] == True:
        #                 # print(f"x{i + 1} = {e + 1}")
        #                 pos[i][0] = e + 1
        #             if e == 0 and result[f"px{i + 1},{e}"] == True:
        #                 # print(f"x{i + 1} = 0")
        #                 pos[i][0] = 0
        #         for f in range(strip[1] - rectangles[i][1] + 1):
        #             if result[f"py{i + 1},{f}"] == False and result[f"py{i + 1},{f + 1}"] == True:
        #                 # print(f"y{i + 1} = {f + 1}")
        #                 pos[i][1] = f + 1
        #             if f == 0 and result[f"py{i + 1},{f}"] == True:
        #                 # print(f"y{i + 1} = 0")
        #                 pos[i][1] = 0
                        
        #     print(["sat", pos])
        #     return "sat"
        # else:
        #     print("unsat")
        #     return "unsat"

# def max_profit_solution():
#     rectangles = [(70, 86, 6020), (66, 148, 9768), (83, 140, 11620), (87, 141, 12267), (114, 118, 13452), (120, 160, 19200), (143, 166, 23738), (167, 152, 25384), (167, 184, 30728)]
#     sorted_rectangles = sorted(rectangles, key=lambda x: x[2])

#     for i in range(0, len(sorted_rectangles)):
#         print("New array")
#         print(sorted_rectangles)

#         sum = 0
#         status = opp_solution(sorted_rectangles)
#         # if (status == "sat"):
#         #     for i in range(0, len(sorted_rectangles)):
#         #         sum += sorted_rectangles[i][2]
#         #     print("Profit: ", sum)
#         #     break
#         # else:
#         #     sorted_rectangles.pop(0)

def max_profit_solution():
    with open('dataset.txt', 'r') as file:
        lines = file.readlines()
        len_file = len(lines)

        for i in range(0, len_file, 4):
            result = ""
            time = 0
            list_strip = list(map(int,lines[i].strip().split()))
            widths = list(map(int, lines[i + 1].strip().split()))
            heights = list(map(int, lines[i + 2].strip().split()))
            profits = list(map(int, lines[i + 3].strip().split()))
            strip = (list_strip[0], list_strip[1])
            num_items = len(widths)

            rectangles = []
            for j in range (0, num_items):
                item = (widths[j], heights[j], profits[j])
                rectangles.append(item)

            sorted_rectangles = sorted(rectangles, key=lambda x: x[2])
            result = [(x[0], x[1]) for x in sorted_rectangles]
            print(result)


            start = timead.time()
            elap_time = 0
            solution = ""

            for i in range(0, len(result)):
                status = opp_solution(result, strip, num_items)
                if (len(result) == 0):
                    elap_time = timead.time() - start
                    solution = "unsat"
                    break
                if (status == "sat"):
                    elap_time = timead.time() - start
                    solution = "sat"
                    break
                elif (status == "unsat"):
                    result.pop(0)

            solve_by_pysat(str(num_items), str(elap_time), solution, 0, 0, "Glucose3")


max_profit_solution()
