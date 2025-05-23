from pysat.formula import CNF
from pysat.solvers import Solver
from pysat.solvers import Glucose3
from pypblib import pblib
from pypblib.pblib import PBConfig, Pb2cnf
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import matplotlib.pyplot as plt

# This version solve rotation and symmetry by using PB_LIB.

id_counter = 0

def display_solution(strip, rectangles, pos_circuits, rotation):
    # define Matplotlib figure and axis
    ax = plt.subplots()
    ax = plt.gca()
    plt.title(strip)
    n = len(rectangles)

    if len(pos_circuits) > 0:
        for i in range(n):
            rect = plt.Rectangle(pos_circuits[i],
                                 rectangles[i][0] if not rotation[i] else rectangles[i][1],
                                 rectangles[i][1] if not rotation[i] else rectangles[i][0],
                                 edgecolor="#333")
            ax.add_patch(rect)

    ax.set_xlim(0, strip[0])
    ax.set_ylim(0, strip[1] + 1)
    ax.set_xticks(range(strip[0] + 1))
    ax.set_yticks(range(strip[1] + 1))
    ax.set_xlabel('width')
    ax.set_ylabel('height')
    # display plot
    plt.show()

def positive_range(end):
    if (end < 0):
        return []
    return range(end)

def opp_solution(rectangles, strip):
# Define the variables
    cnf = CNF()
    width = strip[0]
    height = strip[1]
    variables = {}
    counter = 1
    n = len(rectangles)

    # create lr, ud, px, py variables from 1 to N.
    # SAT Encoding of 2OPP
    for i in range(n):
        for j in range(n):
            if i != j:
                variables[f"lr{i + 1},{j + 1}"] = counter  # lri,rj
                counter += 1
                variables[f"ud{i + 1},{j + 1}"] = counter  # uri,rj
                counter += 1
        for e in range(width):
            variables[f"px{i + 1},{e}"] = counter  # pxi,e
            counter += 1
        for f in range(height):
            variables[f"py{i + 1},{f}"] = counter  # pyi,f
            counter += 1

    # Rotated variables
    for i in range(n):
        variables[f"r{i + 1}"] = counter
        counter += 1

    # Add the 2-literal axiom clauses (order constraint)
    # Formula (3).
    for i in range(n):
        for e in range(width - 1):  # -1 because we're using e+1 in the clause
            cnf.append([-variables[f"px{i + 1},{e}"],
                        variables[f"px{i + 1},{e + 1}"]])
        for f in range(height - 1):  # -1 because we're using f+1 in the clause
            cnf.append([-variables[f"py{i + 1},{f}"],
                        variables[f"py{i + 1},{f + 1}"]])
    # Add the 3-literal non-overlapping constraints
    # Formula (4).
    def non_overlapping(rotated, i, j, h1, h2, v1, v2):
        if not rotated:
            i_width = rectangles[i][0]
            i_height = rectangles[i][1]
            j_width = rectangles[j][0]
            j_height = rectangles[j][1]
            i_rotation = variables[f"r{i + 1}"]
            j_rotation = variables[f"r{j + 1}"]
        else:
            i_width = rectangles[i][1]
            i_height = rectangles[i][0]
            j_width = rectangles[j][1]
            j_height = rectangles[j][0]
            i_rotation = -variables[f"r{i + 1}"]
            j_rotation = -variables[f"r{j + 1}"]

        # Square symmertry breaking, if i is square than it cannot be rotated
        if i_width == i_height and rotated:
            i_square = True
            cnf.append([-variables[f"r{i + 1}"]])
        else:
            i_square = False

        if j_width == j_height and rotated:
            j_square = True
            cnf.append([-variables[f"r{j + 1}"]])
        else:
            j_square = False

        # lri,j v lrj,i v udi,j v udj,i
        four_literal = []
        if h1: four_literal.append(variables[f"lr{i + 1},{j + 1}"])
        if h2: four_literal.append(variables[f"lr{j + 1},{i + 1}"])
        if v1: four_literal.append(variables[f"ud{i + 1},{j + 1}"])
        if v2: four_literal.append(variables[f"ud{j + 1},{i + 1}"])

        cnf.append(four_literal + [i_rotation])
        cnf.append(four_literal + [j_rotation])

        # ¬lri, j ∨ ¬pxj, e
        if h1 and not i_square:
            for e in range(min(width, i_width)):
                    cnf.append([i_rotation,
                                -variables[f"lr{i + 1},{j + 1}"],
                                -variables[f"px{j + 1},{e}"]])
        # ¬lrj,i ∨ ¬pxi,e
        if h2 and not j_square:
            for e in range(min(width, j_width)):
                    cnf.append([j_rotation,
                                -variables[f"lr{j + 1},{i + 1}"],
                                -variables[f"px{i + 1},{e}"]])
        # ¬udi,j ∨ ¬pyj,f
        if v1 and not i_square:
            for f in range(min(height, i_height)):
                    cnf.append([i_rotation,
                                -variables[f"ud{i + 1},{j + 1}"],
                                -variables[f"py{j + 1},{f}"]])
        # ¬udj, i ∨ ¬pyi, f,
        if v2 and not j_square:
            for f in range(min(height, j_height)):
                    cnf.append([j_rotation,
                                -variables[f"ud{j + 1},{i + 1}"],
                                -variables[f"py{i + 1},{f}"]])

        for e in positive_range(width - i_width):
            # ¬lri,j ∨ ¬pxj,e+wi ∨ pxi,e
            if h1 and not i_square:
                    cnf.append([i_rotation,
                                -variables[f"lr{i + 1},{j + 1}"],
                                variables[f"px{i + 1},{e}"],
                                -variables[f"px{j + 1},{e + i_width}"]])

        for e in positive_range(width - j_width):
            # ¬lrj,i ∨ ¬pxi,e+wj ∨ pxj,e
            if h2 and not j_square:
                    cnf.append([j_rotation,
                                -variables[f"lr{j + 1},{i + 1}"],
                                variables[f"px{j + 1},{e}"],
                                -variables[f"px{i + 1},{e + j_width}"]])

        for f in positive_range(height - i_height):
            # udi,j ∨ ¬pyj,f+hi ∨ pxi,e
            if v1 and not i_square:
                    cnf.append([i_rotation,
                                -variables[f"ud{i + 1},{j + 1}"],
                                variables[f"py{i + 1},{f}"],
                                -variables[f"py{j + 1},{f + i_height}"]])
        for f in positive_range(height - j_height):
            # ¬udj,i ∨ ¬pyi,f+hj ∨ pxj,f
            if v2 and not j_square:
                    cnf.append([j_rotation,
                                -variables[f"ud{j + 1},{i + 1}"],
                                variables[f"py{j + 1},{f}"],
                                -variables[f"py{i + 1},{f + j_height}"]])

    for i in range(n):
        for j in range(i + 1, n):
            # lri,j ∨ lrj,i ∨ udi,j ∨ udj,i
            #Large-rectangles horizontal
            if min(rectangles[i][0], rectangles[i][1]) + min(rectangles[j][0], rectangles[j][1]) > width:
                non_overlapping(False, i, j, False, False, True, True)
                non_overlapping(True, i, j, False, False, True, True)
            # Large rectangles vertical
            elif min(rectangles[i][0], rectangles[i][1]) + min(rectangles[j][0], rectangles[j][1]) > height:
                non_overlapping(False, i, j, True, True, False, False)
                non_overlapping(True, i, j, True, True, False, False)

            # Same rectangle and is a square
            elif rectangles[i] == rectangles[j]:
                if rectangles[i][0] == rectangles[i][1]:
                    cnf.append([-variables[f"r{i + 1}"]])
                    cnf.append([-variables[f"r{j + 1}"]])
                    non_overlapping(False,i ,j, True, True, True, True)
                else:
                    non_overlapping(False, i, j, True, True, True, True)
                    non_overlapping(True, i, j, True, True, True, True)
            # normal rectangles
            else:
                non_overlapping(False, i, j, True, True, True, True)
                non_overlapping(True, i, j, True, True, True, True)


 # Domain encoding to ensure every rectangle stays inside strip's boundary
    for i in range(n):
        if rectangles[i][0] > width: #if rectangle[i]'s width larger than strip's width, it has to be rotated
            cnf.append([variables[f"r{i + 1}"]])
        else:
            for e in range(width - rectangles[i][0], width):
                    cnf.append([variables[f"r{i + 1}"],
                                variables[f"px{i + 1},{e}"]])
        if rectangles[i][1] > height:
            cnf.append([variables[f"r{i + 1}"]])
        else:
            for f in range(height - rectangles[i][1], height):
                    cnf.append([variables[f"r{i + 1}"],
                                variables[f"py{i + 1},{f}"]])

        # Rotated
        if rectangles[i][1] > width:
            cnf.append([-variables[f"r{i + 1}"]])
        else:
            for e in range(width - rectangles[i][1], width):
                    cnf.append([-variables[f"r{i + 1}"],
                                variables[f"px{i + 1},{e}"]])
        if rectangles[i][0] > height:
            cnf.append([-variables[f"r{i + 1}"]])
        else:
            for f in range(height - rectangles[i][0], height):
                cnf.append([-variables[f"r{i + 1}"],
                            variables[f"py{i + 1},{f}"]])

    # add all clauses to SAT solver
    elapse_time = 0
    with Glucose3() as solver:
        solver.append_formula(cnf)
        if solver.solve():
            rotation = []
            elapse_time += solver.time()
            pos = [[0 for i in range(2)] for j in range(n)]
            model = solver.get_model()
            result = {}
            for var in model:
                if var > 0:
                    result[list(variables.keys())[list(variables.values()).index(var)]] = True
                else:
                    result[list(variables.keys())[list(variables.values()).index(-var)]] = False
            #print(result)

            # from SAT result, decode into rectangles' position
            for i in range(n):
                rotation.append(result[f"r{i + 1}"])
                for e in range(width - rectangles[i][0] + 1):
                    if result[f"px{i + 1},{e}"] == False and result[f"px{i + 1},{e + 1}"] == True:
                        print(f"x{i + 1} = {e + 1}")
                        pos[i][0] = e + 1
                    if e == 0 and result[f"px{i + 1},{e}"] == True:
                        print(f"x{i + 1} = 0")
                        pos[i][0] = 0
                for f in range(height - rectangles[i][1] + 1):
                    if result[f"py{i + 1},{f}"] == False and result[f"py{i + 1},{f + 1}"] == True:
                        print(f"y{i + 1} = {f + 1}")
                        pos[i][1] = f + 1
                    if f == 0 and result[f"py{i + 1},{f}"] == True:
                        print(f"y{i + 1} = 0")
                        pos[i][1] = 0

            display_solution(strip, rectangles, pos, rotation)
            return ["sat", elapse_time]

        else:
            return ["unsat", elapse_time]
    

def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path =  'output/'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)

    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

def solve_by_pysat(input, time, result, num_var, num_clause, name_lib):
    global id_counter
    id_counter += 1
    result_dict = {
        "ID": id_counter,
        "Problem": input,
        "Type": name_lib,
        "Time": time,
        "Result": result,
        "Variables": num_var,
        "Clauses": num_clause
    }
    write_to_xlsx(result_dict)

#  find all solutions of clauses
def find_all_solutions(solver, n_items):
    solutions_set = set() 
    elapsed_time = 0
    while solver.solve():
        if (elapsed_time > 100):
            return [solutions_set,"timeout"]
        elapsed_time = elapsed_time +  solver.time()
        solution = solver.get_model()
        print("Solution", solution)

        temp = []
        for i in range (0, n_items):
            if (solution[i] > 0):
                temp.append(solution[i])
        solutions_set.add(tuple(temp))
        
        # Block the current solution
        blocking_clause = [-lit for lit in solution[0:n_items]]
        solver.add_clause(blocking_clause)
        
    return [solutions_set, elapsed_time]

def max_profit_solution(rectangles, weights, profits, weight_bound, strip):
    # input data is [width, height, weight, profit] + [MAX_WEIGHT].
    isSat = ""
    num_items = len(weights)
    vars = list(range(1, num_items + 1))

    # Using PB_LIB
    pbConfig = PBConfig()
    pbConfig.set_PB_Encoder(pblib.PB_BDD)
    pb2 = Pb2cnf(pbConfig)

    # Return formula of constraint <= Weight.
    formula = []
    pb2.encode_leq(weights, vars, weight_bound, formula, num_items+1)

    solver = Glucose3(use_timer=True)

    if formula == []:
        solver.add_clause(vars)

    for clause in formula:
        solver.add_clause(clause)

    num_vars = solver.nof_vars()
    num_clauses = solver.nof_clauses()
    
    print("Number of variables: ", solver.nof_vars())
    print("Number of clauses: ", solver.nof_clauses())
    
    # all solutions that satify leq encoding with Weight include empty solution []
    # arr_solution = [set_solution, elapse_time] of find_all_solution with constraints <= Max_weight
    arr_solution = find_all_solutions(solver, num_items)

    all_solution = arr_solution[0]
    elapse_time = arr_solution[1]

    # if no solution or timeout
    if (all_solution == {()}):
        isSat = "unsat"
        solve_by_pysat(len(weights), elapse_time, isSat, num_vars, num_clauses, "PB_BDD")
        solver.delete()
        return
    elif elapse_time == "timeout":
        isSat = "timeout"
        solve_by_pysat(len(weights), "timeout", isSat, num_vars, num_clauses, "PB_BDD")
        solver.delete()
        return
        
    # If exist solution, sort list of rectangles by profit reverse.
    #  Array that store [solution, max profit of solution].
    result = []
    sum_profit = set()

    for sol in all_solution:
        temp = []
        if sol:
            total_profit = sum(profits[item - 1] for item in sol)
            sum_profit.add(total_profit)
            temp = list(sol)
            temp.append(total_profit)
            result.append(temp)
    
    sorted_result = sorted(result, key=lambda x: x[-1], reverse=True)

    # sorted_profit_rectangles = sorted(sum_profit, reverse=True)
    # print("Sorted profit list", sorted_profit_rectangles)

    num_solution = len(sorted_result)

    for i in range(0, num_solution):
        input = sorted_result[i]
        size = len(input)
        list_rect = []

        for j in range(0, size-1):
            list_rect.append(rectangles[input[j]-1])

        max_profit = sorted_result[i][-1]


        status = opp_solution(list_rect, strip) # Return result and timer
        if (status[0] == "sat"):
            elapse_time += status[1]
            print("Max profit", max_profit)
            print("Time:", elapse_time)
            isSat = "sat"
            break
        elif (status[0] == "unsat"):
            isSat = "unsat"
            elapse_time += status[1]
            continue

    solve_by_pysat(len(weights), elapse_time, isSat, num_vars, num_clauses, "PB_BDD")
    solver.delete()


def solve_KPWL_Pblib():
    with open('../dataset/test_input.txt', 'r') as file:
        # max_weight, min_profit
        # strip
        # list weights
        # list profits
        # width_i
        # height_i
        lines = file.readlines()
        len_file = len(lines)
        for i in range(0, len_file,6):
            bounds = list(map(int,lines[i].strip().split()))
            arr_strip = list(map(int,lines[i+1].strip().split()))
            weights = list(map(int, lines[i + 2].strip().split()))
            profits = list(map(int, lines[i + 3].strip().split()))
            arr_width =  list(map(int, lines[i + 4].strip().split()))
            arr_height =  list(map(int, lines[i + 5].strip().split()))
            rectangles = []
            
            l = len(arr_width)
            for j in range(0, l):
                rectangles.append((arr_width[j], arr_height[j]))
            weight_bound = bounds[0]
            
            strip = (arr_strip[0], arr_strip[1])

            max_profit_solution(rectangles, weights, profits, weight_bound, strip)

solve_KPWL_Pblib()