from cplex import Cplex
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import matplotlib.pyplot as plt
import time

id_counter = 0

def display_solution(strip, rectangles, pos_circuits):
    # ... existing code ...
    ax = plt.gca()
    plt.title(strip)

    if len(pos_circuits) > 0:
        for i in range(len(rectangles)):
            rect = plt.Rectangle(pos_circuits[i], *rectangles[i], edgecolor="#333")
            ax.add_patch(rect)

    ax.set_xlim(0, strip[0])
    ax.set_ylim(0, strip[1] + 1)
    ax.set_xticks(range(strip[0] + 1))
    ax.set_yticks(range(strip[1] + 1))
    ax.set_xlabel('width')
    ax.set_ylabel('height')
    plt.show()

def write_to_xlsx(result_dict):
    # ... existing code ...
    excel_results = []
    excel_results.append(result_dict)

    output_path =  'output_rotate/'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)

    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

def export_csv(problem_name, method_name, time, result, num_var, num_clause, max_profit, weight):
    # ... existing code ...
    global id_counter
    id_counter += 1
    result_dict = {
        "No": id_counter,
        "Problem": problem_name,
        "Type": method_name,
        "Time": time,
        "Result": result,
        "Variables": num_var,
        "Clauses": num_clause,
        "Max profit": max_profit,
        "Weight": weight
    }
    write_to_xlsx(result_dict)


def CPLEX_constraints(rectangles, W, H, profits, mid, C, weights):
    # ... existing initialization code ...
    
    # Create variables with proper types
    n = len(rectangles)
    cpx = Cplex()
    cpx.set_problem_name("Rectangle Packing Problem")
    x_names = [f"x_{i}" for i in range(n)]
    cpx.variables.add(
        obj=[0.0]*n, 
        lb=[0.0]*n, 
        ub=[float(W)]*n, 
        types=["I"]*n,  # Integer variables
        names=x_names
    )
    
    y_names = [f"y_{i}" for i in range(n)]
    cpx.variables.add(
        obj=[0.0]*n,
        lb=[0.0]*n,
        ub=[float(H)]*n,
        types=["I"]*n,
        names=y_names
    )
    
    r_names = [f"r_{i}" for i in range(n)]
    cpx.variables.add(
        obj=[0.0]*n,
        lb=[0.0]*n,
        ub=[1.0]*n,
        types=["B"]*n,  # Binary variables for rotation
        names=r_names
    )
    
    xs_names = [f"xs_{i}" for i in range(n)]
    cpx.variables.add(
        obj=[0.0]*n,
        lb=[0.0]*n,
        ub=[1.0]*n,
        types=["B"]*n,  # Binary variables for selection
        names=xs_names
    )

    # Placement constraints with big-M formulation
    M = max(W, H) + 1  # Big-M value
    for i in range(n):
        wi, hi, _, _ = rectangles[i]

        # Width constraint with selection logic
        cpx.linear_constraints.add(
            lin_expr=[[
                [x_names[i], r_names[i], xs_names[i]],
                [1.0, -(wi-hi), M]
            ]],
            senses=["L"],
            rhs=[W + M - wi]
        )
        
        # Height constraint with selection logic
        cpx.linear_constraints.add(
            lin_expr=[[
                [y_names[i], r_names[i], xs_names[i]],
                [1.0, -(hi-wi), M]
            ]],
            senses=["L"],
            rhs=[H + M - hi]
        )

        # Rotation constraints
        if wi > W or hi > H:
            cpx.linear_constraints.add(
                lin_expr=[[[r_names[i]], [1.0]]],
                senses=["E"],
                rhs=[1.0]
            )
        if wi == hi or (wi > H or hi > W):
            cpx.linear_constraints.add(
                lin_expr=[[[r_names[i]], [1.0]]],
                senses=["E"],
                rhs=[0.0]
            )

    # Non-overlapping constraints
    for i in range(n):
        for j in range(i+1, n):
            wi, hi, _, _ = rectangles[i]
            wj, hj, _, _ = rectangles[j]
            
            # Create 4 binary variables for disjunctions
            b_names = [f"b1_{i}_{j}", f"b2_{i}_{j}", f"b3_{i}_{j}", f"b4_{i}_{j}"]
            cpx.variables.add(
                types=["B"]*4,
                names=b_names
            )
            
            # Corrected horizontal separation constraints
            cpx.linear_constraints.add(
                lin_expr=[[
                    [x_names[i], r_names[i], x_names[j], b_names[0]],
                    [1.0, (hi-wi), -1.0, M]  # Changed last coefficient from -M to M
                ]],
                senses=["L"],
                rhs=[M - wi]
            )
            cpx.linear_constraints.add(
                lin_expr=[[
                    [x_names[j], r_names[j], x_names[i], b_names[1]],
                    [1.0, (hj-wj), -1.0, M]  # Changed last coefficient from -M to M
                ]],
                senses=["L"],
                rhs=[M - wj]
            )
            
            # Corrected vertical separation constraints
            cpx.linear_constraints.add(
                lin_expr=[[
                    [y_names[i], r_names[i], y_names[j], b_names[2]],
                    [1.0, (wi-hi), -1.0, M]  # Changed last coefficient from -M to M
                ]],
                senses=["L"],
                rhs=[M - hi]
            )
            cpx.linear_constraints.add(
                lin_expr=[[
                    [y_names[j], r_names[j], y_names[i], b_names[3]],
                    [1.0, (wj-hj), -1.0, M]  # Changed last coefficient from -M to M
                ]],
                senses=["L"],
                rhs=[M - hj]
            )
            
            # At least one disjunction must hold
            cpx.linear_constraints.add(
                lin_expr=[[
                    b_names + [xs_names[i], xs_names[j]],
                    [1.0]*4 + [-1.0, -1.0]
                ]],
                senses=["G"],
                rhs=[-1.0]
            )

    # ... rest of the code remains same as original CPLEX implementation ...
    # (profit constraint, weight constraint, symmetry breaking, and solution handling)
    
    # Profit constraint: sum(profits[i] * xs[i]) >= mid
    profit_indices = []
    profit_values = []
    for i in range(n):
        profit_indices.append(xs_names[i])
        profit_values.append(float(profits[i]))
    
    cpx.linear_constraints.add(
        lin_expr=[[profit_indices, profit_values]],
        senses=["G"],
        rhs=[float(mid)]
    )
    
    # Weight constraint: sum(weights[i] * xs[i]) <= C
    weight_indices = []
    weight_values = []
    for i in range(n):
        weight_indices.append(xs_names[i])
        weight_values.append(float(weights[i]))
    
    cpx.linear_constraints.add(
        lin_expr=[[weight_indices, weight_values]],
        senses=["L"],
        rhs=[float(C)]
    )
    
    # Symmetry breaking constraints
    # [1] If two rectangles are identical, force lexicographic order
    for i in range(n):
        for j in range(i+1, n):
            if rectangles[i] == rectangles[j]:
                # x[i] <= x[j]
                cpx.linear_constraints.add(
                    lin_expr=[[
                        [x_names[i], x_names[j]], 
                        [1.0, -1.0]
                    ]],
                    senses=["L"],
                    rhs=[0.0]
                )
                
                # y[i] <= y[j]
                cpx.linear_constraints.add(
                    lin_expr=[[
                        [y_names[i], y_names[j]], 
                        [1.0, -1.0]
                    ]],
                    senses=["L"],
                    rhs=[0.0]
                )
    
    # [2] For rectangles with maximum width, restrict horizontal domain
    max_width_val = max(rect[0] for rect in rectangles)
    for i, rect in enumerate(rectangles):
        wi, hi, _, _ = rect
        if wi == max_width_val:
            max_domain = (W - wi) // 2
            cpx.linear_constraints.add(
                lin_expr=[[
                    [x_names[i]], 
                    [1.0]
                ]],
                senses=["L"],
                rhs=[float(max_domain)]
            )
    
    # Solve the model
    try:
        cpx.solve()
        status = cpx.solution.get_status()
        
        # Check if a solution was found (1 = optimal, 101/102 = feasible)
        if status in [1, 101, 102]:
            selected_rectangles = []
            pos = []
            
            # Get the values of the decision variables
            x_vals = cpx.solution.get_values(x_names)
            y_vals = cpx.solution.get_values(y_names)
            r_vals = cpx.solution.get_values(r_names)
            xs_vals = cpx.solution.get_values(xs_names)
            
            total_profit = 0
            total_weight = 0
            
            for i in range(n):
                if xs_vals[i] > 0.5:  # If rectangle i is selected (binary value close to 1)
                    selected_rectangles.append(rectangles[i])
                    pos.append([x_vals[i], y_vals[i]])
                    total_profit += rectangles[i][2]
                    total_weight += rectangles[i][3]
            
            num_vars = cpx.variables.get_num()
            num_constraints = cpx.linear_constraints.get_num()
            
            return ["SAT", num_vars, num_constraints, total_profit, total_weight]
        else:
            num_vars = cpx.variables.get_num()
            num_constraints = cpx.linear_constraints.get_num()
            return ["UNSAT", num_vars, num_constraints]
    
    except Exception as e:
        print(f"Error solving CPLEX model: {e}")
        return ["UNSAT", 0, 0]


def max_profit_solution():
    folder_path = '../dataset/soft'
    for file_name in os.listdir(folder_path):
        file_path = os.path.join(folder_path, file_name)
        if os.path.isfile(file_path) and file_name.endswith('.txt'):
            print(f"Processing file: {file_name}")
            with open(file_path, 'r') as file:
                lines = file.readlines()
                list_strip = list(map(int,lines[0].strip().split()))
                widths = list(map(int, lines[1].strip().split()))
                heights = list(map(int, lines[2].strip().split()))
                profits = list(map(int, lines[3].strip().split()))
                weights = list(map(int, lines[4].strip().split()))
                num_items = len(widths)

                # Get list rectangles from input data.
                rectangles = []
                prev_sat = []
                lower_bound = min(profits)
                upper_bound = sum(profits)
                for j in range (0, num_items):
                    item = (widths[j], heights[j], profits[j], weights[j])
                    rectangles.append(item)

                isExport = ""
                started_time = time.time()
                while (lower_bound + 1 < upper_bound):
                    if time.time() - started_time >= 600: 
                        isExport = "TIMEOUT"
                        if prev_sat == []:
                            export_csv(file_name, "cplex_rotate_symmetry", time.time() - started_time, "TIMEOUT", 0, 0, 0, 0)
                            break
                        export_csv(file_name, "cplex_rotate_symmetry", time.time() - started_time, "TIMEOUT", 0, 0, prev_sat[3], prev_sat[4])
                        break
                    mid = lower_bound + (upper_bound - lower_bound) // 2
                    status = CPLEX_constraints(rectangles, list_strip[0], list_strip[1], profits, mid, list_strip[2], weights)
                    if (status[0] == "UNSAT"):
                        upper_bound = mid
                        continue
                    elif (status[0] == "SAT"):
                        prev_sat = status
                        lower_bound = mid
                        continue
                ended_time = time.time()

                if (isExport != "TIMEOUT"):
                    if prev_sat and (prev_sat[0]) == "SAT":
                        export_csv(file_name, "cplex_rotate_symmetry", ended_time - started_time, "SAT", prev_sat[1], prev_sat[2], prev_sat[3], prev_sat[4])
                    else: export_csv(file_name, "cplex_rotate_symmetry", ended_time - started_time, "UNSAT", 0, 0, 0, 0)

max_profit_solution()