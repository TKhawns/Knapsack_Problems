from ortools.sat.python import cp_model
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import matplotlib.pyplot as plt
import time

id_counter = 0

def display_solution(strip, rectangles, pos_circuits):
    ax = plt.subplots()
    ax = plt.gca()
    plt.title(strip)

    if len(pos_circuits) > 0:
        for i in range(len(rectangles)):
            rect = plt.Rectangle(pos_circuits[i], *rectangles[i], edgecolor="#333")
            ax.add_patch(rect)

    ax.set_xlim(0, strip[0])
    ax.set_ylim(0, strip[1] + 1)
    ax.set_xticks(range(strip[0] + 1))
    ax.set_yticks(range(strip[1] + 1))
    ax.set_xlabel('width')
    ax.set_ylabel('height')
    plt.show()

def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path =  'output_non_rotate/'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)

    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

def export_csv(problem_name, method_name, time, result, num_var, num_clause, max_profit, weight):
    global id_counter
    id_counter += 1
    result_dict = {
        "No": id_counter,
        "Problem": problem_name,
        "Type": method_name,
        "Time": time,
        "Result": result,
        "Variables": num_var,
        "Clauses": num_clause,
        "Max profit": max_profit,
        "Weight": weight
    }
    write_to_xlsx(result_dict)


def CPSAT_constraints(rectangles, W, H, profits, mid, C, weights):
    model = cp_model.CpModel()
    n = len(rectangles)
    
    # Decision variables
    x = [model.NewIntVar(0, W, f'x_{i}') for i in range(n)]
    y = [model.NewIntVar(0, H, f'y_{i}') for i in range(n)]
    xs = [model.NewBoolVar(f'sel_{i}') for i in range(n)]  # Selection variable
    
    # Placement constraints (ensuring rectangles stay within the bin)
    for i in range(n):
        wi, hi, _, _ = rectangles[i]
        model.Add(x[i] + wi <= W).OnlyEnforceIf(xs[i])
        model.Add(y[i] + hi <= H).OnlyEnforceIf(xs[i])
    
    # Non-overlapping constraints
    for i in range(n):
        for j in range(i + 1, n):
            wi, hi, _, _ = rectangles[i]
            wj, hj, _, _ = rectangles[j]
            
            # Create Boolean literals for the four disjuncts
            b1 = model.NewBoolVar(f'non_overlap_{i}_{j}_1')  # i is to the left of j
            b2 = model.NewBoolVar(f'non_overlap_{i}_{j}_2')  # j is to the left of i
            b3 = model.NewBoolVar(f'non_overlap_{i}_{j}_3')  # i is below j
            b4 = model.NewBoolVar(f'non_overlap_{i}_{j}_4')  # j is below i
            
            # Enforce non-overlapping conditions
            model.Add(x[i] + wi <= x[j]).OnlyEnforceIf(b1)
            model.Add(x[i] + wi > x[j]).OnlyEnforceIf(b1.Not())
            
            model.Add(x[j] + wj <= x[i]).OnlyEnforceIf(b2)
            model.Add(x[j] + wj > x[i]).OnlyEnforceIf(b2.Not())
            
            model.Add(y[i] + hi <= y[j]).OnlyEnforceIf(b3)
            model.Add(y[i] + hi > y[j]).OnlyEnforceIf(b3.Not())
            
            model.Add(y[j] + hj <= y[i]).OnlyEnforceIf(b4)
            model.Add(y[j] + hj > y[i]).OnlyEnforceIf(b4.Not())
            
            # Ensure at least one separation condition is met
            model.AddBoolOr([xs[i].Not(), xs[j].Not(), b1, b2, b3, b4])
    
    # Profit constraint
    model.Add(sum(profits[i] * xs[i] for i in range(n)) >= mid)
    model.Add(sum(weights[i] * xs[i] for i in range(n)) <= C)
    

    # Symmetry breaking constraints.
    # [1] If two rectangles are identical then force a lexicographic order.
    for i in range(n):
        for j in range(i + 1, n):
            if rectangles[i] == rectangles[j]:
                model.Add(x[i] <= x[j])
                model.Add(y[i] <= y[j])

    # [2] For rectangles with the maximum width, restrict their horizontal domain.
    max_width_val = max(rect[0] for rect in rectangles)  # Assuming rect[0] is the width.
    for i, rect in enumerate(rectangles):
        wi, hi, _, _ = rect
        if wi == max_width_val:
            max_domain = (W - wi) // 2
            model.Add(x[i] <= max_domain)
    # Constraints max total profit
    # Not work.
    # model.maximize(sum(x * v for x, v in zip(xs, profits)))

    # CP Solver
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 600.0

    status = solver.Solve(model)
    num_vars = len(model.proto.variables)
    num_clauses = len(model.proto.constraints)
    
    if status == cp_model.OPTIMAL:
        selected_rectangles = []
        pos = []
        
        for i, xs_val in enumerate(xs):
            if solver.value(xs_val):
                selected_rectangles.append(rectangles[i])
                pos.append([solver.value(x[i]), solver.value(y[i])])
        profits = sum(rec[2] for rec in selected_rectangles)
        total_weight = sum(rec[3] for rec in selected_rectangles)
        return ["SAT", num_vars, num_clauses, profits, total_weight]
    else:
        return ["UNSAT", num_vars, num_clauses]


def max_profit_solution():
    folder_path = '../dataset/all_data_weight/'
    for file_name in os.listdir(folder_path):
        file_path = os.path.join(folder_path, file_name)
        if os.path.isfile(file_path) and file_name.endswith('.txt'):
            print(f"Processing file: {file_name}")
            with open(file_path, 'r') as file:
                lines = file.readlines()
                list_strip = list(map(int,lines[0].strip().split()))
                widths = list(map(int, lines[1].strip().split()))
                heights = list(map(int, lines[2].strip().split()))
                profits = list(map(int, lines[3].strip().split()))
                weights = list(map(int, lines[4].strip().split()))
                num_items = len(widths)

                # Get list rectangles from input data.
                rectangles = []
                prev_sat = []
                lower_bound = min(profits)
                upper_bound = sum(profits)
                for j in range (0, num_items):
                    item = (widths[j], heights[j], profits[j], weights[j])
                    rectangles.append(item)

                isExport = ""
                started_time = time.time()
                while (lower_bound + 1 < upper_bound):
                    if time.time() - started_time >= 600: 
                        isExport = "TIMEOUT"
                        if prev_sat == []:
                            export_csv(file_name, "cpsat_non_rotate", time.time() - started_time, "TIMEOUT", 0, 0, 0, 0)
                            break
                        export_csv(file_name, "cpsat_non_rotate", time.time() - started_time, "TIMEOUT", 0, 0, 0, 0)
                        break
                    mid = lower_bound + (upper_bound - lower_bound) // 2
                    status = CPSAT_constraints(rectangles, list_strip[0], list_strip[1], profits, mid, list_strip[2], weights)
                    if (status[0] == "UNSAT"):
                        upper_bound = mid
                        continue
                    elif (status[0] == "SAT"):
                        prev_sat = status
                        lower_bound = mid
                        continue
                ended_time = time.time()

                if isExport != "TIMEOUT":
                    if (prev_sat[0]) == "SAT":
                        export_csv(file_name, "cpsat_non_rotate", ended_time - started_time, "SAT", prev_sat[1], prev_sat[2], prev_sat[3], prev_sat[4])
                    else: export_csv(file_name, "cpsat_non_rotate", ended_time - started_time, "UNSAT", 0, 0, 0, 0)

max_profit_solution()