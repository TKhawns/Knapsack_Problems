from ortools.sat.python import cp_model
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import matplotlib.pyplot as plt

id_counter = 0

def display_solution(strip, rectangles, pos_circuits):
    # define Matplotlib figure and axis
    ax = plt.subplots()
    ax = plt.gca()
    plt.title(strip)

    if len(pos_circuits) > 0:
        for i in range(len(rectangles)):
            rect = plt.Rectangle(pos_circuits[i], *rectangles[i], edgecolor="#333")
            ax.add_patch(rect)

    ax.set_xlim(0, strip[0])
    ax.set_ylim(0, strip[1] + 1)
    ax.set_xticks(range(strip[0] + 1))
    ax.set_yticks(range(strip[1] + 1))
    ax.set_xlabel('width')
    ax.set_ylabel('height')
    # display plot
    plt.show()


def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path =  'output/'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)

    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

def solve_by_pysat(input, time, result, num_var, num_clause, name_lib):
    global id_counter
    id_counter += 1
    result_dict = {
        "ID": id_counter,
        "Problem": input,
        "Type": name_lib,
        "Time": time,
        "Result": result,
        "Variables": num_var,
        "Clauses": num_clause
    }
    write_to_xlsx(result_dict)

def cpSAT_OPP(rectangles, W, H, weights, profits, capacity):
    model = cp_model.CpModel()
    n = len(rectangles)
    
    # Variable encoding (x[i], y[i]) = (w[i], h[i])
    x = [model.NewIntVar(0, W, f'x_{i}') for i in range(n)]
    y = [model.NewIntVar(0, H, f'y_{i}') for i in range(n)]
    
    # Define xs and rotation variables
    xs = [model.NewBoolVar(f"a_{i}") for i in range(len(weights))]
    # Add rotation variables (True = rotated 90 degrees)
    rotated = [model.NewBoolVar(f"r_{i}") for i in range(n)]
    
    # Find maximum width and height among rectangles
    max_width = max(max(rect[0], rect[1]) for rect in rectangles)  # Modified to consider both dimensions
    max_height = max(max(rect[0], rect[1]) for rect in rectangles)  # Modified to consider both dimensions

    # (1) Domain of rectangle (x[i], y[i]) with rotation
    for i in range(n):
        wi, hi = rectangles[i]
        # If not rotated: width = wi, height = hi
        # If rotated: width = hi, height = wi
        model.Add(x[i] + wi <= W).OnlyEnforceIf(rotated[i].Not())
        model.Add(y[i] + hi <= H).OnlyEnforceIf(rotated[i].Not())
        model.Add(x[i] + hi <= W).OnlyEnforceIf(rotated[i])
        model.Add(y[i] + wi <= H).OnlyEnforceIf(rotated[i])

    # Add symmetry breaking constraints
    # 1. Domain reduction for maximum width rectangle
    max_width_idx = max(range(n), key=lambda i: max(rectangles[i][0], rectangles[i][1]))
    max_width_rect = max(rectangles[max_width_idx][0], rectangles[max_width_idx][1])
    model.Add(x[max_width_idx] <= (W - max_width_rect) // 2)

    # 2. Handle identical rectangles
    for i in range(n):
        for j in range(i + 1, n):
            if rectangles[i] == rectangles[j]:
                model.Add(x[i] <= x[j]).OnlyEnforceIf([xs[i], xs[j]])
                model.Add(y[i] <= y[j]).OnlyEnforceIf([xs[i], xs[j]])
    
    # 3. Large rectangles constraint
    for i in range(n):
        for j in range(i + 1, n):
            wi, hi = rectangles[i]
            wj, hj = rectangles[j]
            if min(wi + wj, wi + hj, hi + wj, hi + hj) > W:
                # These rectangles cannot be placed side by side in any rotation
                model.Add(y[i] + hi <= y[j]).OnlyEnforceIf([xs[i], xs[j], rotated[i].Not()])
                model.Add(y[i] + wi <= y[j]).OnlyEnforceIf([xs[i], xs[j], rotated[i]])
                model.Add(y[j] + hj <= y[i]).OnlyEnforceIf([xs[i], xs[j], rotated[j].Not()])
                model.Add(y[j] + wj <= y[i]).OnlyEnforceIf([xs[i], xs[j], rotated[j]])

    #(2) Non-overlapping constraints with rotation
    for i in range(n):
        for j in range(i + 1, n):
            wi, hi = rectangles[i]
            wj, hj = rectangles[j]

            # Boolean variables
            no_overlap_1 = model.NewBoolVar(f'nonoverlap_1_{i}_{j}')
            no_overlap_2 = model.NewBoolVar(f'nonoverlap_2_{i}_{j}')
            no_overlap_3 = model.NewBoolVar(f'nonoverlap_3_{i}_{j}')
            no_overlap_4 = model.NewBoolVar(f'nonoverlap_4_{i}_{j}')

            # For rectangle i
            wi_expr = model.NewIntVar(0, max(wi, hi), f'wi_expr_{i}')
            hi_expr = model.NewIntVar(0, max(wi, hi), f'hi_expr_{i}')
            model.Add(wi_expr == wi).OnlyEnforceIf(rotated[i].Not())
            model.Add(hi_expr == hi).OnlyEnforceIf(rotated[i].Not())
            model.Add(wi_expr == hi).OnlyEnforceIf(rotated[i])
            model.Add(hi_expr == wi).OnlyEnforceIf(rotated[i])

            # For rectangle j
            wj_expr = model.NewIntVar(0, max(wj, hj), f'wj_expr_{j}')
            hj_expr = model.NewIntVar(0, max(wj, hj), f'hj_expr_{j}')
            model.Add(wj_expr == wj).OnlyEnforceIf(rotated[j].Not())
            model.Add(hj_expr == hj).OnlyEnforceIf(rotated[j].Not())
            model.Add(wj_expr == hj).OnlyEnforceIf(rotated[j])
            model.Add(hj_expr == wj).OnlyEnforceIf(rotated[j])

            # Non-overlapping constraints with rotation
            model.Add(x[i] + wi_expr <= x[j]).OnlyEnforceIf(no_overlap_1)
            model.Add(x[j] + wj_expr <= x[i]).OnlyEnforceIf(no_overlap_2)
            model.Add(y[i] + hi_expr <= y[j]).OnlyEnforceIf(no_overlap_3)
            model.Add(y[j] + hj_expr <= y[i]).OnlyEnforceIf(no_overlap_4)

            model.AddBoolOr([no_overlap_1, no_overlap_2, no_overlap_3, no_overlap_4])

    # Constraint about Weight
    model.Add(sum(x * w for x, w in zip(xs, weights)) <= capacity)

    # Constraints max total profit
    model.maximize(sum(x * v for x, v in zip(xs, profits)))

    # CP Solver
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 100.0
    status = solver.Solve(model)
    elapse_time = solver.user_time
    isSat = ""
    
    if (elapse_time > 100):
        print("timeout")
        return
    
    if status == cp_model.OPTIMAL:
        isSat = "sat"
        total = 0
        selected_rectangles = []
        pos = []
        
        for i, xs_val in enumerate(xs):
            if solver.Value(xs_val):
                total += profits[i]
                # If rotated, swap width and height
                if solver.Value(rotated[i]):
                    selected_rectangles.append((rectangles[i][1], rectangles[i][0]))
                else:
                    selected_rectangles.append(rectangles[i])
                pos.append([solver.Value(x[i]), solver.Value(y[i])])
                
        print("Total profit:", total)
        solve_by_pysat(len(weights), elapse_time, isSat, 0, 0, "CPSAT")
        
        # Visualize solution
        strip = (W, H)
        display_solution(strip, selected_rectangles, pos)
    else:
        print('No solution found.')
        isSat = "unsat"
        solve_by_pysat(len(weights), elapse_time, isSat, 0, 0, "CPSAT")


def solve_KPWL_CPSAT():
    with open('../dataset/test_input.txt', 'r') as file:
        # max_weight, min_profit
        # strip
        # list weights
        # list profits
        # width_i
        # height_i
        lines = file.readlines()
        len_file = len(lines)
        for i in range(0, len_file,6):
            bounds = list(map(int,lines[i].strip().split()))
            arr_strip = list(map(int,lines[i+1].strip().split()))
            weights = list(map(int, lines[i + 2].strip().split()))
            profits = list(map(int, lines[i + 3].strip().split()))
            arr_width =  list(map(int, lines[i + 4].strip().split()))
            arr_height =  list(map(int, lines[i + 5].strip().split()))
            rectangles = []
            
            l = len(arr_width)
            for j in range(0, l):
                rectangles.append((arr_width[j], arr_height[j]))
            weight_bound = bounds[0]
            
            cpSAT_OPP(rectangles, arr_strip[0], arr_strip[1], weights, profits, weight_bound)

solve_KPWL_CPSAT()