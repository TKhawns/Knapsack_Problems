import os
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from ortools.linear_solver import pywraplp
import matplotlib.pyplot as plt
import time

id_counter = 0

def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path =  'output/'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)

    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

def solve_by_pysat(input, time, result, num_var, num_clause, name_lib):
    global id_counter
    id_counter += 1
    result_dict = {
        "ID": id_counter,
        "Problem": input,
        "Type": name_lib,
        "Time": time,
        "Result": result,
        "Variables": num_var,
        "Clauses": num_clause
    }
    write_to_xlsx(result_dict)

def find_all_solution(weights, profits, capacity, rectangles):
    # MIP SOLVER
    solver = pywraplp.Solver.CreateSolver('SAT')

    n = len(weights)
    xs = [solver.BoolVar(f'x_{i}') for i in range(n)]

    # (1) Weight constraints
    solver.Add(sum(xs[i] * weights[i] for i in range(n)) <= capacity)

    # (2) Maximize profits.
    solver.Maximize(sum(xs[i] * profits[i] for i in range(n)))

    solutions = []
    start_time = time.time()
    while True:
        status = solver.Solve()
        elapse_time = time.time() - start_time

        if elapse_time >= 100 or status != pywraplp.Solver.OPTIMAL:
            print(elapse_time)
            break 
        
        # Current solution
        current_solution = [int(xs[i].solution_value()) for i in range(n)]
        total_profit = sum(profits[i] * current_solution[i] for i in range(n))

        selected_rectangles = [rectangles[i] for i in range(n) if current_solution[i] == 1]

        # New weights list.
        new_weights = [weights[i] for i in range(n) if int(xs[i].solution_value()) != 0 ]
        solutions.append((selected_rectangles, total_profit, new_weights))

        solver.Add(solver.Sum((1 - xs[i]) if current_solution[i] == 1 else xs[i] for i in range(n)) >= 1)

    solutions.sort(reverse=True, key=lambda x: x[1])

    # In ra các nghiệm đã sắp xếp
    # for selected_rectangles, profit in solutions:
    #     print(f"Selected Rectangles: {selected_rectangles}, Profit: {profit}")
    return [solutions, elapse_time]


def MIP_OPP(rectangles, W, H):
    strip = (W, H)

    # MIP Solver
    solver = pywraplp.Solver.CreateSolver('SAT')
    if not solver:
        print("Solver not found.")
        return

    n = len(rectangles)

    # Define variables for positions of each rectangle (x[i], y[i])
    x = [solver.IntVar(0, W, f'x_{i}') for i in range(n)]
    y = [solver.IntVar(0, H, f'y_{i}') for i in range(n)]
    is_rotated = [solver.BoolVar(f'rot_{i}') for i in range(n)]


    # (1) Domain constraints for rectangles within the bin dimensions
    for i in range(n):
        wi, hi = rectangles[i]
        can_rotate = wi <= H and hi <= W

        solver.Add(x[i] + wi <= W)  # Ensure rectangle width fits within W
        solver.Add(y[i] + hi <= H)  # Ensure rectangle height fits within H

        if can_rotate:
            # Width constraint with rotation
            solver.Add(x[i] + wi <= W + (hi - wi) * is_rotated[i])
            # Height constraint with rotation
            solver.Add(y[i] + hi <= H + (wi - hi) * is_rotated[i])
        else:
            # Original constraints if rotation not possible
            solver.Add(x[i] + wi <= W)
            solver.Add(y[i] + hi <= H)

    # Modified symmetry reduction constraints
    # (2) Symmetry reduction constraints
    max_width = max(rect[0] for rect in rectangles)
    max_height = max(rect[1] for rect in rectangles)
    for i in range(n):
        wi, hi = rectangles[i]
        effective_width = solver.IntVar(0, max(W, H), f'eff_w_{i}')
        effective_height = solver.IntVar(0, max(W, H), f'eff_h_{i}')
        
        # Set effective dimensions based on rotation
        solver.Add(effective_width == wi + (hi - wi) * is_rotated[i])
        solver.Add(effective_height == hi + (wi - hi) * is_rotated[i])

        # Domain reduction using horizontal/vertical symmetry
        if wi == max_width or hi == max_width:
            solver.Add(2 * x[i] <= W - effective_width)
        if hi == max_height or wi == max_height:
            solver.Add(2 * y[i] <= H - effective_height)

    # (4) Same rectangle symmetry
    for i in range(n):
        for j in range(i + 1, n):
            if rectangles[i] == rectangles[j]:
                # Fix positional relationship to reduce symmetry
                solver.Add(x[i] <= x[j])
                solver.Add(y[i] <= y[j])

    
    # Modified non-overlapping constraints
    for i in range(n):
        for j in range(i + 1, n):
            wi, hi = rectangles[i]
            wj, hj = rectangles[j]

            # Effective dimensions for rectangle i
            eff_wi = wi + (hi - wi) * is_rotated[i]
            eff_hi = hi + (wi - hi) * is_rotated[i]
            
            # Effective dimensions for rectangle j
            eff_wj = wj + (hj - wj) * is_rotated[j]
            eff_hj = hj + (wj - hj) * is_rotated[j]

            # Binary variables for overlap conditions
            no_overlap_1 = solver.BoolVar(f'no_overlap_1_{i}_{j}')
            no_overlap_2 = solver.BoolVar(f'no_overlap_2_{i}_{j}')
            no_overlap_3 = solver.BoolVar(f'no_overlap_3_{i}_{j}')
            no_overlap_4 = solver.BoolVar(f'no_overlap_4_{i}_{j}')

            # Check if rectangles can't fit horizontally or vertically in any rotation
            max_width_i = max(wi, hi)
            max_width_j = max(wj, hj)
            if max_width_i + max_width_j > W:
                # Force vertical arrangement
                solver.Add(y[i] + eff_hi <= y[j] + H * no_overlap_3)
                solver.Add(y[j] + eff_hj <= y[i] + H * no_overlap_4)

            max_height_i = max(wi, hi)
            max_height_j = max(wj, hj)
            if max_height_i + max_height_j > H:
                # Force horizontal arrangement
                solver.Add(x[i] + eff_wi <= x[j] + W * no_overlap_1)
                solver.Add(x[j] + eff_wj <= x[i] + W * no_overlap_2)

            # Non-overlapping constraints with rotation
            solver.Add(x[i] + eff_wi <= x[j] + W * (1 - no_overlap_1))
            solver.Add(x[j] + eff_wj <= x[i] + W * (1 - no_overlap_2))
            solver.Add(y[i] + eff_hi <= y[j] + H * (1 - no_overlap_3))
            solver.Add(y[j] + eff_hj <= y[i] + H * (1 - no_overlap_4))

            # Ensure at least one no-overlap condition holds
            solver.Add(no_overlap_1 + no_overlap_2 + no_overlap_3 + no_overlap_4 >= 1)

    # Solve the model
    status = solver.Solve()
    print(status)
    # postition of result rectangles, input of visualize.
    pos = []
    print("POS: ", pos)
    selected_rectangles = []

    if status == pywraplp.Solver.OPTIMAL:
        print("Optimal solution found.")
        for i in range(n):
                wi, hi = rectangles[i]
                if x[i].solution_value() >= 0 and y[i].solution_value() >= 0:
                    pos.append([x[i].solution_value(), y[i].solution_value()])

                print(f"Rectangle: ({wi}, {hi})", f"Position (x, y) = ({x[i].solution_value()}, {y[i].solution_value()})"),  
                selected_rectangles.append(rectangles[i])
            
        print("New pos:", pos)
        print("Result rectangles: ", selected_rectangles)
        # return ["sat", solver.wall_time()/1000]
        display_solution(strip, selected_rectangles, pos)

    else:
        print("No optimal solution found.")
        return ["unsat",solver.wall_time() / 1000]  # Convert to seconds



def solve_KPWL_CPSAT():
    with open('../dataset/test_input.txt', 'r') as file:
        # max_weight, min_profit
        # strip
        # list weights
        # list profits
        # width_i
        # height_i
        lines = file.readlines()
        len_file = len(lines)
        for i in range(0, len_file,6):
            bounds = list(map(int,lines[i].strip().split()))
            arr_strip = list(map(int,lines[i+1].strip().split()))
            weights = list(map(int, lines[i + 2].strip().split()))
            profits = list(map(int, lines[i + 3].strip().split()))
            arr_width =  list(map(int, lines[i + 4].strip().split()))
            arr_height =  list(map(int, lines[i + 5].strip().split()))
            rectangles = []
            
            l = len(arr_width)
            for j in range(0, l):
                rectangles.append((arr_width[j], arr_height[j]))

            result = find_all_solution(weights, profits, bounds[0], rectangles)
            solutions = result[0]
            elapse_time = result[1]

            if (elapse_time >= 100):
                solve_by_pysat(len(weights), "timeout", "unsat 2CT", 0, 0, "MIP")
                print("timeout")
                continue

            for item in solutions:
                res = MIP_OPP(item[0], arr_strip[0], arr_strip[1])
                elapse_time += res[1]
                if (res[0] == "sat"):
                    solve_by_pysat(len(weights), elapse_time, "sat", 0, 0, "MIP")
                    break
                elif (res[0] != "unsat"):
                    solve_by_pysat(len(weights), "timeout", "unsat opp", 0, 0, "MIP")
                    continue
                if (elapse_time >= 100):
                    solve_by_pysat(len(weights), "timeout", "unsat", 0, 0, "MIP")
                    break

                print(item)
            # MIP_OPP(rectangles, arr_strip[0], arr_strip[1], weights, profits, weight_bound)

def display_solution(strip, rectangles, pos_circuits):
    # define Matplotlib figure and axis
    fig, ax = plt.subplots()
    ax = plt.gca()
    plt.title(strip)

    if len(pos_circuits) > 0:
        for i in range(len(rectangles)):
            rect = plt.Rectangle(pos_circuits[i], *rectangles[i], edgecolor="#333")
            ax.add_patch(rect)

    ax.set_xlim(0, strip[0])
    ax.set_ylim(0, strip[1] + 1)
    ax.set_xticks(range(strip[0] + 1))
    ax.set_yticks(range(strip[1] + 1))
    ax.set_xlabel('width')
    ax.set_ylabel('height')
    # display plot
    plt.show()

solve_KPWL_CPSAT()