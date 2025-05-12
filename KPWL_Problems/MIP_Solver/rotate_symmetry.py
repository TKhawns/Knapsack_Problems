import os
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from ortools.linear_solver import pywraplp
import matplotlib.pyplot as plt
import time

id_counter = 0

def display_solution(strip, rectangles, pos_circuits):
    ax = plt.subplots()
    ax = plt.gca()
    plt.title(strip)

    if len(pos_circuits) > 0:
        for i in range(len(rectangles)):
            rect = plt.Rectangle(pos_circuits[i], *rectangles[i], edgecolor="#333")
            ax.add_patch(rect)

    ax.set_xlim(0, strip[0])
    ax.set_ylim(0, strip[1] + 1)
    ax.set_xticks(range(strip[0] + 1))
    ax.set_yticks(range(strip[1] + 1))
    ax.set_xlabel('width')
    ax.set_ylabel('height')
    plt.show()

def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path =  'out_rotate/'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)

    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

def export_csv(problem_name, method_name, time, result, num_var, num_clause, max_profit, weights):
    global id_counter
    id_counter += 1
    result_dict = {
        "No": id_counter,
        "Problem": problem_name,
        "Type": method_name,
        "Time": time,
        "Result": result,
        "Variables": num_var,
        "Clauses": num_clause,
        "Max profit": max_profit,
        "Weights": weights
    }
    write_to_xlsx(result_dict)


def mip_constraints(rectangles, W, H, profits, mid, C, weights):
    # Create the SCIP solver.
    solver = pywraplp.Solver.CreateSolver('SCIP')

    n = len(rectangles)
    
    # Decision variables:
    # x[i], y[i] – lower‐left coordinates (integers)
    # r[i] – rotation binary variable (0: no rotation, 1: rotated)
    # xs[i] – selection binary variable (1 if rectangle i is used)
    x = [solver.IntVar(0, W, f'x_{i}') for i in range(n)]
    y = [solver.IntVar(0, H, f'y_{i}') for i in range(n)]
    r = [solver.BoolVar(f'r_{i}') for i in range(n)]
    xs = [solver.BoolVar(f'xs_{i}') for i in range(n)]
    
    # For each rectangle, we define the “effective” width and height as a linear expression.
    # If not rotated (r[i]==0): width = wi, height = hi.
    # If rotated    (r[i]==1): width = hi, height = wi.
    # We write: width = wi + (hi - wi)*r[i], and similarly for height.
    for i in range(n):
        wi, hi, _, _ = rectangles[i]
        width_expr  = wi + (hi - wi) * r[i]
        height_expr = hi + (wi - hi) * r[i]
        
        # Big-M values for boundary constraints.
        # When xs[i]==0, the constraint is relaxed.
        big_M_x = W + max(wi, hi)
        big_M_y = H + max(wi, hi)
        
        # If rectangle i is selected then it must lie completely within the bin:
        # x[i] + width_expr <= W   (when xs[i]==1)
        # We enforce this via: x[i] + width_expr <= W + big_M*(1 - xs[i])
        solver.Add(x[i] + width_expr <= W + big_M_x * (1 - xs[i]))
        solver.Add(y[i] + height_expr <= H + big_M_y * (1 - xs[i]))
        
        # Enforce rotation rules:
        # (a) If the rectangle in its original orientation is out‐of‐bin, force rotation.
        if wi > W or hi > H:
            solver.Add(r[i] == 1)
        # (b) If it is a square or “non‐rotatable” (e.g. out‐of‐bin when rotated in the other way),
        # then force r[i] to 0.
        if wi == hi or (wi > H or hi > W):
            solver.Add(r[i] == 0)
    
    # For non-overlap, if both rectangles i and j are selected, then one of the following must hold:
    #   [1] Rectangle i is to the left of j,
    #   [2] Rectangle j is to the left of i,
    #   [3] Rectangle i is below j, or
    #   [4] Rectangle j is below i.
    # We introduce four binary variables b_{i,j,k} (k=1,2,3,4) to select one of these.
    # Then, using big-M constraints, we write for instance:
    #   x[i] + width_i <= x[j] + M*(1 - b_{i,j,1})
    # and finally require that if both are selected, then:
    #   b_{i,j,1} + b_{i,j,2} + b_{i,j,3} + b_{i,j,4} >= xs[i] + xs[j] - 1.
    #
    # We choose a global big-M for horizontal and vertical separation.
    global_big_M_x = W + max(max(rect[0], rect[1]) for rect in rectangles)
    global_big_M_y = H + max(max(rect[0], rect[1]) for rect in rectangles)
    
    # Create and add non-overlap constraints for each pair (i,j)
    for i in range(n):
        for j in range(i+1, n):
            wi, hi, _, _ = rectangles[i]
            wj, hj, _, _ = rectangles[j]
            width_i  = wi + (hi - wi) * r[i]
            height_i = hi + (wi - hi) * r[i]
            width_j  = wj + (hj - wj) * r[j]
            height_j = hj + (wj - hj) * r[j]
            
            # Create four binary variables for the four disjuncts.
            b1 = solver.BoolVar(f'b_{i}_{j}_1')
            b2 = solver.BoolVar(f'b_{i}_{j}_2')
            b3 = solver.BoolVar(f'b_{i}_{j}_3')
            b4 = solver.BoolVar(f'b_{i}_{j}_4')
            
            # [1] i is to the left of j:
            solver.Add(x[i] + width_i <= x[j] + global_big_M_x * (1 - b1))
            # [2] j is to the left of i:
            solver.Add(x[j] + width_j <= x[i] + global_big_M_x * (1 - b2))
            # [3] i is below j:
            solver.Add(y[i] + height_i <= y[j] + global_big_M_y * (1 - b3))
            # [4] j is below i:
            solver.Add(y[j] + height_j <= y[i] + global_big_M_y * (1 - b4))
            
            # If both rectangles are selected (xs[i] = xs[j] = 1) then at least one of the four
            # disjuncts must hold.
            solver.Add(b1 + b2 + b3 + b4 >= xs[i] + xs[j] - 1)
    
    # Total profit constraint: selected rectangles must achieve profit at least mid.
    solver.Add(solver.Sum([xs[i] * profits[i] for i in range(n)]) >= mid)
    solver.Add(solver.Sum([xs[i] * weights[i] for i in range(n)]) <= C)
    
    # --- Symmetry constraints ---
    # [1] If two rectangles are identical, fix a consistent ordering.
    for i in range(n):
        for j in range(i+1, n):
            if rectangles[i] == rectangles[j]:
                solver.Add(x[i] <= x[j])
                solver.Add(y[i] <= y[j])
    # [2] Domain symmetry: restrict the horizontal domain for some rectangles.
    # (Using the second element of the tuple to compute max_width as in your original code.)
    max_width = max(rect[1] for rect in rectangles)
    for i, (wi, hi, _, _) in enumerate(rectangles):
        if wi == max_width:
            solver.Add(x[i] <= (W - wi) // 2)
    
    # Objective: maximize total profit.
    objective = solver.Sum([xs[i] * profits[i] for i in range(n)])
    solver.Maximize(objective)
    
    # Set a time limit (SCIP expects milliseconds, so 1200 sec = 1,200,000 ms)
    solver.set_time_limit(600 * 1000)
    
    status = solver.Solve()
    
    if status == pywraplp.Solver.OPTIMAL or status == pywraplp.Solver.FEASIBLE:
        selected_rectangles = []
        positions = []
        for i in range(n):
            # Use a threshold (0.5) to interpret the binary decision variable.
            if xs[i].solution_value() > 0.5:
                selected_rectangles.append(rectangles[i])
                positions.append((x[i].solution_value(), y[i].solution_value()))
        num_var = solver.NumVariables()
        num_constr = solver.NumConstraints()
        total_profit = sum(rect[2] for rect in selected_rectangles)
        total_weight = sum(rect[3] for rect in selected_rectangles)
        return ['SAT', total_profit, num_var, num_constr, total_weight]
    else:
        return ["UNSAT"]
def max_profit_solution():
    folder_path = '../dataset/soft'
    for file_name in os.listdir(folder_path):
        file_path = os.path.join(folder_path, file_name)
        if os.path.isfile(file_path) and file_name.endswith('.txt'):
            print(f"Processing file: {file_name}")
            with open(file_path, 'r') as file:
                lines = file.readlines()
                list_strip = list(map(int,lines[0].strip().split()))
                widths = list(map(int, lines[1].strip().split()))
                heights = list(map(int, lines[2].strip().split()))
                profits = list(map(int, lines[3].strip().split()))
                weights = list(map(int, lines[4].strip().split()))
                num_items = len(widths)

                # Get list rectangles from input data.
                rectangles = []
                prev_sat = []
                lower_bound = min(profits)
                upper_bound = sum(profits)
                for j in range (0, num_items):
                    item = (widths[j], heights[j], profits[j],weights[j])
                    rectangles.append(item)

                isExport = ""
                started_time = time.time()
                while (lower_bound + 1 < upper_bound):
                    if (time.time() - started_time >= 600): 
                        isExport = "TIMEOUT"
                        if prev_sat == []:
                            export_csv(file_name, "mip_rotate_symmetry", time.time() - started_time, "TIMEOUT", 0, 0, 0, 0)
                            break
                        export_csv(file_name, "mip_rotate_symmetry", time.time() - started_time, "TIMEOUT", prev_sat[1], prev_sat[2], prev_sat[3], prev_sat[4])
                        break
                    mid = lower_bound + (upper_bound - lower_bound) // 2
                    status = mip_constraints(rectangles, list_strip[0], list_strip[1], profits, mid, list_strip[2], weights)
                    if (status[0] == "UNSAT"):
                        upper_bound = mid
                        continue
                    elif (status[0] == "SAT"):
                        prev_sat = status
                        lower_bound = mid
                        continue
                ended_time = time.time()

                if isExport != "TIMEOUT":
                    if (prev_sat != [] and prev_sat[0]) == "SAT":
                        export_csv(file_name, "mip_rotate_symmetry", ended_time - started_time, "SAT", prev_sat[1], prev_sat[2], prev_sat[3], prev_sat[4])
                    else: export_csv(file_name, "mip_rotate_symmetry", ended_time - started_time, "UNSAT", 0, 0, 0, 0)

max_profit_solution()