import os
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import matplotlib.pyplot as plt
import time
from docplex.cp.model import CpoModel

# Counter STT to export csv file.
id_counter = 0
# Display solution.
def display_solution(strip, rectangles, pos_circuits):
    ax = plt.subplots()
    ax = plt.gca()
    plt.title(strip)

    if len(pos_circuits) > 0:
        for i in range(len(rectangles)):
            rect = plt.Rectangle(pos_circuits[i], *rectangles[i], edgecolor="#333")
            ax.add_patch(rect)

    ax.set_xlim(0, strip[0])
    ax.set_ylim(0, strip[1] + 1)
    ax.set_xticks(range(strip[0] + 1))
    ax.set_yticks(range(strip[1] + 1))
    ax.set_xlabel('width')
    ax.set_ylabel('height')
    plt.show()
# Export file to csv file.
def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path =  'output/'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)

    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)
def export_csv(problem_name, method_name, time, result, num_var, num_clause, max_profit):
    global id_counter
    id_counter += 1
    result_dict = {
        "No": id_counter,
        "Problem": problem_name,
        "Type": method_name,
        "Time": time,
        "Result": result,
        "Variables": num_var,
        "Clauses": num_clause,
        "Max profit": max_profit
    }
    write_to_xlsx(result_dict)

def cplex_constraints(rectangles, W, H, profits, mid):
    # Create the CPO model
    model = CpoModel()
    n = len(rectangles)

    # Variables for position and rotation
    x = [model.integer_var(0, W, name=f'x_{i}') for i in range(n)]
    y = [model.integer_var(0, H, name=f'y_{i}') for i in range(n)]
    r = [model.binary_var(name=f'r_{i}') for i in range(n)]  # Rotation variable
    xs = [model.binary_var(name=f's_{i}') for i in range(n)]  # Selection variable

    # Constraints for each rectangle
    for i in range(n):
        wi, hi, _ = rectangles[i]

        # Width and height constraints based on rotation
        model.add(xs[i] * (x[i] + (1 - r[i]) * wi + r[i] * hi) <= W)
        model.add(xs[i] * (y[i] + (1 - r[i]) * hi + r[i] * wi) <= H)

        # Enforce rotation if out of bounds
        if wi > W or hi > H:
            model.add(r[i] == 1)
        # Prevent rotation for squares or non-rotatable rectangles
        if wi == hi or (wi > H or hi > W):
            model.add(r[i] == 0)

    # Non-overlapping constraints
    for i in range(n):
        for j in range(i + 1, n):
            wi, hi, _ = rectangles[i]
            wj, hj, _ = rectangles[j]

            # Non-overlapping constraints with rotation
            cond1 = (x[i] + (1 - r[i]) * wi + r[i] * hi <= x[j])
            cond2 = (x[j] + (1 - r[j]) * wj + r[j] * hj <= x[i])
            cond3 = (y[i] + (1 - r[i]) * hi + r[i] * wi <= y[j])
            cond4 = (y[j] + (1 - r[j]) * hj + r[j] * wj <= y[i])

            # If either rectangle is not selected, or one of the non-overlap conditions holds.
            model.add(model.logical_or([
            xs[i] == 0,
            xs[j] == 0,
            cond1,
            cond2,
            cond3,
            cond4
            ]))

    model.add(sum(xs[i] * profits[i] for i in range(n)) >= mid)

    # # Maximize total profit
    model.add(model.maximize(sum(xs[i] * profits[i] for i in range(n))))

    # Solve the model
    solution = model.solve(TimeLimit=600)

    if solution:
        selected_rectangles = []
        positions = []

        for i in range(n):
                if solution.get_value(xs[i]) == 1:
                    selected_rectangles.append(rectangles[i])
                    positions.append((solution.get_value(x[i]), solution.get_value(y[i])))
        num_var, num_clause = len(model.get_all_variables()), len(model.get_all_expressions())
        profits = sum(rec[2] for rec in selected_rectangles)
        return ['SAT', selected_rectangles, profits, num_var, num_clause]
    else:
        return ["UNSAT"]

def max_profit_solution():
    folder_path = '../../dataset/soft/'
    for file_name in os.listdir(folder_path):
        file_path = os.path.join(folder_path, file_name)
        if os.path.isfile(file_path) and file_name.endswith('.txt') and file_name == "okp1.txt":
            print(f"Processing file: {file_name}")
            with open(file_path, 'r') as file:
                lines = file.readlines()
                list_strip = list(map(int,lines[0].strip().split()))
                widths = list(map(int, lines[1].strip().split()))
                heights = list(map(int, lines[2].strip().split()))
                profits = list(map(int, lines[3].strip().split()))
                num_items = len(widths)

                # Get list rectangles from input data.
                rectangles = []
                prev_sat = []
                lower_bound = min(profits)
                upper_bound = sum(profits)
                for j in range (0, num_items):
                    item = (widths[j], heights[j], profits[j])
                    rectangles.append(item)

                started_time = time.time()
                isExport = ""
                while (lower_bound + 1 < upper_bound):
                    if (time.time() - started_time >= 600):
                        isExport = "TIMEOUT"
                        if (prev_sat != []):
                            export_csv(file_name, "cplex_non_symmetry_studio", time.time() - started_time, "TIMEOUT", prev_sat[3], prev_sat[4], prev_sat[2])
                            break
                        export_csv(file_name, "cplex_non_symmetry_studio", time.time() - started_time, "TIMEOUT", 0, 0, 0)
                        break
                    mid = lower_bound + (upper_bound - lower_bound) // 2
                    status = cplex_constraints(rectangles, list_strip[0], list_strip[1], profits, mid)
                    if (status[0] == "UNSAT"):
                        upper_bound = mid
                        continue
                    elif (status[0] == "SAT"):
                        prev_sat = status
                        lower_bound = mid
                        continue
                ended_time = time.time()

                if (isExport != "TIMEOUT"):
                    if (prev_sat[0]) == "SAT":
                        export_csv(file_name, "cplex_non_symmetry_studio", ended_time - started_time, "SAT", prev_sat[3], prev_sat[4], prev_sat[2])
                    else: export_csv(file_name, "cplex_non_symmetry_studio", ended_time - started_time, "UNSAT", 0, 0, 0)

max_profit_solution()