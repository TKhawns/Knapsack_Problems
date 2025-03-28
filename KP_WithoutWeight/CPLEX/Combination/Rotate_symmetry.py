import os
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import matplotlib.pyplot as plt
from itertools import combinations
import time
from docplex.cp.model import CpoModel


id_counter = 0

def display_solution(strip, rectangles, pos_circuits):
    ax = plt.subplots()
    ax = plt.gca()
    plt.title(strip)

    if len(pos_circuits) > 0:
        for i in range(len(rectangles)):
            rect = plt.Rectangle(pos_circuits[i], *rectangles[i], edgecolor="#333")
            ax.add_patch(rect)

    ax.set_xlim(0, strip[0])
    ax.set_ylim(0, strip[1] + 1)
    ax.set_xticks(range(strip[0] + 1))
    ax.set_yticks(range(strip[1] + 1))
    ax.set_xlabel('width')
    ax.set_ylabel('height')
    plt.show()

def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path =  'output/'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)

    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

def solve_by_pysat(input, time, result, num_var, num_clause, name_lib):
    global id_counter
    id_counter += 1
    result_dict = {
        "ID": id_counter,
        "Problem": input,
        "Type": name_lib,
        "Time": time,
        "Result": result,
        "Variables": num_var,
        "Clauses": num_clause
    }
    write_to_xlsx(result_dict)

def findListSumOfProfits(rectangles, numberOfItems):
    # parameter rectangles: include profit [(width, height, profit)...].
    # Output is [[listOfRects, profitOfThisListRects], ...[]]
    result_for_loop = []
    started_time = time.time()
    # Timeout is 60s.
    while numberOfItems > 0:
        if time.time() - started_time >= 120:
            return []
        comb = combinations(rectangles, numberOfItems)
        for var in comb:
            temp = []
            rect_in_comb = [(x[0], x[1]) for x in list(var)]
            list_profit = [(x[2]) for x in list(var)]
            profit = sum(x[2] for x in list(var))
            temp.append(rect_in_comb)
            temp.append(list_profit)
            temp.append(profit)
            result_for_loop.append(temp)

        numberOfItems -= 1
    result_for_loop = sorted(result_for_loop, reverse=True, key=lambda x: x[2])
    return result_for_loop

def cpSAT_OPP(rectangles, W, H, profits):
    # Create the CPO model
    model = CpoModel()
    n = len(rectangles)

    # Variables for position and rotation
    x = [model.integer_var(0, W, name=f'x_{i}') for i in range(n)]
    y = [model.integer_var(0, H, name=f'y_{i}') for i in range(n)]
    r = [model.binary_var(name=f'r_{i}') for i in range(n)]  # Rotation variable
    xs = [model.binary_var(name=f's_{i}') for i in range(n)]  # Selection variable

    # Constraints for each rectangle
    for i in range(n):
        wi, hi = rectangles[i]

        # Width and height constraints based on rotation
        model.add(x[i] + (1 - r[i]) * wi + r[i] * hi <= W)
        model.add(y[i] + (1 - r[i]) * hi + r[i] * wi <= H)

        # Enforce rotation if out of bounds
        if wi > W or hi > H:
            model.add(r[i] == 1)
        # Prevent rotation for squares or non-rotatable rectangles
        if wi == hi or (wi > H or hi > W):
            model.add(r[i] == 0)

    # Non-overlapping constraints
    for i in range(n):
        for j in range(i + 1, n):
            wi, hi = rectangles[i]
            wj, hj = rectangles[j]

            # Non-overlapping constraints with rotation
            no_overlap_1 = (x[i] + (1 - r[i]) * wi + r[i] * hi <= x[j])
            no_overlap_2 = (x[j] + (1 - r[j]) * wj + r[j] * hj <= x[i])
            no_overlap_3 = (y[i] + (1 - r[i]) * hi + r[i] * wi <= y[j])
            no_overlap_4 = (y[j] + (1 - r[j]) * hj + r[j] * wj <= y[i])

            # Ensure at least one non-overlapping condition
            model.add(no_overlap_1 | no_overlap_2 | no_overlap_3 | no_overlap_4)
    
    # # Same rectangles:
    # for i in range(n):
    #     for j in range(i + 1, n):
    #         if rectangles[i] == rectangles[j]:
    #             # Fix positional relationship
    #             model.add(x[i] <= x[j])
    #             model.add(y[i] <= y[j])
    
    max_width = max(rect[1] for rect in rectangles)

    for i, (wi, hi) in enumerate(rectangles):
        if wi == max_width:
            # Restrict horizontal domain
            max_domain = (W - wi) // 2
            model.add(x[i] <= max_domain)


    # Maximize total profit
    model.add(model.maximize(sum(xs[i] * profits[i] for i in range(n))))

    # Solve the model
    solution = model.solve(TimeLimit=240)

    if solution:
        selected_rectangles = []
        positions = []

        for i in range(n):
                if solution.get_value(xs[i]) == 1:
                    selected_rectangles.append(rectangles[i])
                    positions.append((solution.get_value(x[i]), solution.get_value(y[i])))
        print(model.get_all_variables())
        return ["SAT", len(model.get_all_variables()), len(model.get_all_expressions())]
    else:
        return ["UNSAT", 0, [], []]


def max_profit_solution():
    with open('../dataset/dataset8.txt', 'r') as file:
        lines = file.readlines()

        list_strip = list(map(int,lines[0].strip().split()))
        widths = list(map(int, lines[1].strip().split()))
        heights = list(map(int, lines[2].strip().split()))
        profits = list(map(int, lines[3].strip().split()))
        num_items = len(widths)
        is_sol = ""

        # Get list rectangles from input data.
        rectangles = []
        for j in range (0, num_items):
            item = (widths[j], heights[j], profits[j])
            rectangles.append(item)
        
        result_vars = 0
        result_clauses = 0
        started_time = time.time()
        # Update code
        # input_list sample = [ [ [(1, 3), (2, 1), (2, 1), (1, 1), (1, 2), (1, 1), (1, 1), (6, 2), (2, 1), (1, 1)], 33], ...[] ]
        input_list = findListSumOfProfits(rectangles, num_items)
        if input_list == []:
            print("UNSAT because combination timeout")
            is_sol = "UNSAT"
            solve_by_pysat(num_items, "TIMEOUT_COMB", is_sol, result_vars, result_clauses, "CPLEX")
            return
        for input in input_list:
            print("Start solving after combination:")
            status = cpSAT_OPP(input[0], list_strip[0], list_strip[1], input[1])
            if status[0] == "SAT":
                    is_sol = "SAT"
                    result_vars = status[1]
                    result_clauses = status[2]
                    print("Max profit is: ", input[2])
                    solve_by_pysat(num_items, time.time() - started_time, is_sol, result_vars, result_clauses, "CPLEX")        
                    break

            elif status[0] == "UNSAT":
                is_sol = "UNSAT"
                print("\n", "UNSAT this loop because OPP constraints. ")
                continue
        if (is_sol == "UNSAT"):
            solve_by_pysat(num_items, time.time() - started_time, is_sol, 0, 0, "CPLEX")        
            print("UNSAT this dataset.")

max_profit_solution()