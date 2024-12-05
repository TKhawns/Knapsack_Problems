from ortools.sat.python import cp_model
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import matplotlib.pyplot as plt
from itertools import combinations 

id_counter = 0

def display_solution(strip, rectangles, pos_circuits):
    # define Matplotlib figure and axis
    ax = plt.subplots()
    ax = plt.gca()
    plt.title(strip)

    if len(pos_circuits) > 0:
        for i in range(len(rectangles)):
            rect = plt.Rectangle(pos_circuits[i], *rectangles[i], edgecolor="#333")
            ax.add_patch(rect)

    ax.set_xlim(0, strip[0])
    ax.set_ylim(0, strip[1] + 1)
    ax.set_xticks(range(strip[0] + 1))
    ax.set_yticks(range(strip[1] + 1))
    ax.set_xlabel('width')
    ax.set_ylabel('height')
    # display plot
    plt.show()

def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path =  'output/'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)

    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

def solve_by_pysat(input, time, result, num_var, num_clause, name_lib):
    global id_counter
    id_counter += 1
    result_dict = {
        "ID": id_counter,
        "Problem": input,
        "Type": name_lib,
        "Time": time,
        "Result": result,
        "Variables": num_var,
        "Clauses": num_clause
    }
    write_to_xlsx(result_dict)

def findListSumOfProfits(rectangles, numberOfItems):
    # parameter rectangles: include profit [(width, height, profit)...].
    result_for_loop = []
    # Output is [[listOfRects, profitOfThisListRects], ...[]]
    while numberOfItems > 0:
        comb = combinations(rectangles, numberOfItems)
        for var in comb:
            temp = []
            rect_in_comb = [(x[0], x[1]) for x in list(var)]
            profit = sum(x[2] for x in list(var))
            temp.append(rect_in_comb)
            temp.append(profit)
            result_for_loop.append(temp)

        numberOfItems -= 1

    result_for_loop = sorted(result_for_loop, reverse=True, key=lambda x: x[1])


    print("Result of function:", result_for_loop)
    return result_for_loop

def cpSAT_OPP(rectangles, W, H, profits):
    model = cp_model.CpModel()

    n = len(rectangles)
    
    # Variables for position and rotation
    x = [model.NewIntVar(0, W, f'x_{i}') for i in range(n)]
    y = [model.NewIntVar(0, H, f'y_{i}') for i in range(n)]
    r = [model.NewBoolVar(f'r_{i}') for i in range(n)]   # Rotation variable
    xs = [model.NewBoolVar(f'a_{i}') for i in range(n)]  # Selection variable

    # (1) Domain of rectangles (x[i], y[i]) with rotation
    for i in range(n):
        wi, hi = rectangles[i]

        # Constraint: width and height based on rotation
        model.Add(x[i] + (1 - r[i]) * wi + r[i] * hi <= W)
        model.Add(y[i] + (1 - r[i]) * hi + r[i] * wi <= H)

        # If it out of strip, it must be rotated
        if wi > W or hi > H:
            model.Add(r[i] == 1)
        # Prevent rotation for squares and strip.
        if wi == hi or (wi > H or hi > W):
            model.Add(r[i] == 0)
    
    # (2) Large Rectangles Optimization
    for i in range(n):
        for j in range(i + 1, n):
            wi, hi = rectangles[i]
            wj, hj = rectangles[j]

            # Horizontal packing not possible
            if wi + wj > W:
                model.Add(x[i] + (1 - r[i]) * wi + r[i] * hi <= x[j]).OnlyEnforceIf(xs[i].Not())
                model.Add(x[j] + (1 - r[j]) * wj + r[j] * hj <= x[i]).OnlyEnforceIf(xs[j].Not())

            # Vertical packing not possible
            if hi + hj > H:
                model.Add(y[i] + (1 - r[i]) * hi + r[i] * wi <= y[j]).OnlyEnforceIf(xs[i].Not())
                model.Add(y[j] + (1 - r[j]) * hj + r[j] * wj <= y[i]).OnlyEnforceIf(xs[j].Not())

    # (3) Same Rectangles Optimization
    for i in range(n):
        for j in range(i + 1, n):
            if rectangles[i] == rectangles[j]:
                # Fix positional relationship: r_i is always to the left or below r_j
                model.Add(x[i] <= x[j])
                model.Add(y[i] <= y[j])

    # (2) Non-overlapping constraints
    for i in range(n):
        for j in range(i + 1, n):
            wi, hi = rectangles[i]
            wj, hj = rectangles[j]

            # Non-overlapping constraints with rotation
            no_overlap_1 = model.NewBoolVar(f'no_overlap_1_{i}_{j}')  # x[i] + width[i] <= x[j]
            no_overlap_2 = model.NewBoolVar(f'no_overlap_2_{i}_{j}')  # x[j] + width[j] <= x[i]
            no_overlap_3 = model.NewBoolVar(f'no_overlap_3_{i}_{j}')  # y[i] + height[i] <= y[j]
            no_overlap_4 = model.NewBoolVar(f'no_overlap_4_{i}_{j}')  # y[j] + height[j] <= y[i]

            model.Add(x[i] + (1 - r[i]) * wi + r[i] * hi <= x[j]).OnlyEnforceIf(no_overlap_1)
            model.Add(x[j] + (1 - r[j]) * wj + r[j] * hj <= x[i]).OnlyEnforceIf(no_overlap_2)
            model.Add(y[i] + (1 - r[i]) * hi + r[i] * wi <= y[j]).OnlyEnforceIf(no_overlap_3)
            model.Add(y[j] + (1 - r[j]) * hj + r[j] * wj <= y[i]).OnlyEnforceIf(no_overlap_4)

            # Ensure at least one non-overlapping condition holds
            model.AddBoolOr([no_overlap_1, no_overlap_2, no_overlap_3, no_overlap_4])

    # (1) Domain reduction for the maximum rectangle by width
    max_width = max(rect[1] for rect in rectangles)

    for i, (wi, hi) in enumerate(rectangles):
        if wi == max_width:
            # Restrict horizontal domain
            max_domain = (W - wi) // 2
            model.Add(x[i] <= max_domain)

    # Constraints max total profit
    # Not work.
    # model.maximize(sum(x * v for x, v in zip(xs, profits)))

    # CP Solver
    solver = cp_model.CpSolver()
    status = solver.Solve(model)
    
    if status == cp_model.OPTIMAL:
        total = 0
        selected_rectangles = []
        pos = []
        
        for i, xs_val in enumerate(xs):
            if solver.value(xs_val):
                total += profits[i]
                selected_rectangles.append(rectangles[i])
                pos.append([solver.value(x[i]), solver.value(y[i])])
                
        print("Total profit:", total)

        # Visualize solution
        strip = (W, H)
        return "sat"
        display_solution(strip, selected_rectangles, pos)
    else:
        print('No solution found.')
        return "unsat"


def max_profit_solution():
    with open('../dataset/dataset.txt', 'r') as file:
        lines = file.readlines()
        len_file = len(lines)

        is_Sol = ""

        for i in range(0, len_file, 4):
            list_strip = list(map(int,lines[i].strip().split()))
            widths = list(map(int, lines[i + 1].strip().split()))
            heights = list(map(int, lines[i + 2].strip().split()))
            profits = list(map(int, lines[i + 3].strip().split()))
            num_items = len(widths)

            rectangles = []
            for j in range (0, num_items):
                item = (widths[j], heights[j], profits[j])
                rectangles.append(item)
            
            # Update code
            # input_list sample = [ [ [(1, 3), (2, 1), (2, 1), (1, 1), (1, 2), (1, 1), (1, 1), (6, 2), (2, 1), (1, 1)], 33], ...[] ]
            input_list = findListSumOfProfits(rectangles, num_items)
            for input in input_list:
                status = cpSAT_OPP(input[0], list_strip[0], list_strip[1], profits)
                if status == "sat":
                      is_Sol = "ok"
                      print("Rectangles: ", input[0])
                      print("Max profit is: ", input[1])
                      break
                    
                elif status == "unsat":
                    print("\n", "Unsat OPP constraint in this input: ", input[0])
                    continue
        if is_Sol != "ok":
            print("No solution found")            

max_profit_solution()