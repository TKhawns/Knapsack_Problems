from pysat.formula import WCNF
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import timeit
from pysat.examples.rc2 import RC2
import matplotlib.pyplot as plt

time_budget = 100
id_counter = 0

def display_solution(strip, rectangles, pos_circuits):
    # define Matplotlib figure and axis
    fig, ax = plt.subplots()
    ax = plt.gca()
    plt.title(strip)

    if len(pos_circuits) > 0:
        for i in range(len(rectangles)):
            rect = plt.Rectangle(pos_circuits[i], *rectangles[i], edgecolor="#333")
            ax.add_patch(rect)

    ax.set_xlim(0, strip[0])
    ax.set_ylim(0, strip[1] + 1)
    ax.set_xticks(range(strip[0] + 1))
    ax.set_yticks(range(strip[1] + 1))
    ax.set_xlabel('width')
    ax.set_ylabel('height')
    # display plot
    plt.show()

def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path =  'output/'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)

    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

def solve_by_pysat(input, time, result, num_var, num_clause, name_lib):
    global id_counter
    id_counter += 1
    result_dict = {
        "ID": id_counter,
        "Problem": input,
        "Type": name_lib,
        "Time": time,
        "Result": result,
        "Variables": num_var,
        "Clauses": num_clause
    }
    write_to_xlsx(result_dict)


def positive_range(end):
    if (end < 0):
        return []
    return range(end)

def opp_solution(rectangles, strip, profits):
    # Define the variables
    cnf = WCNF()

    width = strip[0]
    height = strip[1]
    variables = {}
    counter = 1
    n = len(rectangles)

    # X_i = 1 if item i is selected.
    for i in range(n):
        variables[f"X{i + 1}"] = counter
        counter += 1

    # create lr, ud, px, py variables from 1 to N.
    # SAT Encoding of 2OPP
    for i in range(n):
        for j in range(n):
            if i != j:
                variables[f"lr{i + 1},{j + 1}"] = counter  # lri,rj
                counter += 1
                variables[f"ud{i + 1},{j + 1}"] = counter  # uri,rj
                counter += 1
        for e in range(width):
            variables[f"px{i + 1},{e}"] = counter  # pxi,e
            counter += 1
        for f in range(height):
            variables[f"py{i + 1},{f}"] = counter  # pyi,f
            counter += 1

    # Rotated variables
    for i in range(n):
        variables[f"r{i + 1}"] = counter
        counter += 1

    # Add the 2-literal axiom clauses (order constraint)
    # Formula (3).
    for i in range(n):
        for e in range(width - 1):  # -1 because we're using e+1 in the clause
            cnf.append([-variables[f"X{i + 1}"], -variables[f"px{i + 1},{e}"],
                        variables[f"px{i + 1},{e + 1}"]])
        for f in range(height - 1):  # -1 because we're using f+1 in the clause
            cnf.append([-variables[f"X{i + 1}"], -variables[f"py{i + 1},{f}"],
                        variables[f"py{i + 1},{f + 1}"]])
    # Add the 3-literal non-overlapping constraints
    # Formula (4).
    def non_overlapping(rotated, i, j, h1, h2, v1, v2):
        if not rotated:
            i_width = rectangles[i][0]
            i_height = rectangles[i][1]
            j_width = rectangles[j][0]
            j_height = rectangles[j][1]
            i_rotation = variables[f"r{i + 1}"]
            j_rotation = variables[f"r{j + 1}"]
        else:
            i_width = rectangles[i][1]
            i_height = rectangles[i][0]
            j_width = rectangles[j][1]
            j_height = rectangles[j][0]
            i_rotation = -variables[f"r{i + 1}"]
            j_rotation = -variables[f"r{j + 1}"]

        # Square symmertry breaking, if i is square than it cannot be rotated
        if i_width == i_height and rotated:
            i_square = True
            cnf.append([-variables[f"r{i + 1}"]])
        else:
            i_square = False

        if j_width == j_height and rotated:
            j_square = True
            cnf.append([-variables[f"r{j + 1}"]])
        else:
            j_square = False

        # Pick item constraint.
        extra_cnf = [-variables[f"X{i + 1}"], -variables[f"X{j + 1}"]]
        # lri,j v lrj,i v udi,j v udj,i
        four_literal = []
        if h1: four_literal.append(variables[f"lr{i + 1},{j + 1}"])
        if h2: four_literal.append(variables[f"lr{j + 1},{i + 1}"])
        if v1: four_literal.append(variables[f"ud{i + 1},{j + 1}"])
        if v2: four_literal.append(variables[f"ud{j + 1},{i + 1}"])

        cnf.append(four_literal + [i_rotation]  + extra_cnf)
        cnf.append(four_literal + [j_rotation]  + extra_cnf)

        # ¬lri, j ∨ ¬pxj, e
        if h1 and not i_square:
            for e in range(min(width, i_width)):
                    cnf.append(extra_cnf + [i_rotation,
                                -variables[f"lr{i + 1},{j + 1}"],
                                -variables[f"px{j + 1},{e}"]])
        # ¬lrj,i ∨ ¬pxi,e
        if h2 and not j_square:
            for e in range(min(width, j_width)):
                    cnf.append(extra_cnf + [j_rotation,
                                -variables[f"lr{j + 1},{i + 1}"],
                                -variables[f"px{i + 1},{e}"]]  )
        # ¬udi,j ∨ ¬pyj,f
        if v1 and not i_square:
            for f in range(min(height, i_height)):
                    cnf.append(extra_cnf + [i_rotation,
                                -variables[f"ud{i + 1},{j + 1}"],
                                -variables[f"py{j + 1},{f}"]] )
        # ¬udj, i ∨ ¬pyi, f,
        if v2 and not j_square:
            for f in range(min(height, j_height)):
                    cnf.append(extra_cnf + [j_rotation,
                                -variables[f"ud{j + 1},{i + 1}"],
                                -variables[f"py{i + 1},{f}"]])

        for e in positive_range(width - i_width):
            # ¬lri,j ∨ ¬pxj,e+wi ∨ pxi,e
            if h1 and not i_square:
                    cnf.append(extra_cnf + [i_rotation,
                                -variables[f"lr{i + 1},{j + 1}"],
                                variables[f"px{i + 1},{e}"],
                                -variables[f"px{j + 1},{e + i_width}"]])

        for e in positive_range(width - j_width):
            # ¬lrj,i ∨ ¬pxi,e+wj ∨ pxj,e
            if h2 and not j_square:
                    cnf.append(extra_cnf + [j_rotation,
                                -variables[f"lr{j + 1},{i + 1}"],
                                variables[f"px{j + 1},{e}"],
                                -variables[f"px{i + 1},{e + j_width}"]])

        for f in positive_range(height - i_height):
            # udi,j ∨ ¬pyj,f+hi ∨ pxi,e
            if v1 and not i_square:
                    cnf.append(extra_cnf + [i_rotation,
                                -variables[f"ud{i + 1},{j + 1}"],
                                variables[f"py{i + 1},{f}"],
                                -variables[f"py{j + 1},{f + i_height}"]])
        for f in positive_range(height - j_height):
            # ¬udj,i ∨ ¬pyi,f+hj ∨ pxj,f
            if v2 and not j_square:
                    cnf.append(extra_cnf + [j_rotation,
                                -variables[f"ud{j + 1},{i + 1}"],
                                variables[f"py{j + 1},{f}"],
                                -variables[f"py{i + 1},{f + j_height}"]])

    for i in range(n):
        for j in range(i + 1, n):
            # lri,j ∨ lrj,i ∨ udi,j ∨ udj,i
            #Large-rectangles horizontal
            if min(rectangles[i][0], rectangles[i][1]) + min(rectangles[j][0], rectangles[j][1]) > width:
                non_overlapping(False, i, j, False, False, True, True)
                non_overlapping(True, i, j, False, False, True, True)
            # Large rectangles vertical
            elif min(rectangles[i][0], rectangles[i][1]) + min(rectangles[j][0], rectangles[j][1]) > height:
                non_overlapping(False, i, j, True, True, False, False)
                non_overlapping(True, i, j, True, True, False, False)

            # Same rectangle and is a square
            elif rectangles[i] == rectangles[j]:
                if rectangles[i][0] == rectangles[i][1]:
                    cnf.append([-variables[f"r{i + 1}"]])
                    cnf.append([-variables[f"r{j + 1}"]])
                    non_overlapping(False,i ,j, True, True, True, True)
                else:
                    non_overlapping(False, i, j, True, True, True, True)
                    non_overlapping(True, i, j, True, True, True, True)
            # normal rectangles
            else:
                non_overlapping(False, i, j, True, True, True, True)
                non_overlapping(True, i, j, True, True, True, True)


   # Domain encoding to ensure every rectangle stays inside strip's boundary
    for i in range(n):
        if rectangles[i][0] > width: #if rectangle[i]'s width larger than strip's width, it has to be rotated
            cnf.append([variables[f"r{i + 1}"]])
        else:
            for e in range(width - rectangles[i][0], width):
                    cnf.append([variables[f"r{i + 1}"],
                                variables[f"px{i + 1},{e}"]])
        if rectangles[i][1] > height:
            cnf.append([variables[f"r{i + 1}"]])
        else:
            for f in range(height - rectangles[i][1], height):
                    cnf.append([variables[f"r{i + 1}"],
                                variables[f"py{i + 1},{f}"]])

        # Rotated
        if rectangles[i][1] > width:
            cnf.append([-variables[f"r{i + 1}"]])
        else:
            for e in range(width - rectangles[i][1], width):
                    cnf.append([-variables[f"r{i + 1}"],
                                variables[f"px{i + 1},{e}"]])
        if rectangles[i][0] > height:
            cnf.append([-variables[f"r{i + 1}"]])
        else:
            for f in range(height - rectangles[i][0], height):
                cnf.append([-variables[f"r{i + 1}"],
                            variables[f"py{i + 1},{f}"]])
                
    # weight constraint of MAXSAT
    # for i in range(n):
    #     cnf.append([variables[f"X{i + 1}"]])

    for i in range(n):
        cnf.append([variables[f"X{i + 1}"]], weight=profits[i])
        

    # add all clauses to SAT solver
    start = timeit.default_timer()
    with RC2(cnf) as rc2: #add all cnf to solver
        model = rc2.compute()
    
        print("Cost: ", rc2.cost)
        print("model: ", model)
        if model:
            pos = [[0 for i in range(2)] for j in range(len(rectangles))]
            rotation = []
            print("SAT")
            result = {}
            solver_time = format(timeit.default_timer() - start, ".3f")
            for var in model:
                if var > 0:
                    result[list(variables.keys())[list(variables.values()).index(var)]] = True
                else:
                    result[list(variables.keys())[list(variables.values()).index(-var)]] = False

            print("\nRectangle Placements:")
            print("-" * 50)
            selected_rectangles = []
            total_profit = 0
            for i in range(len(rectangles)):
                rotation.append(result[f"r{i + 1}"])
                for e in range(width - 1):
                    if result[f"px{i + 1},{e}"] == False and result[f"px{i + 1},{e + 1}"] == True:
                        pos[i][0] = e + 1
                    if e == 0 and result[f"px{i + 1},{e}"] == True:
                        pos[i][0] = 0
                for f in range(height - 1):
                    if result[f"py{i + 1},{f}"] == False and result[f"py{i + 1},{f + 1}"] == True:
                        pos[i][1] = f + 1
                    if f == 0 and result[f"py{i + 1},{f}"] == True:
                        pos[i][1] = 0
                
                # Print detailed information for each rectangle
                orig_width, orig_height = rectangles[i][0], rectangles[i][1]
                if rotation[i]:
                    orig_width = rectangles[i][1]
                    orig_height = rectangles[i][0]
                selected_rectangles.append([orig_width, orig_height])
                actual_width = orig_height if rotation[i] else orig_width
                actual_height = orig_width if rotation[i] else orig_height
                print(f"Rectangle {i+1}:")
                print(f"  Position: ({pos[i][0]}, {pos[i][1]})")
                print(f"  Dimensions: {actual_width}x{actual_height} {'(Rotated)' if rotation[i] else ''}")
                print(f"  Profit: {profits[i]}")
                total_profit += profits[i]
                print("-" * 50)
            print(f"Total profit: {total_profit}")
            display_solution(strip, selected_rectangles, pos)
            return(["sat", pos, rotation, solver_time])
        else:
            print("No solution")
            return "unsat"

def max_profit_solution():
    with open('../dataset/dataset.txt', 'r') as file:
        lines = file.readlines()
        len_file = len(lines)

        for i in range(0, len_file, 4):
            result = ""
            list_strip = list(map(int,lines[i].strip().split()))
            widths = list(map(int, lines[i + 1].strip().split()))
            heights = list(map(int, lines[i + 2].strip().split()))
            profits = list(map(int, lines[i + 3].strip().split()))
            strip = [list_strip[0], list_strip[1]]
            num_items = len(widths)

            rectangles = []
            for j in range (0, num_items):
                item = (widths[j], heights[j])
                rectangles.append(item)
            print("Rectangles: ", rectangles)

            result = opp_solution(rectangles, strip, profits)
            print("Result of MAXSAT: ", result)

max_profit_solution()
