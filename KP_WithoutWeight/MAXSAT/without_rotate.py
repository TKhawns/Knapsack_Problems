from pysat.formula import WCNF
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from pysat.examples.rc2 import RC2
import matplotlib.pyplot as plt
import time
import signal
time_budget = 100
id_counter = 0

def interrupt(s):
    s.interrupt()

def display_solution(strip, rectangles, pos_circuits):
    ax = plt.subplots()
    ax = plt.gca()
    plt.title(strip)
    n = len(rectangles)
    if len(pos_circuits) > 0:
        for i in range(n):
            rect = plt.Rectangle(pos_circuits[i], *rectangles[i], edgecolor="#333", facecolor="#69b3a2", alpha=0.5)
            ax.add_patch(rect)
    else:
        print("No circuits to display.")

    ax.set_xlim(0, strip[0])
    ax.set_ylim(0, strip[1] + 1)
    ax.set_xticks(range(strip[0] + 1))
    ax.set_yticks(range(strip[1] + 1))
    ax.set_xlabel('width')
    ax.set_ylabel('height')

    plt.show()

def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path =  'output/'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)

    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

def export_csv(problem_name, method_name, time, result, num_var, num_clause, max_profit):
    global id_counter
    id_counter += 1
    result_dict = {
        "No": id_counter,
        "Problem": problem_name,
        "Type": method_name,
        "Time": time,
        "Result": result,
        "Variables": num_var,
        "Clauses": num_clause,
        "Max profit": max_profit
    }
    write_to_xlsx(result_dict)

def positive_range(end):
    if (end < 0):
        return []
    return range(end)

def maxsat_constraints(rectangles, strip, profits, file_name):
    # Define the variables
    cnf = WCNF()
    variables = {}
    counter = 1
    n = len(rectangles)
    constraint_count = 0
    width = strip[0]
    height = strip[1]
    # find max height and width rectangles for largest rectangle symmetry breaking
    max_height = max([int(rectangle[1]) for rectangle in rectangles])
    max_width = max([int(rectangle[0]) for rectangle in rectangles])

    # a_i = 1 if item i is selected.
    for i in range(n):
        variables[f"a{i + 1}"] = counter
        counter += 1

    for i in range(n):
        for j in range(n):
            if i != j:
                variables[f"lr{i + 1},{j + 1}"] = counter  # lri,rj
                counter += 1
                variables[f"ud{i + 1},{j + 1}"] = counter  # uri,rj
                counter += 1
        for e in range(width):
            variables[f"px{i + 1},{e}"] = counter  # pxi,e
            counter += 1
        for f in range(height):
            variables[f"py{i + 1},{f}"] = counter  # pyi,f
            counter += 1


    # Add the 2-literal axiom clauses (order constraint)
    for i in range(n):
        for e in range(width - 1):  # -1 because we're using e+1 in the clause
            cnf.append([-variables[f"a{i + 1}"], -variables[f"px{i + 1},{e}"],
                        variables[f"px{i + 1},{e + 1}"]])
            constraint_count += 1
        for f in range(height - 1):
            cnf.append([-variables[f"a{i + 1}"], -variables[f"py{i + 1},{f}"],
                        variables[f"py{i + 1},{f + 1}"]])
            constraint_count += 1

     # Add the 3-literal non-overlapping constraints
    def non_overlapping(i, j, h1, h2, v1, v2):
        i_width = rectangles[i][0]
        i_height = rectangles[i][1]
        j_width = rectangles[j][0]
        j_height = rectangles[j][1]

        # lri, j ∨ lrj, i ∨ udi, j ∨ udj, i
        extra_cnf = [-variables[f"a{i + 1}"], -variables[f"a{j + 1}"]]

        four_literal = []
        if h1: four_literal.append(variables[f"lr{i + 1},{j + 1}"])
        if h2: four_literal.append(variables[f"lr{j + 1},{i + 1}"])
        if v1: four_literal.append(variables[f"ud{i + 1},{j + 1}"])
        if v2: four_literal.append(variables[f"ud{j + 1},{i + 1}"])
        cnf.append(extra_cnf + four_literal)

        # ¬lri, j ∨ ¬pxj, e
        if h1:
            for e in range(i_width):
                if f"px{j + 1},{e}" in variables:
                    cnf.append(extra_cnf + [-variables[f"lr{i + 1},{j + 1}"],
                                -variables[f"px{j + 1},{e}"]])
        # ¬lrj,i ∨ ¬pxi,e
        if h2:
            for e in range(j_width):
                if f"px{i + 1},{e}" in variables:
                    cnf.append(extra_cnf + [-variables[f"lr{j + 1},{i + 1}"],
                                -variables[f"px{i + 1},{e}"]])
        # ¬udi,j ∨ ¬pyj,f
        if v1:
            for f in range(i_height):
                if f"py{j + 1},{f}" in variables:
                    cnf.append(extra_cnf + [-variables[f"ud{i + 1},{j + 1}"],
                                -variables[f"py{j + 1},{f}"]])
        # ¬udj, i ∨ ¬pyi, f,
        if v2:
            for f in range(j_height):
                if f"py{i + 1},{f}" in variables:
                    cnf.append(extra_cnf + [-variables[f"ud{j + 1},{i + 1}"],
                                -variables[f"py{i + 1},{f}"]])

        for e in positive_range(width - i_width):
            # ¬lri,j ∨ ¬pxj,e+wi ∨ pxi,e
            if h1:
                if f"px{j + 1},{e + i_width}" in variables:
                    cnf.append(extra_cnf + [-variables[f"lr{i + 1},{j + 1}"],
                                variables[f"px{i + 1},{e}"],
                                -variables[f"px{j + 1},{e + i_width}"]])
            # ¬lrj,i ∨ ¬pxi,e+wj ∨ pxj,e
            if h2:
                if f"px{i + 1},{e + j_width}" in variables:
                    cnf.append(extra_cnf + [-variables[f"lr{j + 1},{i + 1}"],
                                variables[f"px{j + 1},{e}"],
                                -variables[f"px{i + 1},{e + j_width}"]])

        for f in positive_range(height - i_height):
            # udi,j ∨ ¬pyj,f+hi ∨ pxi,e
            if v1:
                if f"py{j + 1},{f + i_height}" in variables:
                    cnf.append(extra_cnf + [-variables[f"ud{i + 1},{j + 1}"],
                                variables[f"py{i + 1},{f}"],
                                -variables[f"py{j + 1},{f + i_height}"]])
            # ¬udj,i ∨ ¬pyi,f+hj ∨ pxj,f
            if v2:
                if f"py{i + 1},{f + j_height}" in variables:
                    cnf.append(extra_cnf + [-variables[f"ud{j + 1},{i + 1}"],
                                variables[f"py{j + 1},{f}"],
                                -variables[f"py{i + 1},{f + j_height}"]])

    for i in range(len(rectangles)):
        for j in range(i + 1, len(rectangles)):
            # lri,j ∨ lrj,i ∨ udi,j ∨ udj,i
            #Large-rectangles horizontal
            if rectangles[i][0] + rectangles[j][0] > width:
                non_overlapping(i, j, False, False, True, True)

            #Large-rectangles vertical
            if rectangles[i][1] + rectangles[j][1] > height:
                non_overlapping(i, j, True, True, False, False)

            #Same-sized rectangles
            elif rectangles[i] == rectangles[j]:
                non_overlapping(i, j, True, False, True, True)
            #
            #largest width rectangle
            elif rectangles[i][0] == max_width and rectangles[j][0] > (width - max_width) / 2:
                non_overlapping(i, j, False, True, True, True)
            #
            #largest height rectangle
            elif rectangles[i][1] == max_height and rectangles[j][1] > (height - max_height) / 2:
                non_overlapping(i, j, True, True, False, True)

           #normal rectangles
            else:
                non_overlapping(i, j, True, True, True, True)

   # Domain encoding to ensure every rectangle stays inside strip's boundary
    for i in range(n):
        # cnf.append([-variables[f"a{i + 1}"]])
        if rectangles[i][0] > width: #if rectangle[i]'s width larger than strip's width, it has to be rotated
            cnf.append([-variables[f"a{i + 1}"]])
            constraint_count += 1
        else:
            for e in range(width - rectangles[i][0], width):
                    cnf.append([
                                variables[f"px{i + 1},{e}"], -variables[f"a{i + 1}"]])
                    constraint_count += 1
        if rectangles[i][1] > height:
            cnf.append([-variables[f"a{i + 1}"]])
            constraint_count += 1
        else:
            for f in range(height - rectangles[i][1], height):
                    cnf.append([
                                variables[f"py{i + 1},{f}"], -variables[f"a{i + 1}"]])
                    constraint_count += 1
                
    # WCNF weight constraint of MAXSAT
    for i in range(n):
        cnf.append([variables[f"a{i + 1}"]], weight=profits[i])


    # os.makedirs("wcnf_no_rotate", exist_ok=True)
    # output_file = os.path.join("wcnf_no_rotate", f"{os.path.splitext(file_name)[0]}.wcnf")
    # cnf.to_file(output_file)

    def handler(signum, frame):
        raise TimeoutError("RC2 solver timed out")

    signal.signal(signal.SIGALRM, handler)
    signal.alarm(600)  # 5 seconds timeout
    total_profit = 0

    try:
        # add all clauses to SAT solver
        with RC2(cnf) as rc2: # add all cnf to solver
            model = rc2.compute()  # RC2 natively supports time limits
        
            if model is None:
                return ["TIMEOUT"]
            if model:
                # Initial result position of rectangles by [-1, -1]
                # pos = [[-1 for i in range(2)] for j in range(len(rectangles))]
                result = {}

                for var in model:
                    if var > 0:
                        result[list(variables.keys())[list(variables.values()).index(var)]] = True
                    else:
                        result[list(variables.keys())[list(variables.values()).index(-var)]] = False

                # rotation = []
                selected_rectangles = []
                # result_rectangle = []
                # res_pos = []

                for i in range(n):
                    # rotation.append(result[f"r{i + 1}"])
                    selected_rectangles.append(result[f"a{i + 1}"])

                    # Print detailed information for each rectangle
                    if selected_rectangles[i] == True:
                        total_profit += profits[i]
     
                return ["SAT", counter, len(cnf.hard) + len(cnf.soft), total_profit]
            else:
                return ["UNSAT"]
    except TimeoutError:
        print("Solver exceeded time limit")
        return ["UNSAT"]
    finally:
        print("aa")
        signal.alarm(0)  # Disable the alarm
        if (total_profit == 0):
            return ["UNSAT"]
        return ["SAT", counter, len(cnf.hard) + len(cnf.soft), total_profit]

def max_profit_solution():
    folder_path = './miss_data/'
    for file_name in os.listdir(folder_path):
        file_path = os.path.join(folder_path, file_name)
        if os.path.isfile(file_path) and file_name.endswith('.txt') and file_name == "gcut1.txt":
            print(f"Processing file: {file_name}")
            with open(file_path, 'r') as file:
                lines = file.readlines()
                list_strip = list(map(int,lines[0].strip().split()))
                widths = list(map(int, lines[1].strip().split()))
                heights = list(map(int, lines[2].strip().split()))
                profits = list(map(int, lines[3].strip().split()))
                num_items = len(widths)

                # Get list rectangles from input data.
                rectangles = []
    
                for j in range (0, num_items):
                    item = (widths[j], heights[j], profits[j])
                    rectangles.append(item)

                started_time = time.time()
                status = maxsat_constraints(rectangles, list_strip, profits, file_name)
                end_time = time.time()
                print(file_name, end_time - started_time)
                # if status[0] == "TIMEOUT":
                #     export_csv(file_name, "maxsat_non_rotate", end_time - started_time, "TIMEOUT", 0, 0, 0)
                # if status[0] == "SAT":
                #     export_csv(file_name, "maxsat_non_rotate", end_time - started_time, "SAT", status[1], status[2], status[3])
                # if status[0] == "UNSAT":
                #     export_csv(file_name, "maxsat_non_rotate", end_time - started_time, "UNSAT", 0, 0, 0)

max_profit_solution()

