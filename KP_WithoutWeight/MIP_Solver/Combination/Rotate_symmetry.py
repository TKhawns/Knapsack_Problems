import os
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from ortools.linear_solver import pywraplp
import matplotlib.pyplot as plt
from itertools import combinations 
import time

id_counter = 0

def display_solution(strip, rectangles, pos_circuits):
    # define Matplotlib figure and axis
    ax = plt.subplots()
    ax = plt.gca()
    plt.title(strip)

    if len(pos_circuits) > 0:
        for i in range(len(rectangles)):
            rect = plt.Rectangle(pos_circuits[i], *rectangles[i], edgecolor="#333")
            ax.add_patch(rect)

    ax.set_xlim(0, strip[0])
    ax.set_ylim(0, strip[1] + 1)
    ax.set_xticks(range(strip[0] + 1))
    ax.set_yticks(range(strip[1] + 1))
    ax.set_xlabel('width')
    ax.set_ylabel('height')
    plt.show()

def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path =  'output/'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)

    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

def solve_by_pysat(input, time, result, num_var, num_clause, name_lib):
    global id_counter
    id_counter += 1
    result_dict = {
        "ID": id_counter,
        "Problem": input,
        "Type": name_lib,
        "Time": time,
        "Result": result,
        "Variables": num_var,
        "Clauses": num_clause
    }
    write_to_xlsx(result_dict)


def MIP_OPP(rectangles, W, H):
    # Define MIP Solver
    strip = (W, H)
    solver = pywraplp.Solver.CreateSolver('SCIP')
    n = len(rectangles)

    # Define variables for positions of each rectangle (x[i], y[i])
    x = [solver.IntVar(0, W, f'x_{i}') for i in range(n)]
    y = [solver.IntVar(0, H, f'y_{i}') for i in range(n)]
    # Rotation variable
    r = [solver.BoolVar(f'r_{i}') for i in range(n)]  
    xs = [solver.BoolVar(f'a_{i}') for i in range(n)]  


    # (1) Domain of rectangles (x[i], y[i]) with rotation
    for i in range(n):
        wi, hi = rectangles[i]

        # Adjust width and height based on rotation
        solver.Add(x[i] + (1 - r[i]) * wi + r[i] * hi <= W)
        solver.Add(y[i] + (1 - r[i]) * hi + r[i] * wi <= H)

        if wi > W or hi > H:  # Must rotate to fit
            solver.Add(r[i] == 1)
        elif wi == hi or (wi > H or hi > W):  # Rotation not allowed if it still won't fit
            solver.Add(r[i] == 0)
    
    # (2) Large Rectangles Optimization
    for i in range(n):
        for j in range(i + 1, n):
            wi, hi = rectangles[i]
            wj, hj = rectangles[j]

            # Horizontal packing not possible
            if wi + wj > W:
                solver.Add(x[i] + (1 - r[i]) * wi + r[i] * hi <= x[j] + W * (1 - xs[j]))
                solver.Add(x[j] + (1 - r[j]) * wj + r[j] * hj <= x[i] + W * (1 - xs[i]))

            # Vertical packing not possible
            if hi + hj > H:
                solver.Add(y[i] + (1 - r[i]) * hi + r[i] * wi <= y[j] + H * (1 - xs[j]))
                solver.Add(y[j] + (1 - r[j]) * hj + r[j] * wj <= y[i] + H * (1 - xs[i]))

    # (3) Same Rectangles Optimization
    for i in range(n):
        for j in range(i + 1, n):
            if rectangles[i] == rectangles[j]:
                # Same rotation
                solver.Add(r[i] == r[j])

                # Fix positional relationship: Ensure r_i is always to the left or below r_j
                solver.Add(x[i] <= x[j])
                solver.Add(y[i] <= y[j])

    # (5) Non-overlapping constraints between rectangles
    for i in range(n):
        for j in range(i + 1, n):
            wi, hi = rectangles[i]
            wj, hj = rectangles[j]

            # Binary variables for overlap conditions
            no_overlap_1 = solver.BoolVar(f'no_overlap_1_{i}_{j}')  # x[i] + wi <= x[j]
            no_overlap_2 = solver.BoolVar(f'no_overlap_2_{i}_{j}')  # x[j] + wj <= x[i]
            no_overlap_3 = solver.BoolVar(f'no_overlap_3_{i}_{j}')  # y[i] + hi <= y[j]
            no_overlap_4 = solver.BoolVar(f'no_overlap_4_{i}_{j}')  # y[j] + hj <= y[i]

            solver.Add(x[i] + (1 - r[i]) * wi + r[i] * hi <= x[j] + W * (1 - no_overlap_1))
            solver.Add(x[j] + (1 - r[j]) * wj + r[j] * hj <= x[i] + W * (1 - no_overlap_2))
            solver.Add(y[i] + (1 - r[i]) * hi + r[i] * wi <= y[j] + H * (1 - no_overlap_3))
            solver.Add(y[j] + (1 - r[j]) * hj + r[j] * wj <= y[i] + H * (1 - no_overlap_4))

            # Ensure at least one no-overlap condition holds
            solver.Add(no_overlap_1 + no_overlap_2 + no_overlap_3 + no_overlap_4 >= 1)

    max_width = max(rect[1] for rect in rectangles)
    # [1] same rectangle.
    for i in range(n):
        for j in range(i + 1, n):
            if rectangles[i] == rectangles[j]:
                # Fix positional relationship
                solver.add(x[i] <= x[j])
                solver.add(y[i] <= y[j])

    for i, (wi, hi) in enumerate(rectangles):
        if wi == max_width:
            # Restrict horizontal domain
            max_domain = (W - wi) // 2
            solver.Add(x[i] <= max_domain)

    # Solve the model
    status = solver.Solve()
    # postition of result rectangles, input of visualize.
    pos = []
    selected_rectangles = []

    if status == pywraplp.Solver.OPTIMAL:
        print("Optimal solution found.")
        for i in range(n):
                wi, hi = rectangles[i]
                if x[i].solution_value() >= 0 and y[i].solution_value() >= 0:
                    pos.append([x[i].solution_value(), y[i].solution_value()])
                new_w = wi
                new_h = hi
                if r[i].solution_value() >= 0:
                    new_w = hi
                    new_h = wi
                selected_rectangles.append([new_w, new_h])

                # print(f"Rectangle: ({wi}, {hi})", f"Position (x, y) = ({x[i].solution_value()}, {y[i].solution_value()})"),  
                # selected_rectangles.append(rectangles[i])
            
        # print("Result rectangles: ", selected_rectangles)
        return ["SAT", solver.NumVariables(), solver.NumConstraints()]
        # display_solution(strip, selected_rectangles, pos)

    else:
        print("No optimal solution found.")
        return ["UNSAT", 0, 0]


# function to generate all combinations of list rectangle by numberOfItems.
def findListSumOfProfits(rectangles, numberOfItems):
    # parameter rectangles: include profit [(width, height, profit)...].
    result_for_loop = []
    # Output is [[listOfRects, profitOfThisListRects], ...[]]
    while numberOfItems > 0:
        comb = combinations(rectangles, numberOfItems)
        for var in comb:
            temp = []
            rect_in_comb = [(x[0], x[1]) for x in list(var)]
            profit = sum(x[2] for x in list(var))
            temp.append(rect_in_comb)
            temp.append(profit)
            result_for_loop.append(temp)

        numberOfItems -= 1

    result_for_loop = sorted(result_for_loop, reverse=True, key=lambda x: x[1])

    # print("Result of function:", result_for_loop)
    return result_for_loop

def max_profit_solution():
    with open('../dataset/dataset8.txt', 'r') as file:
        lines = file.readlines()

        list_strip = list(map(int,lines[0].strip().split()))
        widths = list(map(int, lines[1].strip().split()))
        heights = list(map(int, lines[2].strip().split()))
        profits = list(map(int, lines[3].strip().split()))
        num_items = len(widths)
        is_sol = ""

        rectangles = []
        for j in range (0, num_items):
            item = (widths[j], heights[j], profits[j])
            rectangles.append(item)

        started_time = time.time()
        
        # Update code
        # input_list sample = [ [ [(1, 3), (2, 1), (2, 1), (1, 1), (1, 2), (1, 1), (1, 1), (6, 2), (2, 1), (1, 1)], 33], ...[] ]
        input_list = findListSumOfProfits(rectangles, num_items)
        for input in input_list:
            status = MIP_OPP(input[0], list_strip[0], list_strip[1])
            if status[0] == "SAT":
                is_sol = "SAT"
                #   print("Index: ", input_list.index(input))
                #   print("Time is:", time.time() - started_time)
                #   print("Rectangles: ", input[0])
                print("Max profit is: ", input[1])
                solve_by_pysat(num_items, time.time() - started_time, is_sol, status[1], status[2], "MIP")
                break
                
            elif status[0] == "UNSAT":
                print("\n", "Unsat OPP constraint in this input: ", input[0])
                continue          

max_profit_solution()