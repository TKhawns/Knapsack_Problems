import os
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from ortools.linear_solver import pywraplp
import matplotlib.pyplot as plt
import time

id_counter = 0

def display_solution(strip, rectangles, pos_circuits):
    ax = plt.subplots()
    ax = plt.gca()
    plt.title(strip)

    if len(pos_circuits) > 0:
        for i in range(len(rectangles)):
            rect = plt.Rectangle(pos_circuits[i], *rectangles[i], edgecolor="#333")
            ax.add_patch(rect)

    ax.set_xlim(0, strip[0])
    ax.set_ylim(0, strip[1] + 1)
    ax.set_xticks(range(strip[0] + 1))
    ax.set_yticks(range(strip[1] + 1))
    ax.set_xlabel('width')
    ax.set_ylabel('height')
    plt.show()

def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path =  'non_rotate/'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)

    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

def export_csv(problem_name, method_name, time, result, num_var, num_clause, max_profit):
    global id_counter
    id_counter += 1
    result_dict = {
        "No": id_counter,
        "Problem": problem_name,
        "Type": method_name,
        "Time": time,
        "Result": result,
        "Variables": num_var,
        "Clauses": num_clause,
        "Max profit": max_profit
    }
    write_to_xlsx(result_dict)


def mip_constraints(rectangles, W, H, profits, mid):
    # Create the SCIP solver.
    solver = pywraplp.Solver.CreateSolver('SCIP')
    solver.set_time_limit(600)

    n = len(rectangles)
    
    # Decision variables:
    # x[i], y[i] – lower‐left coordinates (integers)
    # xs[i] – selection binary variable (1 if rectangle i is used)
    x = [solver.IntVar(0, W, f'x_{i}') for i in range(n)]
    y = [solver.IntVar(0, H, f'y_{i}') for i in range(n)]
    xs = [solver.BoolVar(f'xs_{i}') for i in range(n)]
    
    # For each rectangle, the effective width and height are its original dimensions.
    for i in range(n):
        wi, hi, _ = rectangles[i]
        width_expr  = wi
        height_expr = hi
        
        # Big-M values for boundary constraints.
        big_M_x = W + max(wi, hi)
        big_M_y = H + max(wi, hi)
        
        # If rectangle i is selected then it must lie completely within the bin:
        solver.Add(x[i] + width_expr <= W + big_M_x * (1 - xs[i]))
        solver.Add(y[i] + height_expr <= H + big_M_y * (1 - xs[i]))
    
    # For non-overlap, if both rectangles i and j are selected, then one of the following must hold:
    #   [1] Rectangle i is to the left of j,
    #   [2] Rectangle j is to the left of i,
    #   [3] Rectangle i is below j, or
    #   [4] Rectangle j is below i.
    global_big_M_x = W + max(max(rect[0], rect[1]) for rect in rectangles)
    global_big_M_y = H + max(max(rect[0], rect[1]) for rect in rectangles)
    
    # Create and add non-overlap constraints for each pair (i, j)
    for i in range(n):
        for j in range(i+1, n):
            wi, hi, _ = rectangles[i]
            wj, hj, _ = rectangles[j]
            width_i  = wi
            height_i = hi
            width_j  = wj
            height_j = hj
            
            # Create four binary variables for the four disjuncts.
            b1 = solver.BoolVar(f'b_{i}_{j}_1')
            b2 = solver.BoolVar(f'b_{i}_{j}_2')
            b3 = solver.BoolVar(f'b_{i}_{j}_3')
            b4 = solver.BoolVar(f'b_{i}_{j}_4')
            
            # [1] i is to the left of j:
            solver.Add(x[i] + width_i <= x[j] + global_big_M_x * (1 - b1))
            # [2] j is to the left of i:
            solver.Add(x[j] + width_j <= x[i] + global_big_M_x * (1 - b2))
            # [3] i is below j:
            solver.Add(y[i] + height_i <= y[j] + global_big_M_y * (1 - b3))
            # [4] j is below i:
            solver.Add(y[j] + height_j <= y[i] + global_big_M_y * (1 - b4))
            
            # If both rectangles are selected then at least one of the four disjuncts must hold.
            solver.Add(b1 + b2 + b3 + b4 >= xs[i] + xs[j] - 1)
    
    # Total profit constraint: selected rectangles must achieve profit at least mid.
    solver.Add(solver.Sum([xs[i] * profits[i] for i in range(n)]) >= mid)
    
    # --- Symmetry constraints ---
    # [1] If two rectangles are identical, fix a consistent ordering.
    for i in range(n):
        for j in range(i+1, n):
            if rectangles[i] == rectangles[j]:
                solver.Add(x[i] <= x[j])
                solver.Add(y[i] <= y[j])
    # [2] Domain symmetry: restrict the horizontal domain for some rectangles.
    max_width = max(rect[1] for rect in rectangles)
    for i, (wi, hi, _) in enumerate(rectangles):
        if wi == max_width:
            solver.Add(x[i] <= (W - wi) // 2)
    
    # Objective: maximize total profit.
    objective = solver.Sum([xs[i] * profits[i] for i in range(n)])
    solver.Maximize(objective)
    
    # Set a time limit (SCIP expects milliseconds, so 1200 sec = 1,200,000 ms)
    solver.set_time_limit(600 * 1000)
    
    status = solver.Solve()
    
    if status == pywraplp.Solver.OPTIMAL or status == pywraplp.Solver.FEASIBLE:
        selected_rectangles = []
        positions = []
        for i in range(n):
            # Use a threshold (0.5) to interpret the binary decision variable.
            if xs[i].solution_value() > 0.5:
                selected_rectangles.append(rectangles[i])
                positions.append((x[i].solution_value(), y[i].solution_value()))
        num_var = solver.NumVariables()
        num_constr = solver.NumConstraints()
        total_profit = sum(rect[2] for rect in selected_rectangles)
        return ['SAT', total_profit, num_var, num_constr]
    else:
        return ["UNSAT"]
def max_profit_solution():
    folder_path = '../miss_data/'
    for file_name in os.listdir(folder_path):
        file_path = os.path.join(folder_path, file_name)
        if os.path.isfile(file_path) and file_name.endswith('.txt'):
            print(f"Processing file: {file_name}")
            with open(file_path, 'r') as file:
                lines = file.readlines()
                list_strip = list(map(int,lines[0].strip().split()))
                widths = list(map(int, lines[1].strip().split()))
                heights = list(map(int, lines[2].strip().split()))
                profits = list(map(int, lines[3].strip().split()))
                num_items = len(widths)

                # Get list rectangles from input data.
                rectangles = []
                prev_sat = []
                lower_bound = min(profits)
                upper_bound = sum(profits)
                for j in range (0, num_items):
                    item = (widths[j], heights[j], profits[j])
                    rectangles.append(item)

                isExport = ""
                started_time = time.time()
                while (lower_bound + 1 < upper_bound):
                    if (time.time() - started_time >= 600): 
                        isExport = "TIMEOUT"
                        if prev_sat == []:
                            export_csv(file_name, "mip_non_rotate", time.time() - started_time, "TIMEOUT", 0, 0, 0)
                            break
                        export_csv(file_name, "mip_non_rotate", time.time() - started_time, "TIMEOUT", prev_sat[1], prev_sat[2], prev_sat[3])
                        break
                    mid = lower_bound + (upper_bound - lower_bound) // 2
                    status = mip_constraints(rectangles, list_strip[0], list_strip[1], profits, mid)
                    if (status[0] == "UNSAT"):
                        upper_bound = mid
                        continue
                    elif (status[0] == "SAT"):
                        prev_sat = status
                        lower_bound = mid
                        continue
                ended_time = time.time()

                if isExport != "TIMEOUT":
                    if (prev_sat != [] and prev_sat[0]) == "SAT":
                        export_csv(file_name, "mip_non_rotate", ended_time - started_time, "SAT", prev_sat[1], prev_sat[2], prev_sat[3])
                    else: export_csv(file_name, "mip_non_rotate", ended_time - started_time, "UNSAT", 0, 0, 0)

max_profit_solution()