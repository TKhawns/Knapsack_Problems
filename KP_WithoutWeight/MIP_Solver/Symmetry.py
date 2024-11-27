import os
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from ortools.linear_solver import pywraplp
import matplotlib.pyplot as plt
import time


id_counter = 0

def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path =  'output/'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)

    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

def solve_by_pysat(input, time, result, num_var, num_clause, name_lib):
    global id_counter
    id_counter += 1
    result_dict = {
        "ID": id_counter,
        "Problem": input,
        "Type": name_lib,
        "Time": time,
        "Result": result,
        "Variables": num_var,
        "Clauses": num_clause
    }
    write_to_xlsx(result_dict)



def MIP_OPP(rectangles, W, H):
    strip = (W, H)

    # MIP Solver
    solver = pywraplp.Solver.CreateSolver('SAT')
    if not solver:
        print("Solver not found.")
        return

    n = len(rectangles)

    # Define variables for positions of each rectangle (x[i], y[i])
    x = [solver.IntVar(0, W, f'x_{i}') for i in range(n)]
    y = [solver.IntVar(0, H, f'y_{i}') for i in range(n)]

    # (1) Domain constraints for rectangles within the bin dimensions
    for i in range(n):
        wi, hi = rectangles[i]
        solver.Add(x[i] + wi <= W)  # Ensure rectangle width fits within W
        solver.Add(y[i] + hi <= H)  # Ensure rectangle height fits within H

    # (2) Symmetry reduction constraints
    max_width = max(rect[0] for rect in rectangles)
    max_height = max(rect[1] for rect in rectangles)

    for i in range(n):
        wi, hi = rectangles[i]

        # Domain reduction using horizontal symmetry
        if wi == max_width:
            solver.Add(x[i] <= (W - wi) // 2)

        # Domain reduction using vertical symmetry
        if hi == max_height:
            solver.Add(y[i] <= (H - hi) // 2)

    # (4) Same rectangle symmetry
    for i in range(n):
        for j in range(i + 1, n):
            if rectangles[i] == rectangles[j]:
                # Fix positional relationship to reduce symmetry
                solver.Add(x[i] <= x[j])
                solver.Add(y[i] <= y[j])

    # (5) Non-overlapping constraints between rectangles
    for i in range(n):
        for j in range(i + 1, n):
            wi, hi = rectangles[i]
            wj, hj = rectangles[j]

            # Binary variables for overlap conditions
            no_overlap_1 = solver.BoolVar(f'no_overlap_1_{i}_{j}')  # x[i] + wi <= x[j]
            no_overlap_2 = solver.BoolVar(f'no_overlap_2_{i}_{j}')  # x[j] + wj <= x[i]
            no_overlap_3 = solver.BoolVar(f'no_overlap_3_{i}_{j}')  # y[i] + hi <= y[j]
            no_overlap_4 = solver.BoolVar(f'no_overlap_4_{i}_{j}')  # y[j] + hj <= y[i]


            solver.Add(x[i] + wi <= x[j] + W * (1 - no_overlap_1))
            solver.Add(x[j] + wj <= x[i] + W * (1 - no_overlap_2))
            solver.Add(y[i] + hi <= y[j] + H * (1 - no_overlap_3))
            solver.Add(y[j] + hj <= y[i] + H * (1 - no_overlap_4))

            # Ensure at least one no-overlap condition holds
            solver.Add(no_overlap_1 + no_overlap_2 + no_overlap_3 + no_overlap_4 >= 1)

    # Solve the model
    status = solver.Solve()
    print(status)
    # postition of result rectangles, input of visualize.
    pos = []
    print("POS: ", pos)
    selected_rectangles = []

    if status == pywraplp.Solver.OPTIMAL:
        print("Optimal solution found.")
        for i in range(n):
                wi, hi = rectangles[i]
                if x[i].solution_value() >= 0 and y[i].solution_value() >= 0:
                    pos.append([x[i].solution_value(), y[i].solution_value()])

                print(f"Rectangle: ({wi}, {hi})", f"Position (x, y) = ({x[i].solution_value()}, {y[i].solution_value()})"),  
                selected_rectangles.append(rectangles[i])
            
        print("New pos:", pos)
        print("Result rectangles: ", selected_rectangles)
        # return ["sat", solver.wall_time()/1000]
        display_solution(strip, selected_rectangles, pos)

    else:
        print("No optimal solution found.")
        return ["unsat",solver.wall_time() / 1000]  # Convert to seconds


def max_profit_solution():
    with open('../dataset/dataset.txt', 'r') as file:
        lines = file.readlines()
        len_file = len(lines)

        for i in range(0, len_file, 4):
            result = ""
            list_strip = list(map(int,lines[i].strip().split()))
            widths = list(map(int, lines[i + 1].strip().split()))
            heights = list(map(int, lines[i + 2].strip().split()))
            profits = list(map(int, lines[i + 3].strip().split()))
            strip = (list_strip[0], list_strip[1])
            num_items = len(widths)

            rectangles = []
            for j in range (0, num_items):
                item = (widths[j], heights[j], profits[j])
                rectangles.append(item)

            sorted_rectangles = sorted(rectangles, key=lambda x: x[2])
            result = [(x[0], x[1]) for x in sorted_rectangles]
            profits_sorted = [x[2] for x in sorted_rectangles]  # Keep track of sorted profits

            print(result)


            start = time.time()
            elap_time = 0
            solution = ""

            for i in range(0, len(result)):
                status = MIP_OPP(result, list_strip[0], list_strip[1])
                if (len(result) == 0):
                    elap_time = time.time() - start
                    solution = "unsat"
                    break
                if (status == "sat"):
                    elap_time = time.time() - start
                    solution = "sat"
                    total_profit = sum(profits_sorted[i:])
                    print(f"Maximum profit achieved: {total_profit}")
                    break
                elif (status == "unsat"):
                    result.pop(0)

            solve_by_pysat(str(num_items), str(elap_time), solution, 0, 0, "Glucose3")

def display_solution(strip, rectangles, pos_circuits):
    # define Matplotlib figure and axis
    fig, ax = plt.subplots()
    ax = plt.gca()
    plt.title(strip)

    if len(pos_circuits) > 0:
        for i in range(len(rectangles)):
            rect = plt.Rectangle(pos_circuits[i], *rectangles[i], edgecolor="#333")
            ax.add_patch(rect)

    ax.set_xlim(0, strip[0])
    ax.set_ylim(0, strip[1] + 1)
    ax.set_xticks(range(strip[0] + 1))
    ax.set_yticks(range(strip[1] + 1))
    ax.set_xlabel('width')
    ax.set_ylabel('height')
    # display plot
    plt.show()

max_profit_solution()