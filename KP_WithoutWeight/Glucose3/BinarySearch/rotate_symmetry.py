from pysat.formula import CNF
from pysat.solvers import Glucose3
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import matplotlib.pyplot as plt
import time
from threading import Timer

id_counter = 0

def interrupt(s): s.interrupt()

def display_solution(strip, rectangles, pos_circuits):
    ax = plt.subplots()
    ax = plt.gca()
    plt.title(strip)

    if len(pos_circuits) > 0:
        for i in range(len(rectangles)):
            rect = plt.Rectangle(pos_circuits[i], *rectangles[i], edgecolor="#333", facecolor="#69b3a2", alpha=0.5)
            ax.add_patch(rect)
    else:
        print("No circuits to display.")

    ax.set_xlim(0, strip[0])
    ax.set_ylim(0, strip[1] + 1)
    ax.set_xticks(range(strip[0] + 1))
    ax.set_yticks(range(strip[1] + 1))
    ax.set_xlabel('width')
    ax.set_ylabel('height')

    # display plot
    plt.show()

def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path =  'rotate_symmetry/'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)

    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

def export_csv(problem_name, method_name, time, result, num_var, num_clause,num_var_solver,num_clause_solver,  max_profit):
    global id_counter
    id_counter += 1
    result_dict = {
        "No": id_counter,
        "Problem": problem_name,
        "Type": method_name,
        "Time": time,
        "Result": result,
        "Variables count": num_var,
        "Clauses count": num_clause,
        "Variables solver": num_var_solver,
        "Clauses solver": num_clause_solver,
        "Max profit": max_profit
    }
    write_to_xlsx(result_dict)

def positive_range(end):
    if (end < 0):
        return []
    return range(end)

def pos_i(i, k, profit):
    if i == 0:
        return 0
    if i < k:
        sum_w = sum(profit[1:i+1])
        return min(k, sum_w)
    else:
        return k

def glucose_constraints(rectangles, width, height, k, profit):
    # Define the variables
    strip = (width, height)
    cnf = CNF()
    variables = {}
    counter = 1
    constraint_count = 0
    n = len(rectangles)

    # SCPB Constraints definition
    profit = [0] + profit
    n_scpb = n
    # Create map_register to hold the auxiliary variables
    map_register = [[0 for _ in range(k + 1)] for _ in range(n_scpb + 1)]

    # a_i = True (1) if item i is selected.
    for i in range(n):
        variables[f"a{i + 1}"] = counter
        counter += 1

    # SAT Encoding of 2 0PP.
    # create lr, ud, px, py variables from 1 to N.
    for i in range(n):
        for j in range(n):
            if i != j:
                variables[f"lr{i + 1},{j + 1}"] = counter  # lri,rj
                counter += 1
                variables[f"ud{i + 1},{j + 1}"] = counter  # uri,rj
                counter += 1
        for e in range(width):
            variables[f"px{i + 1},{e}"] = counter  # pxi,e
            counter += 1
        for f in range(height):
            variables[f"py{i + 1},{f}"] = counter  # pyi,f
            counter += 1

    # Rotated variables of 2OPP.
    for i in range(n):
        variables[f"r{i + 1}"] = counter
        counter += 1

    ended_loop = counter
    # Init value of map_register in SCPB.
    for i in range(1, n_scpb):
        n_bits = pos_i(i, k, profit)
        for j in range(1, n_bits + 1):
            map_register[i][j] = counter
            counter += 1

    # (1) X_i -> R_i,j for j = 1 to w_i
    for i in range(1, n_scpb):
        for j in range(1, profit[i] + 1):
            if j <= pos_i(i, k, profit):
                cnf.append([-variables[f"a{i}"], map_register[i][j]])
                constraint_count += 1

    # (2) R_{i-1,j} -> R_i,j for j = 1 to pos_{i-1}
    for i in range(2, n_scpb):
        for j in range(1, pos_i(i - 1, k, profit) + 1):
            cnf.append([-map_register[i - 1][j], map_register[i][j]])
            constraint_count += 1

    # (3) X_i ^ R_{i-1,j} -> R_i,j+w_i for j = 1 to pos_{i-1}
    for i in range(2, n_scpb):
        for j in range(1, pos_i(i - 1, k, profit) + 1):
            if j + profit[i] <= k and j + profit[i] <= pos_i(i, k, profit):
                cnf.append([-variables[f"a{i}"], -map_register[i - 1][j], map_register[i][j + profit[i]]])
                constraint_count += 1

    # (4) ¬X_i ^ ¬R_{i-1,j} -> ¬R_i,j for j = 1 to pos_{i-1}
    for i in range(2, n_scpb):
        for j in range(1, pos_i(i - 1, k, profit) + 1):
            cnf.append([variables[f"a{i}"], map_register[i - 1][j], -map_register[i][j]])
            constraint_count += 1

    # (5) ¬X_i -> ¬R_i,j for j = 1 + pos_{i-1} to pos_i
    for i in range(1, n_scpb):
        # if pos_i(i - 1, k, weight) < k:
            for j in range(1 + pos_i(i - 1, k, profit), pos_i(i, k, profit) + 1):
                cnf.append([variables[f"a{i}"], -map_register[i][j]])
                constraint_count += 1

    # (6) ¬R_{i-1,j} -> ¬R_i,j+w_i for j = 1 to pos_{i-1}
    for i in range(2, n_scpb):
        # if pos_i(i - 1, k, weight) < k:
            for j in range(1, pos_i(i - 1, k, profit) + 1):
                if j + profit[i] <= k and j + profit[i] <= pos_i(i, k, profit):
                    cnf.append([map_register[i - 1][j], -map_register[i][j + profit[i]]])
                    constraint_count += 1

    # (7) R_{n-1,k} v (X_n ^ R_{n-1,k-w_n})
    if k > pos_i(n_scpb - 1, k, profit):
        if k - profit[n_scpb] > 0 and k - profit[n_scpb] <= pos_i(n_scpb - 1, k, profit):
            cnf.append([variables[f"a{n_scpb}"]])
            constraint_count += 1
            cnf.append([map_register[n_scpb - 1][k - profit[n_scpb]]])
            constraint_count += 1
    else:
        cnf.append([map_register[n_scpb - 1][k], variables[f"a{n_scpb}"]])
        constraint_count += 1
        if k - profit[n_scpb] > 0 and k - profit[n_scpb] <= pos_i(n_scpb - 1, k, profit):
            cnf.append([map_register[n_scpb - 1][k], map_register[n_scpb - 1][k - profit[n_scpb]]])
            constraint_count += 1

    # Add the 2-literal axiom clauses (order constraint)
    # Formula (3).
    # [Update] if a_i = 1 -> order-encoding.
    for i in range(n):
        for e in range(width - 1):  # -1 because we're using e+1 in the clause
            cnf.append([-variables[f"a{i + 1}"], -variables[f"px{i + 1},{e}"],
                        variables[f"px{i + 1},{e + 1}"]])
            constraint_count += 1
        for f in range(height - 1):
            cnf.append([-variables[f"a{i + 1}"], -variables[f"py{i + 1},{f}"],
                        variables[f"py{i + 1},{f + 1}"]])
            constraint_count += 1
            
    # Add the 3-literal non-overlapping constraints
    # Formula (4).
    def non_overlapping(rotated, i, j, h1, h2, v1, v2, constraint_count):
        if not rotated:
            i_width = rectangles[i][0]
            i_height = rectangles[i][1]
            j_width = rectangles[j][0]
            j_height = rectangles[j][1]
            i_rotation = variables[f"r{i + 1}"]
            j_rotation = variables[f"r{j + 1}"]
        else:
            i_width = rectangles[i][1]
            i_height = rectangles[i][0]
            j_width = rectangles[j][1]
            j_height = rectangles[j][0]
            i_rotation = -variables[f"r{i + 1}"]
            j_rotation = -variables[f"r{j + 1}"]

        # Square symmertry breaking, if i is square than it cannot be rotated
        if i_width == i_height and rotated:
            i_square = True
            cnf.append([-variables[f"r{i + 1}"]])
            constraint_count += 1
        else:
            i_square = False

        if j_width == j_height and rotated:
            j_square = True
            cnf.append([-variables[f"r{j + 1}"]])
            constraint_count += 1
        else:
            j_square = False

        extra_cnf = [-variables[f"a{i + 1}"], -variables[f"a{j + 1}"]]
        # lri,j v lrj,i v udi,j v udj,i
        four_literal = []
        if h1: 
            four_literal.append(variables[f"lr{i + 1},{j + 1}"])
            constraint_count += 1
        if h2: 
            four_literal.append(variables[f"lr{j + 1},{i + 1}"])
            constraint_count += 1
        if v1: 
            four_literal.append(variables[f"ud{i + 1},{j + 1}"])
            constraint_count += 1
        if v2: 
            four_literal.append(variables[f"ud{j + 1},{i + 1}"])
            constraint_count += 1

        cnf.append(extra_cnf + four_literal + [i_rotation])
        constraint_count += 1
        cnf.append(extra_cnf + four_literal + [j_rotation])
        constraint_count += 1

        # ¬lri, j ∨ ¬pxj, e
        if h1 and not i_square:
            for e in range(min(width, i_width)):
                    cnf.append(extra_cnf + [i_rotation,
                                -variables[f"lr{i + 1},{j + 1}"],
                                -variables[f"px{j + 1},{e}"]])
                    constraint_count += 1
        # ¬lrj,i ∨ ¬pxi,e
        if h2 and not j_square:
            for e in range(min(width, j_width)):
                    cnf.append(extra_cnf + [j_rotation,
                                -variables[f"lr{j + 1},{i + 1}"],
                                -variables[f"px{i + 1},{e}"]])
                    constraint_count += 1
        # ¬udi,j ∨ ¬pyj,f
        if v1 and not i_square:
            for f in range(min(height, i_height)):
                    cnf.append(extra_cnf + [i_rotation,
                                -variables[f"ud{i + 1},{j + 1}"],
                                -variables[f"py{j + 1},{f}"]])
                    constraint_count += 1
        # ¬udj, i ∨ ¬pyi, f,
        if v2 and not j_square:
            for f in range(min(height, j_height)):
                    cnf.append(extra_cnf + [j_rotation,
                                -variables[f"ud{j + 1},{i + 1}"],
                                -variables[f"py{i + 1},{f}"]])
                    constraint_count += 1

        for e in positive_range(width - i_width):
            # ¬lri,j ∨ ¬pxj,e+wi ∨ pxi,e
            if h1 and not i_square:
                    cnf.append(extra_cnf + [i_rotation,
                                -variables[f"lr{i + 1},{j + 1}"],
                                variables[f"px{i + 1},{e}"],
                                -variables[f"px{j + 1},{e + i_width}"]])
                    constraint_count += 1

        for e in positive_range(width - j_width):
            # ¬lrj,i ∨ ¬pxi,e+wj ∨ pxj,e
            if h2 and not j_square:
                    cnf.append(extra_cnf + [j_rotation,
                                -variables[f"lr{j + 1},{i + 1}"],
                                variables[f"px{j + 1},{e}"],
                                -variables[f"px{i + 1},{e + j_width}"]])
                    constraint_count += 1

        for f in positive_range(height - i_height):
            # udi,j ∨ ¬pyj,f+hi ∨ pxi,e
            if v1 and not i_square:
                    cnf.append(extra_cnf + [i_rotation,
                                -variables[f"ud{i + 1},{j + 1}"],
                                variables[f"py{i + 1},{f}"],
                                -variables[f"py{j + 1},{f + i_height}"]])
                    constraint_count += 1
        for f in positive_range(height - j_height):
            # ¬udj,i ∨ ¬pyi,f+hj ∨ pxj,f
            if v2 and not j_square:
                cnf.append(extra_cnf + [j_rotation,
                            -variables[f"ud{j + 1},{i + 1}"],
                            variables[f"py{j + 1},{f}"],
                            -variables[f"py{i + 1},{f + j_height}"]])
                constraint_count += 1
    for i in range(n):
        for j in range(i + 1, n):
            # lri,j ∨ lrj,i ∨ udi,j ∨ udj,i
            #Large-rectangles horizontal
            # cnf.append([-variables[f"a{i + 1}"], -variables[f"a{j + 1}"]])
            if min(rectangles[i][0], rectangles[i][1]) + min(rectangles[j][0], rectangles[j][1]) > width:
                non_overlapping(False, i, j, False, False, True, True, constraint_count)
                non_overlapping(True, i, j, False, False, True, True , constraint_count)
            # Large rectangles vertical
            elif min(rectangles[i][0], rectangles[i][1]) + min(rectangles[j][0], rectangles[j][1]) > height:
                non_overlapping(False, i, j, True, True, False, False, constraint_count)
                non_overlapping(True, i, j, True, True, False, False, constraint_count)

            # Same rectangle and is a square
            elif rectangles[i] == rectangles[j]:
                if rectangles[i][0] == rectangles[i][1]:
                    cnf.append([-variables[f"r{i + 1}"]])
                    constraint_count += 1
                    cnf.append([-variables[f"r{j + 1}"]])
                    constraint_count += 1
                    non_overlapping(False,i ,j, True, True, True, True, constraint_count)
                else:
                    non_overlapping(False, i, j, True, True, True, True, constraint_count)
                    non_overlapping(True, i, j, True, True, True, True, constraint_count)
            # normal rectangles
            else:
                non_overlapping(False, i, j, True, True, True, True, constraint_count)
                non_overlapping(True, i, j, True, True, True, True, constraint_count)


   # Domain encoding to ensure every rectangle stays inside strip's boundary
    for i in range(n):
        # cnf.append([-variables[f"a{i + 1}"]])
        if rectangles[i][0] > width: #if rectangle[i]'s width larger than strip's width, it has to be rotated
            cnf.append([variables[f"r{i + 1}"], -variables[f"a{i + 1}"]])
            constraint_count += 1
        else:
            for e in range(width - rectangles[i][0], width):
                    cnf.append([variables[f"r{i + 1}"],
                                variables[f"px{i + 1},{e}"], -variables[f"a{i + 1}"]])
                    constraint_count += 1
        if rectangles[i][1] > height:
            cnf.append([variables[f"r{i + 1}"], -variables[f"a{i + 1}"]])
            constraint_count += 1
        else:
            for f in range(height - rectangles[i][1], height):
                    cnf.append([variables[f"r{i + 1}"],
                                variables[f"py{i + 1},{f}"], -variables[f"a{i + 1}"]])
                    constraint_count += 1

        # Rotated
        if rectangles[i][1] > width:
            cnf.append([-variables[f"r{i + 1}"], -variables[f"a{i + 1}"]])
            constraint_count += 1
        else:
            for e in range(width - rectangles[i][1], width):
                    cnf.append([-variables[f"r{i + 1}"],
                                variables[f"px{i + 1},{e}"], -variables[f"a{i + 1}"]])
                    constraint_count += 1
        if rectangles[i][0] > height:
            cnf.append([-variables[f"r{i + 1}"], -variables[f"a{i + 1}"]])
            constraint_count += 1
        else:
            for f in range(height - rectangles[i][0], height):
                cnf.append([-variables[f"r{i + 1}"],
                            variables[f"py{i + 1},{f}"], -variables[f"a{i + 1}"]])
                constraint_count += 1

    # add all clauses to SAT solver
    solver = Glucose3(use_timer=True)
    solver.append_formula(cnf)

    # Uncomment to set time litmit of solver.
    timer = Timer(600, interrupt, [solver])
    timer.start()

    sat_status = solver.solve_limited(expect_interrupt = True)

    print(sat_status)
    num_vars = solver.nof_vars()
    num_clauses = solver.nof_clauses()
    
    if sat_status is False:
        solver.delete()
        return ["UNSAT"]
    else:
        # Initial result position of rectangles by [-1, -1]
        pos = [[-1 for i in range(2)] for j in range(n)]
        result = {}
        rotation = []
        model = solver.get_model()
        if model is None:
            return ["UNSAT"]
        else:
            for index in range(0, ended_loop):
                var = model[index]
                var_key = next((k for k, v in variables.items() if v == abs(var)), None)
                if var_key is None:
                    print(f"Variable {var} not found in variables")
                    continue
                result[var_key] = var > 0

            # From SAT result, decode into rectangles' position.
            selected_rectangles = []
            result_max_profit = 0
            result_rectangle = []
            res_pos = []

            for i in range(n):
                rotation.append(result[f"r{i + 1}"])
                selected_rectangles.append(result[f"a{i + 1}"])

                # # Append position to pos[]
                for e in range(width - 1):
                    if result[f"px{i + 1},{e}"] == False and result[f"px{i + 1},{e + 1}"] == True:
                        pos[i][0] = e + 1
                    if e == 0 and result[f"px{i + 1},{e}"] == True:
                        pos[i][0] = 0
                for f in range(height - 1):
                    if result[f"py{i + 1},{f}"] == False and result[f"py{i + 1},{f + 1}"] == True:
                        pos[i][1] = f + 1
                    if f == 0 and result[f"py{i + 1},{f}"] == True:
                        pos[i][1] = 0
                
                if selected_rectangles[i] == True:
                    # calculate max profit in list result rectangles.
                    result_max_profit += rectangles[i][2]
                    # value dimension of rectangle / rotate rectangle.
                    orig_width, orig_height = rectangles[i][0], rectangles[i][1]
                    if rotation[i]:
                        orig_width = rectangles[i][1]
                        orig_height = rectangles[i][0]
                    result_rectangle.append([orig_width, orig_height])
                    # Verify position of rectangle. Must be check if it is similar with value of init => change to 0.
                    if pos[i][0] > -1 and pos[i][1] == -1:
                        pos[i][1] = 0
                    if pos[i][0] == -1 and pos[i][1] > -1:
                        pos[i][0] = 0
                    res_pos.append([pos[i][0], pos[i][1]])

            display_solution(strip, result_rectangle, res_pos)
            return ["SAT", result_max_profit, counter-1, constraint_count, num_vars, num_clauses]

def max_profit_solution():
    folder_path = '../../dataset/soft/'
    for file_name in os.listdir(folder_path):
        file_path = os.path.join(folder_path, file_name)
        if os.path.isfile(file_path) and file_name.endswith('.txt') and file_name == "cgcut1.txt":
            print(f"Processing file: {file_name}")
            with open(file_path, 'r') as file:
                lines = file.readlines()
                list_strip = list(map(int,lines[0].strip().split()))
                widths = list(map(int, lines[1].strip().split()))
                heights = list(map(int, lines[2].strip().split()))
                profits = list(map(int, lines[3].strip().split()))
                num_items = len(widths)

                # Get list rectangles from input data.
                rectangles = []
                prev_sat = []
                lower_bound = min(profits)
                upper_bound = sum(profits)
                for j in range (0, num_items):
                    item = (widths[j], heights[j], profits[j])
                    rectangles.append(item)

                started_time = time.time()
                isExport = ""
                while (lower_bound + 1 < upper_bound):
                    if (time.time() - started_time >= 600):
                        isExport = "TIMEOUT"
                        export_csv(file_name, "glucose_rotate_symmetry", time.time() - started_time, "TIMEOUT", 0, 0, 0, 0, 0)
                        break
                    mid = lower_bound + (upper_bound - lower_bound) // 2
                    print(mid)
                    status = glucose_constraints(rectangles, list_strip[0], list_strip[1], mid, profits)
                    if (status[0] == "UNSAT"):
                        upper_bound = mid
                        continue
                    elif (status[0] == "SAT"):
                        prev_sat = status
                        lower_bound = mid
                        continue
                ended_time = time.time()

                if (isExport != "TIMEOUT"):
                    if (prev_sat[0]) == "SAT":
                        export_csv(file_name, "glucose_rotate_symmetry", ended_time - started_time, "SAT", prev_sat[2], prev_sat[3], prev_sat[4], prev_sat[5], prev_sat[1])
                    else: export_csv(file_name, "glucose_rotate_symmetry", ended_time - started_time, "UNSAT", 0, 0, 0, 0, 0 )


max_profit_solution()
