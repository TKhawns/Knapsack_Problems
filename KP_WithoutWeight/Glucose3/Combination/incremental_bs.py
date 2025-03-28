from pysat.formula import CNF
from pysat.solvers import Glucose3
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from zipfile import BadZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import matplotlib.pyplot as plt
import time

id_counter = 0

def display_solution(strip, rectangles, pos_circuits):
    ax = plt.subplots()
    ax = plt.gca()
    plt.title(strip)

    if len(pos_circuits) > 0:
        for i in range(len(rectangles)):
            # Add fill color for better visibility
            rect = plt.Rectangle(pos_circuits[i], *rectangles[i], edgecolor="#333", facecolor="#69b3a2", alpha=0.5)
            ax.add_patch(rect)
    else:
        print("No circuits to display.")

    ax.set_xlim(0, strip[0])
    ax.set_ylim(0, strip[1] + 1)
    ax.set_xticks(range(strip[0] + 1))
    ax.set_yticks(range(strip[1] + 1))
    ax.set_xlabel('width')
    ax.set_ylabel('height')

    plt.show()

def write_to_xlsx(result_dict):
    # Append the result to a list
    excel_results = []
    excel_results.append(result_dict)

    output_path =  'output/'

    # Write the results to an Excel file
    if not os.path.exists(output_path): os.makedirs(output_path)

    df = pd.DataFrame(excel_results)
    current_date = datetime.now().strftime('%Y-%m-%d')
    excel_file_path = f"{output_path}/results_{current_date}.xlsx"

    # Check if the file already exists
    if os.path.exists(excel_file_path):
        try:
            book = load_workbook(excel_file_path)
        except BadZipFile:
            book = Workbook()  # Create a new workbook if the file is not a valid Excel file

        # Check if the 'Results' sheet exists
        if 'Results' not in book.sheetnames:
            book.create_sheet('Results')  # Create 'Results' sheet if it doesn't exist

        sheet = book['Results']
        for row in dataframe_to_rows(df, index=False, header=False): sheet.append(row)
        book.save(excel_file_path)

    else: df.to_excel(excel_file_path, index=False, sheet_name='Results', header=False)

def solve_by_pysat(input, time, result, num_var, num_clause, name_lib):
    global id_counter
    id_counter += 1
    result_dict = {
        "ID": id_counter,
        "Problem": input,
        "Type": name_lib,
        "Time": time,
        "Result": result,
        "Variables": num_var,
        "Clauses": num_clause
    }
    write_to_xlsx(result_dict)

def positive_range(end):
    if (end < 0):
        return []
    return range(end)

def pos_i(i, k, profit):
    if i == 0:
        return 0
    if i < k:
        sum_w = sum(profit[1:i+1])
        return min(k, sum_w)
    else:
        return k

def opp_constraint(strip, rectangles):
    # Define the variables
    cnf = CNF()

    width = strip[0]
    height = strip[1]
    variables = {}
    counter = 1
    n = len(rectangles)

    # a_i = True (1) if item i is selected.
    for i in range(n):
        variables[f"a{i + 1}"] = counter
        counter += 1

    # SAT Encoding of 2 0PP.
    # create lr, ud, px, py variables from 1 to N.
    for i in range(n):
        for j in range(n):
            if i != j:
                variables[f"lr{i + 1},{j + 1}"] = counter  # lri,rj
                counter += 1
                variables[f"ud{i + 1},{j + 1}"] = counter  # uri,rj
                counter += 1
        for e in range(width):
            variables[f"px{i + 1},{e}"] = counter  # pxi,e
            counter += 1
        for f in range(height):
            variables[f"py{i + 1},{f}"] = counter  # pyi,f
            counter += 1

    # Rotated variables of 2OPP.
    for i in range(n):
        variables[f"r{i + 1}"] = counter
        counter += 1


    # Add the 2-literal axiom clauses (order constraint)
    # Formula (3).
    # [Update] if a_i = 1 -> order-encoding.
    for i in range(n):
        for e in range(width - 1):  # -1 because we're using e+1 in the clause
            cnf.append([-variables[f"a{i + 1}"], -variables[f"px{i + 1},{e}"],
                        variables[f"px{i + 1},{e + 1}"]])
        for f in range(height - 1):
            cnf.append([-variables[f"a{i + 1}"], -variables[f"py{i + 1},{f}"],
                        variables[f"py{i + 1},{f + 1}"]])
            
    # Add the 3-literal non-overlapping constraints
    # Formula (4).
    def non_overlapping(rotated, i, j, h1, h2, v1, v2):
        if not rotated:
            i_width = rectangles[i][0]
            i_height = rectangles[i][1]
            j_width = rectangles[j][0]
            j_height = rectangles[j][1]
            i_rotation = variables[f"r{i + 1}"]
            j_rotation = variables[f"r{j + 1}"]
        else:
            i_width = rectangles[i][1]
            i_height = rectangles[i][0]
            j_width = rectangles[j][1]
            j_height = rectangles[j][0]
            i_rotation = -variables[f"r{i + 1}"]
            j_rotation = -variables[f"r{j + 1}"]

        # Square symmertry breaking, if i is square than it cannot be rotated
        if i_width == i_height and rotated:
            i_square = True
            cnf.append([-variables[f"r{i + 1}"]])
        else:
            i_square = False

        if j_width == j_height and rotated:
            j_square = True
            cnf.append([-variables[f"r{j + 1}"]])
        else:
            j_square = False

        extra_cnf = [-variables[f"a{i + 1}"], -variables[f"a{j + 1}"]]
        # lri,j v lrj,i v udi,j v udj,i
        four_literal = []
        if h1: four_literal.append(variables[f"lr{i + 1},{j + 1}"])
        if h2: four_literal.append(variables[f"lr{j + 1},{i + 1}"])
        if v1: four_literal.append(variables[f"ud{i + 1},{j + 1}"])
        if v2: four_literal.append(variables[f"ud{j + 1},{i + 1}"])

        cnf.append(extra_cnf + four_literal + [i_rotation])
        cnf.append(extra_cnf + four_literal + [j_rotation])

        # ¬lri, j ∨ ¬pxj, e
        if h1 and not i_square:
            for e in range(min(width, i_width)):
                    cnf.append(extra_cnf + [i_rotation,
                                -variables[f"lr{i + 1},{j + 1}"],
                                -variables[f"px{j + 1},{e}"]])
        # ¬lrj,i ∨ ¬pxi,e
        if h2 and not j_square:
            for e in range(min(width, j_width)):
                    cnf.append(extra_cnf + [j_rotation,
                                -variables[f"lr{j + 1},{i + 1}"],
                                -variables[f"px{i + 1},{e}"]])
        # ¬udi,j ∨ ¬pyj,f
        if v1 and not i_square:
            for f in range(min(height, i_height)):
                    cnf.append(extra_cnf + [i_rotation,
                                -variables[f"ud{i + 1},{j + 1}"],
                                -variables[f"py{j + 1},{f}"]])
        # ¬udj, i ∨ ¬pyi, f,
        if v2 and not j_square:
            for f in range(min(height, j_height)):
                    cnf.append(extra_cnf + [j_rotation,
                                -variables[f"ud{j + 1},{i + 1}"],
                                -variables[f"py{i + 1},{f}"]])

        for e in positive_range(width - i_width):
            # ¬lri,j ∨ ¬pxj,e+wi ∨ pxi,e
            if h1 and not i_square:
                    cnf.append(extra_cnf + [i_rotation,
                                -variables[f"lr{i + 1},{j + 1}"],
                                variables[f"px{i + 1},{e}"],
                                -variables[f"px{j + 1},{e + i_width}"]])

        for e in positive_range(width - j_width):
            # ¬lrj,i ∨ ¬pxi,e+wj ∨ pxj,e
            if h2 and not j_square:
                    cnf.append(extra_cnf + [j_rotation,
                                -variables[f"lr{j + 1},{i + 1}"],
                                variables[f"px{j + 1},{e}"],
                                -variables[f"px{i + 1},{e + j_width}"]])

        for f in positive_range(height - i_height):
            # udi,j ∨ ¬pyj,f+hi ∨ pxi,e
            if v1 and not i_square:
                    cnf.append(extra_cnf + [i_rotation,
                                -variables[f"ud{i + 1},{j + 1}"],
                                variables[f"py{i + 1},{f}"],
                                -variables[f"py{j + 1},{f + i_height}"]])
        for f in positive_range(height - j_height):
            # ¬udj,i ∨ ¬pyi,f+hj ∨ pxj,f
            if v2 and not j_square:
                cnf.append(extra_cnf + [j_rotation,
                            -variables[f"ud{j + 1},{i + 1}"],
                            variables[f"py{j + 1},{f}"],
                            -variables[f"py{i + 1},{f + j_height}"]])

    for i in range(n):
        for j in range(i + 1, n):
            # lri,j ∨ lrj,i ∨ udi,j ∨ udj,i
            #Large-rectangles horizontal
            # cnf.append([-variables[f"a{i + 1}"], -variables[f"a{j + 1}"]])
            if min(rectangles[i][0], rectangles[i][1]) + min(rectangles[j][0], rectangles[j][1]) > width:
                non_overlapping(False, i, j, False, False, True, True)
                non_overlapping(True, i, j, False, False, True, True)
            # Large rectangles vertical
            elif min(rectangles[i][0], rectangles[i][1]) + min(rectangles[j][0], rectangles[j][1]) > height:
                non_overlapping(False, i, j, True, True, False, False)
                non_overlapping(True, i, j, True, True, False, False)

            # Same rectangle and is a square
            elif rectangles[i] == rectangles[j]:
                if rectangles[i][0] == rectangles[i][1]:
                    cnf.append([-variables[f"r{i + 1}"]])
                    cnf.append([-variables[f"r{j + 1}"]])
                    non_overlapping(False,i ,j, True, True, True, True)
                else:
                    non_overlapping(False, i, j, True, True, True, True)
                    non_overlapping(True, i, j, True, True, True, True)
            # normal rectangles
            else:
                non_overlapping(False, i, j, True, True, True, True)
                non_overlapping(True, i, j, True, True, True, True)


    # Domain encoding to ensure every rectangle stays inside strip's boundary
    for i in range(n):
        # cnf.append([-variables[f"a{i + 1}"]])
        if rectangles[i][0] > width: #if rectangle[i]'s width larger than strip's width, it has to be rotated
            cnf.append([variables[f"r{i + 1}"], -variables[f"a{i + 1}"]])
        else:
            for e in range(width - rectangles[i][0], width):
                    cnf.append([variables[f"r{i + 1}"],
                                variables[f"px{i + 1},{e}"], -variables[f"a{i + 1}"]])
        if rectangles[i][1] > height:
            cnf.append([variables[f"r{i + 1}"], -variables[f"a{i + 1}"]])
        else:
            for f in range(height - rectangles[i][1], height):
                    cnf.append([variables[f"r{i + 1}"],
                                variables[f"py{i + 1},{f}"], -variables[f"a{i + 1}"]])

        # Rotated
        if rectangles[i][1] > width:
            cnf.append([-variables[f"r{i + 1}"], -variables[f"a{i + 1}"]])
        else:
            for e in range(width - rectangles[i][1], width):
                    cnf.append([-variables[f"r{i + 1}"],
                                variables[f"px{i + 1},{e}"], -variables[f"a{i + 1}"]])
        if rectangles[i][0] > height:
            cnf.append([-variables[f"r{i + 1}"], -variables[f"a{i + 1}"]])
        else:
            for f in range(height - rectangles[i][0], height):
                cnf.append([-variables[f"r{i + 1}"],
                            variables[f"py{i + 1},{f}"], -variables[f"a{i + 1}"]])
      
    return [cnf, variables, counter]

def opp_solution(rectangles, strip, k, profit, opp_constraint):
    cnf = opp_constraint[0]
    variables = opp_constraint[1]
    counter = opp_constraint[2]

    # Define the variables
    width = strip[0]
    height = strip[1]
    n_scpb = len(rectangles)
    # SCPB Constraints definition
    profit = [0] + profit

    # Create map_register to hold the auxiliary variables
    map_register = [[0 for _ in range(k + 1)] for _ in range(n_scpb + 1)]

    # Init value of map_register in SCPB.
    for i in range(1, n_scpb):
        n_bits = pos_i(i, k, profit)
        for j in range(1, n_bits + 1):
            map_register[i][j] = counter
            counter += 1

    # (1) X_i -> R_i,j for j = 1 to w_i
    for i in range(1, n_scpb):
        for j in range(1, profit[i] + 1):
            if j <= pos_i(i, k, profit):
                cnf.append([-variables[f"a{i}"], map_register[i][j]])

    # (2) R_{i-1,j} -> R_i,j for j = 1 to pos_{i-1}
    for i in range(2, n_scpb):
        for j in range(1, pos_i(i - 1, k, profit) + 1):
            cnf.append([-map_register[i - 1][j], map_register[i][j]])

    # (3) X_i ^ R_{i-1,j} -> R_i,j+w_i for j = 1 to pos_{i-1}
    for i in range(2, n_scpb):
        for j in range(1, pos_i(i - 1, k, profit) + 1):
            if j + profit[i] <= k and j + profit[i] <= pos_i(i, k, profit):
                cnf.append([-variables[f"a{i}"], -map_register[i - 1][j], map_register[i][j + profit[i]]])

    # (4) ¬X_i ^ ¬R_{i-1,j} -> ¬R_i,j for j = 1 to pos_{i-1}
    for i in range(2, n_scpb):
        for j in range(1, pos_i(i - 1, k, profit) + 1):
            cnf.append([variables[f"a{i}"], map_register[i - 1][j], -map_register[i][j]])

    # (5) ¬X_i -> ¬R_i,j for j = 1 + pos_{i-1} to pos_i
    for i in range(1, n_scpb):
        # if pos_i(i - 1, k, weight) < k:
            for j in range(1 + pos_i(i - 1, k, profit), pos_i(i, k, profit) + 1):
                cnf.append([variables[f"a{i}"], -map_register[i][j]])

    # (6) ¬R_{i-1,j} -> ¬R_i,j+w_i for j = 1 to pos_{i-1}
    for i in range(2, n_scpb):
        # if pos_i(i - 1, k, weight) < k:
            for j in range(1, pos_i(i - 1, k, profit) + 1):
                if j + profit[i] <= k and j + profit[i] <= pos_i(i, k, profit):
                    cnf.append([map_register[i - 1][j], -map_register[i][j + profit[i]]])

    # (7) R_{n-1,k} v (X_n ^ R_{n-1,k-w_n})
    if k > pos_i(n_scpb - 1, k, profit):
        if k - profit[n_scpb] > 0 and k - profit[n_scpb] <= pos_i(n_scpb - 1, k, profit):
            cnf.append([variables[f"a{n_scpb}"]])
            cnf.append([map_register[n_scpb - 1][k - profit[n_scpb]]])
    else:
        cnf.append([map_register[n_scpb - 1][k], variables[f"a{n_scpb}"]])
        if k - profit[n_scpb] > 0 and k - profit[n_scpb] <= pos_i(n_scpb - 1, k, profit):
            cnf.append([map_register[n_scpb - 1][k], map_register[n_scpb - 1][k - profit[n_scpb]]])

    # add all clauses to SAT solver
    solver = Glucose3()
    solver.append_formula(cnf)

    # Uncomment to set time litmit of solver.
    # timer = Timer(60, interrupt, [solver])
    # timer.start()

    sat_status = solver.solve()
    num_vars = solver.nof_vars()
    num_clauses = solver.nof_clauses()

    print(num_vars, num_clauses, len(rectangles))
    
    if sat_status is False:
        return ["UNSAT", num_vars, num_clauses, len(rectangles)]
    else:
        # Initial result position of rectangles by [-1, -1]
        pos = [[-1 for i in range(2)] for j in range(n_scpb)]
        result = {}
        rotation = []

        model = solver.get_model()
        if model is None:
            solver.delete()
            return ["TIMEOUT", num_vars, num_clauses]
        else:
            for index in range(0, opp_constraint[2]):
                var = model[index]
                var_key = next((k for k, v in variables.items() if v == abs(var)), None)
                if var_key is None:
                    print(f"Variable {var} not found in variables")
                    continue
                result[var_key] = var > 0

            # From SAT result, decode into rectangles' position.
            selected_rectangles = []
            result_rectangle = []
            res_pos = []
            for i in range(n_scpb):
                rotation.append(result[f"r{i + 1}"])
                selected_rectangles.append(result[f"a{i + 1}"])

                # Append position to pos[]
                for e in range(width - 1):
                    if result[f"px{i + 1},{e}"] == False and result[f"px{i + 1},{e + 1}"] == True:
                        pos[i][0] = e + 1
                    if e == 0 and result[f"px{i + 1},{e}"] == True:
                        pos[i][0] = 0
                for f in range(height - 1):
                    if result[f"py{i + 1},{f}"] == False and result[f"py{i + 1},{f + 1}"] == True:
                        pos[i][1] = f + 1
                    if f == 0 and result[f"py{i + 1},{f}"] == True:
                        pos[i][1] = 0
                
                result_max_profit = 0

                if selected_rectangles[i] == True:
                    # calculate max profit in list result rectangles.
                    result_max_profit += rectangles[i][2]
                    # value dimension of rectangle / rotate rectangle.
                    orig_width, orig_height = rectangles[i][0], rectangles[i][1]
                    if rotation[i]:
                        orig_width = rectangles[i][1]
                        orig_height = rectangles[i][0]
                    result_rectangle.append([orig_width, orig_height])
                    # Original dimension of rectangle before rotate.
                    actual_width, actual_height = rectangles[i][0], rectangles[i][1]
                    # Verify position of rectangle. Must be check if it is similar with value of init => change to 0.
                    if pos[i][0] > -1 and pos[i][1] == -1:
                        pos[i][1] = 0
                    if pos[i][0] == -1 and pos[i][1] > -1:
                        pos[i][0] = 0
                    res_pos.append([pos[i][0], pos[i][1]])

                    # Log the result.
                    # print(f"  Dimensions: {actual_width}x{actual_height} {'(Rotated)' if rotation[i] else ''}")

            # print(f"Total profit: {total_profit}")
            # display_solution(strip, result_rectangle, res_pos)
            # return(["sat", pos, rotation])
            solver.delete()
            return ["SAT", result_max_profit, num_vars, num_clauses]

def max_profit_solution():
    with open('../dataset/dataset1.txt', 'r') as file:
        lines = file.readlines()

        list_strip = list(map(int,lines[0].strip().split()))
        widths = list(map(int, lines[1].strip().split()))
        heights = list(map(int, lines[2].strip().split()))
        profits = list(map(int, lines[3].strip().split()))
        strip = (list_strip[0], list_strip[1])
        num_items = len(widths)

        rectangles = []
        for j in range (0, num_items):
            item = (widths[j], heights[j], profits[j])
            rectangles.append(item)

        # Binary search from min of list profit to total all profits.
        lower_bound = min(profits)
        upper_bound = sum(profits)

        print("Starting solve from binary search:")
        status = []

        started_time = time.time()
        fixed_constraint = opp_constraint(strip, rectangles)
        while (lower_bound + 1 < upper_bound):
            mid = lower_bound + (upper_bound - lower_bound) // 2
            print("Solving this problem: ", mid, " between ", lower_bound, " and ", upper_bound)
            status = opp_solution(rectangles, strip, mid , profits, fixed_constraint)
            print(status)
            if (status[0] == "UNSAT"):
                upper_bound = mid
                continue
            elif (status[0] == "SAT"):
                lower_bound = mid
                continue

        # print("Mid value in loop: ", mid)
        # solve_by_pysat(num_items, time.time() - started_time, status[0], status[2], status[3], "Glucose3")

        if (status[0] == "SAT"):
            print("Max profit is: ", status[1], " Number of variables: ", status[2], " Number of clauses: ", status[3])
        else:
            print("UNSAT this instance")
    return

max_profit_solution()
